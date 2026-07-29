# -*- coding: utf-8 -*-
"""
ANDA 펀드 대시보드
------------------
data 폴더에 아래 두 파일을 펀드별로 넣어두면, 펀드마다 페이지(탭)가 자동으로 생깁니다.

    data/<펀드이름>펀드수익률.xlsx   (No, 일자, 초과수익률, 펀드수익률, BM수익률)
    data/<펀드이름>펀드종목.xlsx     (종목명, 편입비, 평가액, 평가손익률, 업종명, 규모구분 ...)

예) data/ckvf펀드수익률.xlsx + data/ckvf펀드종목.xlsx  ->  'ckvf' 탭이 생성

매일 새 엑셀로 같은 이름의 파일을 덮어쓰고 브라우저를 새로고침하면 자동으로 최신화됩니다.

실행:  python app.py   ->  브라우저에서 http://127.0.0.1:5000 접속
"""

import os
import glob
import json
import socket
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from flask import Flask, redirect, url_for, abort, request, Response

# ---------------------------------------------------------------------------
# 경로 설정
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")


def _load_dotenv():
    """앱 시작 시 .env 를 os.environ 에 반영. 이미 설정된 변수는 건드리지 않음."""
    env_path = os.path.join(BASE_DIR, ".env")
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k and k not in os.environ:
                    os.environ[k] = v
    except Exception as e:
        print(f"[.env] 로드 실패: {e}")


_load_dotenv()

RET_SUFFIX = "펀드수익률"   # 수익률 파일 접미사
HOLD_SUFFIX = "펀드종목"    # 종목 파일 접미사
RISK_FILE = "펀드리스크.xlsx"  # 5개 펀드 리스크지표가 한 파일에 든 파일
KOSDAQ_FILE = "코스닥70.xlsx"  # 코스닥 70/150 후보 분석 파일 ('클로드' 시트)
KOSDAQ_SHEET = "클로드"
REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "reports")
SEMI_FILE = "반도체_소부_총정리.xlsx"  # 종목 마스터 추출용 (시트별 C3=종목명, C4=종가, C5=시총)
STOCK_CODES_CACHE = "_stock_codes.json"  # 종목명→코드 매핑 캐시 (data 폴더 하위)
KOSPI_SECTOR_FILE = "코스피섹터.xlsx"  # KOSPI 업종별 시총 비중 (헤더: 업종명, 비중)

# KOSPI 22개 업종 시총 비중 스냅샷 (KRX 정보데이터시스템 기준 추정치).
# data/코스피섹터.xlsx 가 있으면 그것이 우선이며, 없을 때만 이 값이 쓰인다.
KOSPI_SECTOR_DEFAULT = {
    "전기전자": 35.0, "운수장비": 9.0, "화학": 7.0, "서비스업": 7.0,
    "금융업": 5.5, "의약품": 5.0, "보험": 4.0, "운수창고": 3.0,
    "건설업": 3.0, "철강금속": 3.0, "기계": 2.5, "유통업": 2.5,
    "통신업": 2.5, "의료정밀": 2.5, "음식료품": 2.0, "전기가스업": 1.5,
    "증권": 1.5, "은행": 1.5, "비금속광물": 1.0, "섬유의복": 0.5,
    "종이목재": 0.5, "기타": 1.0,
}

# ===========================================================================
#  공유 설정  ── 아래 3줄만 필요에 맞게 바꾸세요
# ===========================================================================
LOGIN_ID = "anda"                 # 접속 아이디
LOGIN_PW = "12345"       # ★ 접속 비밀번호 — 이 값을 바꿔서 동료에게만 알려주세요
SHARE_ON_NETWORK = True           # True=사내 다른 PC도 접속 / False=내 PC만
# ===========================================================================

app = Flask(__name__)


@app.before_request
def _require_login():
    """모든 페이지 접속 시 아이디/비밀번호 확인 (브라우저 로그인 창)."""
    a = request.authorization
    if not a or a.username != LOGIN_ID or a.password != LOGIN_PW:
        return Response("로그인이 필요합니다.", 401,
                        {"WWW-Authenticate": 'Basic realm="ANDA Dashboard"'})


@app.before_request
def _read_sector_mode():
    """쿠키에서 섹터 분류 모드 읽어 g 에 설정 (detail=세분화 / big=11대분류)."""
    from flask import g
    g.sector_mode = request.cookies.get("sector_mode", "detail")


def _local_ip():
    """동료에게 알려줄 내 PC의 사내 IP 주소를 찾는다."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


# ---------------------------------------------------------------------------
# 1) data 폴더에서 펀드 목록 찾기
# ---------------------------------------------------------------------------
def _nfc(s):
    """한글 자모 분리(NFD)/결합(NFC) 차이를 통일한다 (맥/윈도우 호환)."""
    return unicodedata.normalize("NFC", str(s))


def find_funds():
    """data 폴더를 훑어 펀드 이름 목록을 정렬해서 반환한다.

    파일명 패턴:
      <펀드이름>펀드수익률.xlsx          — 기본 (날짜 없는 파일)
      <펀드이름>펀드수익률<날짜>.xlsx    — 날짜 접미사 (예: '260710')
    같은 펀드에 여러 버전이 있으면 **날짜 접미사가 가장 큰(최신)** 파일 선택.
    수익률 파일과 종목 파일이 둘 다 있어야 반영.
    """
    import re as _re
    funds_candidates = {}  # {펀드이름: {"returns": {suffix: disk_name, ...}, "holdings": {...}}}
    if not os.path.isdir(DATA_DIR):
        return {}

    norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}

    # 정규식으로 (fund_name, suffix, kind) 파싱
    ret_pat = _re.compile(r"^(.+?)" + _re.escape(RET_SUFFIX) + r"(\d{0,8})\.xlsx$")
    hold_pat = _re.compile(r"^(.+?)" + _re.escape(HOLD_SUFFIX) + r"(\d{0,8})\.xlsx$")

    for nfc_name in norm_map:
        for pat, kind in ((ret_pat, "returns"), (hold_pat, "holdings")):
            m = pat.match(nfc_name)
            if not m:
                continue
            fund, suffix = m.group(1), m.group(2)
            if not fund:
                continue
            funds_candidates.setdefault(fund, {"returns": {}, "holdings": {}})
            funds_candidates[fund][kind][suffix] = norm_map[nfc_name]
            break

    funds = {}
    for fund, kinds in funds_candidates.items():
        rets, holds = kinds["returns"], kinds["holdings"]
        if not rets or not holds:
            continue
        # 각각 최신 suffix 선택 (숫자로 비교, 빈 suffix 는 -1 로 취급)
        def _pick_latest(d):
            latest_key = max(d.keys(), key=lambda s: int(s) if s else -1)
            return d[latest_key]
        funds[fund] = {
            "returns":  os.path.join(DATA_DIR, _pick_latest(rets)),
            "holdings": os.path.join(DATA_DIR, _pick_latest(holds)),
        }
    return dict(sorted(funds.items()))


# ---------------------------------------------------------------------------
# 1-2) 리스크 파일 읽기 + 펀드 매칭
# ---------------------------------------------------------------------------
def load_risk():
    """data/펀드리스크.xlsx (또는 펀드리스크<날짜>.xlsx 중 최신) 로부터 리스크지표 로드.

    헤더가 2줄이라 위치(열 번호) 기준으로 읽는다. 파일이 없으면 빈 목록.
    """
    if not os.path.isdir(DATA_DIR):
        return []
    import re as _re
    norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}
    # 펀드리스크<날짜>.xlsx 중 최신 선택
    base = RISK_FILE.replace(".xlsx", "")
    pat = _re.compile(r"^" + _re.escape(base) + r"(\d{0,8})\.xlsx$")
    candidates = {}
    for n in norm_map:
        m = pat.match(n)
        if m:
            candidates[m.group(1)] = norm_map[n]
    if not candidates:
        return []
    latest_key = max(candidates.keys(), key=lambda s: int(s) if s else -1)
    path = os.path.join(DATA_DIR, candidates[latest_key])
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    out = []
    for r in rows[2:]:                       # 앞 2줄은 헤더
        if r is None or len(r) < 23:
            continue
        if r[1] in (None, "") or r[2] in (None, ""):
            continue                          # 코드/명 없으면 건너뜀
        out.append({
            "code": str(r[1]).strip(),
            "name": str(r[2]).strip(),
            "sharpe": r[3], "jensen": r[4], "treynor": r[5],
            "beta_bm": r[6], "r2_bm": r[7], "beta_kospi": r[8], "r2_kospi": r[9],
            "te": r[10], "ir": r[11], "wavg_beta": r[12], "idx_beta": r[13],
            "md": r[14], "sd_fund": r[15], "sd_bm": r[16],
            "var_fund": r[17], "var_bm": r[18], "delta_var": r[19],
            "raroc": r[20],
        })
    return out


# ---------------------------------------------------------------------------
# 1-3) 코스닥 70/150 분석 파일 읽기
# ---------------------------------------------------------------------------
def load_kosdaq70():
    """data/코스닥70.xlsx 의 '클로드' 시트에서 5개 블록을 읽는다.

    같은 시트에 5개 표가 가로로 나란히 배치되어 있다.
    - B~E : 코스닥 70 후보   (Code, Name, 업종, 시가총액)
    - G~K : 코스닥 150       (Code, Name, 업종, 26E 순이익, 시가총액)
    - M~N : 신규편입 후보    (Name, 업종)
    - P~S : 퇴출 후보        (Name, 업종, 26E 순이익, 시총)
    - U~X : 흑자 퇴출 후보   (Name, 업종, 26E 순이익, 시총)

    각 블록은 컬럼 그룹별로 행 길이가 다르다. 모든 컬럼이 비어있는 행을 만나면 종료.
    """
    if not os.path.isdir(DATA_DIR):
        return None
    norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}
    key = _nfc(KOSDAQ_FILE)
    if key not in norm_map:
        return None
    path = os.path.join(DATA_DIR, norm_map[key])

    wb = load_workbook(path, read_only=True, data_only=True)
    ws_name = None
    for s in wb.sheetnames:
        if _nfc(s).strip() == _nfc(KOSDAQ_SHEET):
            ws_name = s
            break
    if not ws_name:
        return None
    ws = wb[ws_name]

    # openpyxl read_only 모드에서는 row 단위 iter 가 가장 빠름
    all_rows = list(ws.iter_rows(values_only=True))

    def slice_block(cols, fields, start_row=4):
        """cols: 1-based column indices. start_row: 0-based row index (header 다음 행부터)."""
        out = []
        for r in range(start_row, len(all_rows)):
            row = all_rows[r]
            vals = [row[c - 1] if (c - 1) < len(row) else None for c in cols]
            if all(v in (None, "") for v in vals):
                break  # 이 블록의 데이터 끝
            item = dict(zip(fields, vals))
            out.append(item)
        return out

    k70 = slice_block([2, 3, 4, 5],          ["code", "name", "sector", "cap"])
    k150 = slice_block([7, 8, 9, 10, 11],    ["code", "name", "sector", "ni", "cap"])
    new = slice_block([13, 14],              ["name", "sector"])
    out = slice_block([16, 17, 18, 19],      ["name", "sector", "ni", "cap"])
    out_pos = slice_block([21, 22, 23, 24],  ["name", "sector", "ni", "cap"])

    # 코드 정리: 'A123456' -> '123456'
    for it in k70:
        c = it.get("code")
        if isinstance(c, str) and c.startswith("A"):
            it["code"] = c[1:]
    for it in k150:
        c = it.get("code")
        if isinstance(c, str) and c.startswith("A"):
            it["code"] = c[1:]

    # 숫자 변환
    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    for it in k70:
        it["cap"] = num(it.get("cap"))
    for it in k150:
        it["cap"] = num(it.get("cap"))
        it["ni"] = num(it.get("ni"))
    for it in out:
        it["cap"] = num(it.get("cap"))
        it["ni"] = num(it.get("ni"))
    for it in out_pos:
        it["cap"] = num(it.get("cap"))
        it["ni"] = num(it.get("ni"))

    # 정렬: 시총 큰 순
    k70.sort(key=lambda x: -(x.get("cap") or 0))
    k150.sort(key=lambda x: -(x.get("cap") or 0))
    out.sort(key=lambda x: -(x.get("cap") or 0))
    out_pos.sort(key=lambda x: -(x.get("cap") or 0))

    # 섹터별 시총 비중 (코스닥 150 기준)
    sec_cap = {}
    for it in k150:
        sec = it.get("sector") or "기타"
        sec_cap[sec] = sec_cap.get(sec, 0.0) + (it.get("cap") or 0.0)
    total = sum(sec_cap.values())
    sectors = []
    # 엑셀 시총/순이익 단위가 '천원' 이므로 억원 환산은 ÷1e5
    for s, v in sorted(sec_cap.items(), key=lambda kv: -kv[1]):
        sectors.append({
            "name": s,
            "cap_eok": round(v / 1e5, 1),
            "pct": round(v / total * 100, 2) if total else 0.0,
        })

    # 기준일: 파일 수정 시각
    import datetime as _dt
    as_of = _dt.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")

    return {
        "as_of": as_of,
        "k70": k70,
        "k150": k150,
        "new": new,
        "out": out,
        "out_pos": out_pos,
        "sectors": sectors,
        "total_cap_eok": round(total / 1e5, 1),
    }


# ---------------------------------------------------------------------------
# 1-4) 종목 마스터 (검색용) + 일봉 OHLCV (FDR)
# ---------------------------------------------------------------------------
_STOCK_MASTER = None  # {code: {name, code, market, price, cap_mil, sheet}}


def _build_stock_master():
    """엑셀에서 종목명 추출 → FDR 로 코드 매핑 → JSON 캐시."""
    if not os.path.isdir(DATA_DIR):
        return {}
    norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}
    key = _nfc(SEMI_FILE)
    if key not in norm_map:
        return {}
    path = os.path.join(DATA_DIR, norm_map[key])

    # 엑셀에서 (종목명, 종가, 시총백만원, 시트명) 추출
    wb = load_workbook(path, read_only=True, data_only=True)
    raw = {}  # {종목명: {price, cap_mil, sheet}}
    for sn in wb.sheetnames:
        if sn.lower().startswith("sheet"):
            continue
        try:
            ws = wb[sn]
            name = ws["C3"].value
            price = ws["C4"].value
            cap_mil = ws["C5"].value
        except Exception:
            continue
        if not name:
            continue
        name = str(name).strip()
        if name in raw:
            continue
        raw[name] = {"price": price, "cap_mil": cap_mil, "sheet": sn}

    # 캐시 파일이 있으면 매핑 결과 재사용
    cache_path = os.path.join(DATA_DIR, STOCK_CODES_CACHE)
    cached = {}
    if os.path.isfile(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cached = json.load(f)
        except Exception:
            cached = {}

    # 누락된 종목만 FDR 호출해서 매핑
    missing = [n for n in raw if _nfc(n) not in cached]
    if missing:
        try:
            import FinanceDataReader as fdr
            listing = fdr.StockListing("KRX")
            listing["_n"] = listing["Name"].astype(str).map(_nfc).str.strip()
            for n in missing:
                nn = _nfc(n).strip()
                hit = listing[listing["_n"] == nn]
                if len(hit) == 0:
                    hit = listing[listing["_n"].str.replace(" ", "") == nn.replace(" ", "")]
                if len(hit) > 0:
                    r = hit.iloc[0]
                    cached[nn] = {"code": str(r["Code"]), "market": str(r.get("Market", ""))}
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(cached, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[stock-master] FDR 매핑 실패: {e}")

    # 최종 master 조립 (code 키)
    master = {}
    for n, info in raw.items():
        c = cached.get(_nfc(n).strip())
        if not c:
            continue
        master[c["code"]] = {
            "code": c["code"],
            "name": n,
            "market": c.get("market", ""),
            "price": info.get("price"),
            "cap_mil": info.get("cap_mil"),
            "sheet": info.get("sheet"),
        }
    return master


def get_stock_master():
    """프로세스 단위 캐시. 첫 호출에서만 빌드.

    반도체 엑셀 + Market_Screening 엑셀의 종목 모두 합쳐 검색 가능하게 한다.
    """
    global _STOCK_MASTER
    if _STOCK_MASTER is None:
        _STOCK_MASTER = _build_stock_master()
        # Market_Screening 엑셀의 종목도 추가 (매일 갱신될 수 있어 매번 시도)
    # 매 호출 시 Market 엑셀 종목 병합 (load_market_data가 캐시 사용하므로 빠름)
    try:
        m = load_market_data()
        if m and m.get("stocks_by_code"):
            merged = dict(_STOCK_MASTER)
            for code, s in m["stocks_by_code"].items():
                if code in merged:
                    continue
                # cap_eok 단위는 억원. price는 원.
                merged[code] = {
                    "code": code,
                    "name": s.get("name", code),
                    "market": "KRX",
                    "price": s.get("price"),
                    "cap_mil": (s.get("cap_eok") or 0) * 100,  # 억원 → 백만원
                    "sheet": None,
                }
            return merged
    except Exception as e:
        print(f"[stock-master] Market 엑셀 병합 실패: {e}")
    return _STOCK_MASTER


# 종목 코드별 야후 suffix(.KS/.KQ) 캐시. 첫 호출 시 자동 감지.
_YF_SUFFIX_CACHE = {}
_OHLCV_CACHE = {}          # (code, years) -> (epoch, bars)
_OHLCV_TTL = 900           # 시세 결과 캐시 유효시간(초). 일봉이라 15분이면 충분

# ---------------------------------------------------------------------------
# KIS (한국투자증권) Open API 클라이언트 — 국내 일봉 조회
# ---------------------------------------------------------------------------
_KIS_TOKEN = {"value": None, "exp": 0}


def _kis_base_url():
    env = os.environ.get("KIS_ENV", "prod").lower()
    if env in ("paper", "vts", "mock", "모의"):
        return "https://openapivts.koreainvestment.com:29443"
    return "https://openapi.koreainvestment.com:9443"


def _kis_get_token():
    """KIS access token (24h 만료) 발급/재사용. 없거나 만료 임박이면 새로 발급."""
    import time as _time
    import json as _json
    import requests

    key = os.environ.get("KIS_APP_KEY")
    secret = os.environ.get("KIS_APP_SECRET")
    if not (key and secret):
        return None
    now = _time.time()
    if _KIS_TOKEN["value"] and now < _KIS_TOKEN["exp"] - 300:
        return _KIS_TOKEN["value"]
    try:
        r = requests.post(
            _kis_base_url() + "/oauth2/tokenP",
            headers={"content-type": "application/json; charset=utf-8"},
            data=_json.dumps({
                "grant_type": "client_credentials",
                "appkey": key, "appsecret": secret,
            }),
            timeout=15,
        )
        if not r.ok:
            print(f"[kis] token 발급 실패 HTTP {r.status_code}: {r.text[:200]}")
            return None
        data = r.json()
        _KIS_TOKEN["value"] = data.get("access_token")
        _KIS_TOKEN["exp"] = now + int(data.get("expires_in", 86400))
        return _KIS_TOKEN["value"]
    except Exception as e:
        print(f"[kis] token 예외: {e}")
        return None


def _kis_daily_ohlcv(code, years=2):
    """KIS 일봉 조회 (수정주가). 필요한 만큼 페이지네이션.

    반환: [{time, open, high, low, close, volume}, ...] 오름차순. 실패 시 None.

    API: /uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice (tr_id FHKST03010100)
      · 한 번에 최대 100영업일. 시작·종료일 지정 시 그 범위의 최근 100건 반환.
      · 더 긴 히스토리는 종료일을 앞당겨 반복 호출로 이어붙임.
    """
    import datetime as _dt
    import requests

    token = _kis_get_token()
    key = os.environ.get("KIS_APP_KEY")
    secret = os.environ.get("KIS_APP_SECRET")
    if not (token and key and secret):
        return None

    url = _kis_base_url() + "/uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice"
    end_dt = _dt.date.today()
    start_dt = end_dt - _dt.timedelta(days=int(365 * years))

    all_rows = {}   # date -> row (중복 방지)
    cursor_end = end_dt
    safety = 0
    # 5년치 ≈ 1250 거래일 = 100건 × 13페이지. 여유 두어 25 로 설정.
    while cursor_end >= start_dt and safety < 25:
        safety += 1
        try:
            r = requests.get(url,
                headers={
                    "content-type": "application/json; charset=utf-8",
                    "authorization": f"Bearer {token}",
                    "appkey": key,
                    "appsecret": secret,
                    "tr_id": "FHKST03010100",
                },
                params={
                    "FID_COND_MRKT_DIV_CODE": "J",
                    "FID_INPUT_ISCD": code,
                    "FID_INPUT_DATE_1": start_dt.strftime("%Y%m%d"),
                    "FID_INPUT_DATE_2": cursor_end.strftime("%Y%m%d"),
                    "FID_PERIOD_DIV_CODE": "D",
                    "FID_ORG_ADJ_PRC": "1",   # 수정주가 반영
                },
                timeout=20,
            )
        except Exception as e:
            print(f"[kis] {code} 요청 예외: {e}")
            return None
        if not r.ok:
            print(f"[kis] {code} HTTP {r.status_code}: {r.text[:200]}")
            return None
        js = r.json() or {}
        if js.get("rt_cd") not in (None, "0"):
            # 오류: 종목 없음 등
            return None
        rows = js.get("output2") or []
        if not rows:
            break
        oldest = None
        for row in rows:
            d = row.get("stck_bsop_date")
            if not d or len(d) < 8:
                continue
            if d in all_rows:
                continue
            all_rows[d] = {
                "time":  f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                "open":  float(row.get("stck_oprc") or 0),
                "high":  float(row.get("stck_hgpr") or 0),
                "low":   float(row.get("stck_lwpr") or 0),
                "close": float(row.get("stck_clpr") or 0),
                "volume": int(float(row.get("acml_vol") or 0)),
            }
            oldest = d if (oldest is None or d < oldest) else oldest
        if not oldest:
            break
        # 이전 창으로 이동 — 가장 오래된 날짜의 하루 전으로
        prev = _dt.datetime.strptime(oldest, "%Y%m%d").date() - _dt.timedelta(days=1)
        if prev < start_dt:
            break
        cursor_end = prev
        if len(rows) < 100:
            break            # 이번 페이지가 100건 미만이면 더 없음

    if not all_rows:
        return None
    return sorted(all_rows.values(), key=lambda x: x["time"])


def load_ohlcv(code, years=2, market=None):
    """국내 일봉 조회. KIS Open API 우선, 실패 시 yfinance fallback.

    KIS 는 정확한 KRX 수정주가·거래량을 반환. 계정 키(.env KIS_APP_KEY/SECRET)
    가 없거나 API 오류 시 yfinance(.KS/.KQ) 로 자동 폴백.

    결과: [{time, open, high, low, close, volume}, ...] 오름차순.
    """
    import datetime as _dt
    import time as _time

    # 1) 결과 캐시 확인 (가장 큰 속도 개선)
    now = _time.time()
    hit = _OHLCV_CACHE.get((code, years))
    if hit and now - hit[0] < _OHLCV_TTL:
        return hit[1]

    # 2) KIS 시도 (계정 키 있을 때만)
    if os.environ.get("KIS_APP_KEY") and os.environ.get("KIS_APP_SECRET"):
        try:
            bars = _kis_daily_ohlcv(code, years=years)
        except Exception as e:
            print(f"[ohlcv] KIS {code} 예외: {e}")
            bars = None
        if bars:
            _OHLCV_CACHE[(code, years)] = (now, bars)
            return bars

    # 3) yfinance fallback
    import yfinance as yf

    end = _dt.date.today()
    start = end - _dt.timedelta(days=int(365 * years))

    # 2) suffix 결정 우선순위: 확정 캐시 > 시장(코스피/코스닥) 힌트 > .KS → .KQ
    cached = _YF_SUFFIX_CACHE.get(code)
    candidates = []
    if cached:
        candidates.append(cached)
    mk = (market or "").upper()
    if "KOSDAQ" in mk or "KONEX" in mk:
        hint = ".KQ"
    elif "KOSPI" in mk or "KS" in mk:
        hint = ".KS"
    else:
        hint = None
    if hint and hint not in candidates:
        candidates.append(hint)
    for s in (".KS", ".KQ"):
        if s not in candidates:
            candidates.append(s)

    df = None
    chosen = None
    for suffix in candidates:
        try:
            d = yf.Ticker(code + suffix).history(
                start=start, end=end + _dt.timedelta(days=1), auto_adjust=False)
            if d is not None and len(d) > 0:
                df = d
                chosen = suffix
                break
        except Exception as e:
            print(f"[ohlcv] {code}{suffix} 실패: {e}")
            continue

    if df is None:
        return []

    _YF_SUFFIX_CACHE[code] = chosen
    df = df.reset_index()
    out = []
    for _, r in df.iterrows():
        d = r["Date"]
        ts = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
        try:
            out.append({
                "time": ts,
                "open": float(r["Open"]),
                "high": float(r["High"]),
                "low": float(r["Low"]),
                "close": float(r["Close"]),
                "volume": int(r["Volume"]) if not pd.isna(r["Volume"]) else 0,
            })
        except (TypeError, ValueError):
            continue
    if out:                                  # 정상 결과만 캐시 (일시적 실패는 캐시 안 함)
        _OHLCV_CACHE[(code, years)] = (now, out)
    return out


def _norm(s):
    return str(s).lower().replace(" ", "").replace("(", "").replace(")", "")


def _is_tmf(fund_name):
    """펀드명에 'tmf'가 포함되면 TMF 펀드로 간주 (대소문자 무시)."""
    return "tmf" in str(fund_name).lower()


# KRX 정보데이터시스템의 KSIC 기반 분류명 → 펀드/KRX 22업종 명칭 매핑.
# 정규화(콤마/공백/가운데점 제거) 후 키와 비교한다.
_KSIC_TO_KRX22 = {
    "운송장비부품": "운수장비",
    "운송장비": "운수장비",
    "운송창고": "운수창고",
    "음식료담배": "음식료품",
    "섬유의류": "섬유의복",
    "기타금융": "금융업",
    "제약": "의약품",
    "기계장비": "기계",
    "유통": "유통업",
    "건설": "건설업",
    "통신": "통신업",
    "IT서비스": "서비스업",
    "일반서비스": "서비스업",
    "오락문화": "서비스업",
    "미디어": "서비스업",
    "교육서비스": "서비스업",
    "부동산": "서비스업",
    "사업서비스": "서비스업",
    "금속": "철강금속",
    "의료정밀기기": "의료정밀",
    "전기가스수도": "전기가스업",
    "전기가스": "전기가스업",
}


def _normalize_sector(s):
    """업종명 정규화: 콤마/공백/가운데점 제거 후 KSIC → KRX 22업종 매핑.

    예) '전기,전자' → '전기전자', '전기·전자' → '전기전자',
        '운송장비·부품' → '운수장비', '기타금융' → '금융업'
    """
    if s is None:
        return "기타"
    s = str(s).strip()
    for ch in (",", " ", "·", "ㆍ", "-", "_", "/"):
        s = s.replace(ch, "")
    if not s:
        return "기타"
    return _KSIC_TO_KRX22.get(s, s)


def load_kospi_sectors():
    """KOSPI 섹터 비중을 GICS 11분류 기준 [{name, pct}] 로 반환.

    우선순위:
      ① Market_Screening 엑셀의 종목 시총을 GICS 대분류로 합산 (마스터)
      ② data/코스피섹터.xlsx (KSIC) — 레거시 fallback
      ③ KOSPI_SECTOR_DEFAULT — 최종 fallback
    """
    # ① Market_Screening 기반 (세분화된 GICS 분류)
    m = load_market_data()
    if m and m.get("stocks_by_code"):
        by_sec = {}
        for s in m["stocks_by_code"].values():
            sector = _gics_to_sector(s.get("gics", ""))
            if not sector:
                continue
            by_sec[sector] = by_sec.get(sector, 0.0) + (s.get("cap_eok") or 0.0)
        total = sum(by_sec.values())
        if total > 0:
            rows = [{"name": n, "pct": round(v / total * 100, 2)}
                    for n, v in by_sec.items()]
            rows.sort(key=lambda x: -x["pct"])
            return rows

    # ② 레거시 — 코스피섹터.xlsx (KSIC)
    rows = None
    if os.path.isdir(DATA_DIR):
        norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}
        key = _nfc(KOSPI_SECTOR_FILE)
        if key in norm_map:
            try:
                df = pd.read_excel(os.path.join(DATA_DIR, norm_map[key]))
                cols = df.columns.tolist()
                name_col = next((c for c in cols if "업종" in str(c) or "섹터" in str(c)), None)
                cap_col = next((c for c in cols if "시가총액" in str(c) or "시총" in str(c)), None)
                market_col = next((c for c in cols if "시장구분" in str(c)), None)
                if name_col and cap_col:
                    work = df.copy()
                    if market_col:
                        work = work[work[market_col].astype(str).str.upper().str.contains("KOSPI", na=False)]
                    work[cap_col] = pd.to_numeric(work[cap_col], errors="coerce")
                    work = work.dropna(subset=[name_col, cap_col])
                    total = float(work[cap_col].sum())
                    if total > 0:
                        rows = []
                        for nm, v in work.groupby(name_col)[cap_col].sum().items():
                            rows.append({"name": _normalize_sector(nm),
                                         "pct": round(float(v) / total * 100, 4)})
            except Exception as e:
                print(f"[kospi-sector] 레거시 xlsx 실패: {e}")

    # ③ 최종 fallback
    if not rows:
        rows = [{"name": _normalize_sector(k), "pct": float(v)}
                for k, v in KOSPI_SECTOR_DEFAULT.items()]

    rows.sort(key=lambda x: -x["pct"])
    return rows


# KRX 22업종 → 세분화된 GICS 매핑. 펀드 엑셀이 KRX 분류라 GICS 로 변환할 때 사용.
# 산업재/IT 는 세분류 (없으면 "기타 산업재"/"기타 IT").
_KRX_TO_GICS = {
    "음식료품": "필수소비재", "음식료담배": "필수소비재",
    "섬유의복": "자유소비재", "섬유의류": "자유소비재",
    "종이목재": "소재",
    "화학": "소재",
    "의약품": "헬스케어", "제약": "헬스케어",
    "비금속광물": "소재", "비금속": "소재",
    "철강금속": "소재", "금속": "소재",
    "기계": "기계", "기계장비": "기계",
    "전기전자": "기타 IT",
    "의료정밀": "헬스케어", "의료정밀기기": "헬스케어",
    "운수장비": "자유소비재", "운송장비": "자유소비재", "운송장비부품": "자유소비재",
    "유통업": "자유소비재", "유통": "자유소비재",
    "전기가스업": "유틸리티", "전기가스수도": "유틸리티", "전기가스": "유틸리티",
    "건설업": "건설", "건설": "건설",
    "운수창고": "항공 및 해운", "운송창고": "항공 및 해운",
    "통신업": "통신서비스", "통신": "통신서비스",
    "금융업": "금융", "기타금융": "금융",
    "은행": "금융", "증권": "금융", "보험": "금융",
    "서비스업": "기타 IT",
    "일반서비스": "기타 IT", "IT서비스": "기타 IT",
    "오락문화": "자유소비재", "미디어": "자유소비재",
    "교육서비스": "자유소비재",
    "사업서비스": "상업서비스와 공급품",
    "부동산": "부동산",
    "제조업": "기타 산업재",
}


# Market_Screening 에 없는 종목들의 수동 GICS 매핑. data/manual_gics.json 으로 확장 가능.
_MANUAL_GICS_BUILTIN = {
    "064400": "기타 IT",                # LG씨엔에스 (IT 서비스)
    "086450": "헬스케어",                # 동국제약
    "178320": "반도체와 반도체장비",      # 서진시스템 (반도체 장비)
    "236200": "기타 IT",                # 슈프리마 (보안 솔루션)
    "365660": "헬스케어",                # 레몬헬스케어
    "394420": "헬스케어",                # 리센스메디컬
    "439960": "기계",                    # 코스모로보틱스공모주 (로봇)
    "461300": "자유소비재",              # 아이스크림미디어
    "473980": "자유소비재",              # 노머스 (엔터테인먼트)
    "493280": "헬스케어",                # 아이엠바이오로직스
    "0082N0": "헬스케어",                # 카나프테라퓨틱스 (신약개발)
    "0007J0": "헬스케어",                # 인벤테라 (바이오)
    "408470": "금융",                    # 한패스 (송금/핀테크)
}


def _load_manual_gics():
    """data/manual_gics.json 에서 사용자 정의 종목코드→GICS 매핑 로드.
    형식: {"종목코드": "섹터명", ...}
    """
    if not os.path.isdir(DATA_DIR):
        return {}
    path = os.path.join(DATA_DIR, "manual_gics.json")
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        # 종목코드 정규화 (6자리 zero-pad)
        return {str(k).zfill(6): v for k, v in data.items() if v in GICS_SECTORS.values()}
    except Exception as e:
        print(f"[manual_gics] 로드 실패: {e}")
        return {}


def get_stock_gics_map():
    """종목코드 → 섹터명. 현재 모드(detail/big) 적용.

    우선순위:
      1) data/manual_gics.json (사용자 수동)
      2) 코드 내 _MANUAL_GICS_BUILTIN
      3) Market_Screening P 시트의 6자리 WICS 코드 → _gics_to_sector
      4) Market_Screening 5.market 시트의 6자리 WICS 코드 → _gics_to_sector
    """
    out = {}
    # 1) 사용자 JSON (값을 모드에 맞게 변환)
    for code, name in _load_manual_gics().items():
        out[code] = _apply_sector_mode(name)
    # 2) 빌트인 수동 매핑
    for code, name in _MANUAL_GICS_BUILTIN.items():
        out.setdefault(code, _apply_sector_mode(name))

    m = load_market_data()
    if not m:
        return out
    # 3) P 시트 (6자리 GICS) — _gics_to_sector 가 모드 자동 적용
    for code, gics_code in (m.get("p_gics") or {}).items():
        if code in out:
            continue
        sector = _gics_to_sector(gics_code)
        if sector:
            out[code] = sector
    # 4) 5.market 시트
    for code, s in (m.get("stocks_by_code") or {}).items():
        if code in out:
            continue
        sector = _gics_to_sector(s.get("gics", ""))
        if sector:
            out[code] = sector
    return out


# 수동 GICS 코드 override — 시장 데이터에 없거나 잘못 분류된 종목을 강제 매핑.
# 참고: `_MANUAL_GICS_BUILTIN` 은 라벨 매핑, 여기는 raw GICS 코드 매핑.
_MANUAL_GICS_CODE_BUILTIN = {
    "064400": "G502050",   # LG씨엔에스 → 네이버와 같은 인터랙티브 미디어 (커뮤니케이션서비스)
    "473980": "G502020",   # 노머스 → 에스엠·JYP 와 같은 엔터테인먼트 (커뮤니케이션서비스)
}


def get_stock_gics_code_map():
    """종목코드 → WICS GICS 코드(예: 'G453010'). wics_level 비교용.

    우선순위:
      1) `_MANUAL_GICS_CODE_BUILTIN` (수동 override)
      2) Market_Screening P 시트
      3) Market_Screening 5.market 시트
      4) 우선주 자동 매칭: 코드 끝자리가 5/7/9 인 종목은 본주(끝자리 0)의 GICS 상속
    """
    out = {}
    # 1) 수동 override (최우선)
    for code, gc in _MANUAL_GICS_CODE_BUILTIN.items():
        out[code] = gc

    m = load_market_data()
    if not m:
        return out
    # 2) P 시트
    for code, gics_code in (m.get("p_gics") or {}).items():
        if code in out:
            continue
        if isinstance(gics_code, str) and gics_code.startswith("G"):
            out[code] = gics_code
    # 3) 5.market
    for code, s in (m.get("stocks_by_code") or {}).items():
        if code in out:
            continue
        gics_code = s.get("gics", "")
        if isinstance(gics_code, str) and gics_code.startswith("G"):
            out[code] = gics_code

    # 4) 우선주 fallback — 마지막 자리가 5/7/9 이면 본주(마지막 0)의 GICS 상속.
    #    예: 005935 삼성전자우 → 005930 삼성전자 의 코드 사용.
    #    이 로직은 이미 out 에 있는 종목 제외, 매핑 없는 종목만 시도.
    for code in list(out.keys()):
        pass  # (dummy)
    # 원본 out 을 훑을 때 새 종목이 추가되지 않도록 base_codes 사용
    base_codes = set(out.keys())
    def _try_pref(code):
        if not code or len(code) != 6 or not code.isdigit():
            return None
        last = code[-1]
        if last in ("5", "7", "9"):
            base = code[:-1] + "0"
            return out.get(base)
        return None
    # 홀딩스에 실제 있는 종목 중 매핑 없는 것들을 뽑기 위해 전체 코드 후보 필요.
    # → 여기선 P 시트 + 5.market 커버리지 밖 우선주만 처리. 홀딩스 종목은 이 함수 호출자가
    #   view_sectors_by_level 에서 매핑 lookup 시 자동으로 처리되도록 아래에서 preferred 저장.
    # 접근 방식 변경: known preferred code (본주에 매핑 있음) 을 스캔해 우선주 코드도 채움.
    for base in list(base_codes):
        if len(base) == 6 and base.isdigit() and base.endswith("0"):
            gc = out[base]
            for suffix in ("5", "7", "9"):
                pref = base[:-1] + suffix
                if pref not in out:
                    out[pref] = gc
    return out


_KOSPI200_CACHE = {"key": None, "codes": None, "stocks": None}


# ---------------------------------------------------------------------------
# 블룸버그 글로벌 피어 밸류에이션 테이블 (팀 공유 폴더 → 로컬 sanitized 캐시)
# ---------------------------------------------------------------------------
BLOOMBERG_SHARED_PATH = r"Z:\Team Investment\00. Screening\★Bloomberg Valuation Table(수정중).xlsx"
BLOOMBERG_SECTORS = [
    "Energy", "Materials", "Industrials", "Consumer Discretionary",
    "Consumer Staples", "Information Technology", "Communication Services",
    "Healthcare", "Financials",
]
_BLOOMBERG_CACHE = {"key": None, "data": None}


def _sanitize_xlsx(src_path, dst_path):
    """openpyxl 이 파일의 docProps/custom.xml 파싱 시 crash 하는 버그 우회.
    xlsx 는 zip 이라서 문제되는 엔트리만 제외하고 재저장.
    """
    import zipfile
    with zipfile.ZipFile(src_path, 'r') as zin:
        with zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == 'docProps/custom.xml':
                    continue
                zout.writestr(item, zin.read(item.filename))


def load_bloomberg_peers():
    """블룸버그 밸류에이션 테이블 로드. 팀 공유폴더 우선, 실패 시 로컬 파일 fallback.
    각 업종 시트에서 (Ticker · Company · Price · 1M/3M/1Y · Mkt cap · PER 26E/27E) 추출.
    원본 mtime 기반 캐시.
    """
    # 소스 선택: Z 드라이브 원본 우선, 없으면 로컬 사본
    src = BLOOMBERG_SHARED_PATH
    if not os.path.isfile(src):
        local_candidates = [
            os.path.join(DATA_DIR, "★Bloomberg Valuation Table.xlsx"),
            os.path.join(DATA_DIR, "Bloomberg Valuation Table.xlsx"),
        ]
        src = next((p for p in local_candidates if os.path.isfile(p)), None)
        if not src:
            return None

    try:
        mtime = os.path.getmtime(src)
    except OSError:
        return None
    if _BLOOMBERG_CACHE.get("key") == (src, mtime) and _BLOOMBERG_CACHE.get("data") is not None:
        return _BLOOMBERG_CACHE["data"]

    # sanitize → 로컬 캐시 파일
    sanitized = os.path.join(DATA_DIR, "_bloomberg_valuation_sanitized.xlsx")
    try:
        _sanitize_xlsx(src, sanitized)
    except Exception as e:
        print(f"[bloomberg] sanitize 실패: {e}")
        return None

    result = {}
    for sec in BLOOMBERG_SECTORS:
        try:
            df = pd.read_excel(sanitized, sheet_name=sec, header=None)
        except Exception as e:
            print(f"[bloomberg] {sec} 읽기 실패: {e}")
            result[sec] = {"subs": []}
            continue
        # 시트 순회 — 각 시트는 [소분류(중분류) 헤더] 뒤에 [종목 행들] 이 오는 반복 구조.
        # 헤더 행: A None, D 문자열, F None (Source: Bloomberg 는 종료 신호).
        # 종목 행: A 에 'XXX EQUITY' ticker.
        subs = []            # [{name, stocks:[...]}, ...]
        current = None       # 현재 서브카테고리
        for i in range(5, len(df)):
            row = df.iloc[i]
            a = row[0] if len(row) > 0 else None
            d = row[3] if len(row) > 3 else None
            f = row[5] if len(row) > 5 else None

            # 헤더 행 인식
            is_header = (pd.isna(a) and isinstance(d, str) and d.strip()
                         and pd.isna(f))
            if is_header:
                name = d.strip()
                if name.lower().startswith("source"):
                    break                # 시트 종료
                current = {"name": name, "stocks": []}
                subs.append(current)
                continue

            # 종목 행
            if isinstance(a, str) and "EQUITY" in a.upper():
                if current is None:
                    # 헤더 없이 종목이 먼저 나오면 대분류명으로 임시 생성
                    current = {"name": sec, "stocks": []}
                    subs.append(current)
                # 외국 종목은 D(Company) 컬럼이 비어있고 Ticker 만 있음.
                # 이 경우 티커의 첫 토큰(GOOGL, META 등)을 name 대용으로 사용.
                if isinstance(d, str) and d.strip():
                    name = d.strip()
                else:
                    name = a.strip().split()[0]
                # 컬럼 매핑 (2026-07-08 갱신본 기준 · 1W 추가로 뒤 컬럼 1칸씩 밀림)
                #   c5 Price · c6 1W · c7 1M · c8 3M · c9 1Y
                #   c11 Mkt cap Local · c12 Mkt cap USDb
                #   c14 PER 26E · c15 PER 27E · c16 PER 28E
                #   c18 PBR 26E · c19 PBR 27E · c20 PBR 28E
                current["stocks"].append({
                    "ticker": a.strip(),
                    "name": name.strip(),
                    "currency": str(row[4]).strip() if len(row) > 4 and pd.notna(row[4]) else "",
                    "price":   _to_float(row[5])  if len(row) > 5 else None,
                    "chg_1w":  _to_float(row[6])  if len(row) > 6 else None,
                    "chg_1m":  _to_float(row[7])  if len(row) > 7 else None,
                    "chg_3m":  _to_float(row[8])  if len(row) > 8 else None,
                    "chg_1yr": _to_float(row[9])  if len(row) > 9 else None,
                    "mkt_cap_local": _to_float(row[11]) if len(row) > 11 else None,
                    "mkt_cap_usd":   _to_float(row[12]) if len(row) > 12 else None,
                    "per_26e": _to_float(row[14]) if len(row) > 14 else None,
                    "per_27e": _to_float(row[15]) if len(row) > 15 else None,
                    "per_28e": _to_float(row[16]) if len(row) > 16 else None,
                    "pbr_26e": _to_float(row[18]) if len(row) > 18 else None,
                })
        # 빈 서브 제거
        subs = [s for s in subs if s["stocks"]]
        result[sec] = {"subs": subs}

    _BLOOMBERG_CACHE["key"] = (src, mtime)
    _BLOOMBERG_CACHE["data"] = {"sectors": result, "source": src, "mtime": mtime}
    return _BLOOMBERG_CACHE["data"]


def load_kospi200():
    """data/ 에서 최신 KODEX_200_*.xls 을 읽어 KOSPI 200 구성 종목 반환.

    반환: {"codes": set[str], "stocks": [{"code","name","weight"}, ...]}
      - codes: 종목코드(6자리) set — 우리 펀드 종목과의 교집합 판별용
      - stocks: KODEX 비중 내림차순 (원화예금 등 종목코드가 6자리 아닌 항목 제외)
    파일 mtime 기반 캐시. 파일 없으면 빈 결과.
    """
    if not os.path.isdir(DATA_DIR):
        return {"codes": set(), "stocks": []}
    candidates = []
    for name in os.listdir(DATA_DIR):
        n = _nfc(name).lower()
        if n.startswith("kodex_200") and (n.endswith(".xls") or n.endswith(".xlsx")):
            candidates.append(os.path.join(DATA_DIR, name))
    if not candidates:
        return {"codes": set(), "stocks": []}
    path = sorted(candidates)[-1]  # 파일명 최신 (KODEX_200_YYYYMMDD)

    mtime = os.path.getmtime(path)
    key = (path, mtime)
    if _KOSPI200_CACHE.get("key") == key and _KOSPI200_CACHE.get("codes") is not None:
        return {"codes": _KOSPI200_CACHE["codes"], "stocks": _KOSPI200_CACHE["stocks"]}

    try:
        df = pd.read_excel(path, header=2)
    except Exception as e:
        print(f"[kospi200] 로드 실패: {e}")
        return {"codes": set(), "stocks": []}

    codes = set()
    stocks = []
    for _, r in df.iterrows():
        raw = r.get("종목코드")
        if raw is None:
            continue
        code = str(raw).strip()
        if not code.isdigit() or len(code) > 6:
            continue                       # 원화예금(KRD...) 등 제외
        code = code.zfill(6)
        codes.add(code)
        stocks.append({
            "code": code,
            "name": str(r.get("종목명") or code).strip(),
            "weight": round(float(r.get("비중(%)") or 0), 4),
        })
    stocks.sort(key=lambda s: -s["weight"])
    _KOSPI200_CACHE["key"] = key
    _KOSPI200_CACHE["codes"] = codes
    _KOSPI200_CACHE["stocks"] = stocks
    return {"codes": codes, "stocks": stocks}


def get_stock_perf_map():
    """종목코드 → 수익률 dict {wk1, m1, m3}. Market_Screening 5.market + P 시트 병합.

    - 5.market 시트: ~450 종목 (안다 관심 종목 + 대형주). 우선순위.
    - P 시트: ~3,800 종목 (KOSPI/KOSDAQ 광범위). 5.market 에 없는 종목만 fallback.
    """
    m = load_market_data()
    if not m:
        return {}
    out = {}
    # 1) 5.market 우선
    for code, s in (m.get("stocks_by_code") or {}).items():
        out[code] = {
            "wk1": s.get("wk1"),
            "m1": s.get("m1"),
            "m3": s.get("m3"),
        }
    # 2) P 시트 fallback — 5.market 에 없는 종목만
    for code, p in (m.get("p_perf") or {}).items():
        if code not in out:
            out[code] = p
    return out


def get_market_benchmark():
    """알파 산점도용 벤치마크 — Market_Screening P 시트의 KODEX 200 (A069500) 행.

    P 시트에서 실제 KODEX 200 ETF의 시장가 기준 수익률을 그대로 사용.
    P 시트에 KODEX 200 행이 없으면 대형주 지수로 fallback.
    """
    m = load_market_data()
    if not m:
        return None
    kdx = m.get("kodex200")
    if kdx and any(kdx.get(k) is not None for k in ("wk1", "m1", "m3")):
        return {
            "name": kdx.get("name") or "KODEX 200",
            "wk1": kdx.get("wk1"),
            "m1":  kdx.get("m1"),
            "m3":  kdx.get("m3"),
        }
    # Fallback: 시장 매크로의 대형주 / KOSPI
    for target in ("대형주", "KOSPI"):
        for it in m.get("dom_idx", []):
            if it.get("name") == target:
                return {
                    "name": it["name"], "wk1": it.get("d5"),
                    "m1": it.get("m1"), "m3": it.get("m3"),
                }
    return None


def build_alpha_scatters(codes_pcts, perf_map, bm):
    """알파 산점도용 데이터 두 종류(1M×3M / 5D×1M) 생성.

    codes_pcts: [(code, name, weight_pct), ...] — 화면 표시 종목 리스트
    perf_map: get_stock_perf_map() 결과
    bm: get_market_benchmark() 결과

    반환: {
      "bm": {"name": ..., "wk1": ..., "m1": ..., "m3": ...},
      "points": [{code, name, weight, wk1, m1, m3}, ...],
    }
    """
    points = []
    for code, name, w in codes_pcts:
        code = str(code or "").zfill(6)
        p = perf_map.get(code)
        if not p:
            continue
        wk1, m1, m3 = p.get("wk1"), p.get("m1"), p.get("m3")
        if wk1 is None and m1 is None and m3 is None:
            continue
        points.append({
            "code": code, "name": name,
            "weight": round(float(w or 0), 2),
            "wk1": round(float(wk1), 2) if wk1 is not None else None,
            "m1":  round(float(m1),  2) if m1  is not None else None,
            "m3":  round(float(m3),  2) if m3  is not None else None,
        })
    return {"bm": bm, "points": points}


def kospi_sectors_by_level(level):
    """Market_Screening 의 stocks_by_code 를 level(big/mid/small)별로 합산해
    [{name, pct}] 반환. load_kospi_sectors 의 wics 일반화 버전.
    """
    m = load_market_data()
    if not (m and m.get("stocks_by_code")):
        return []
    by_sec = {}
    for s in m["stocks_by_code"].values():
        sec = _wics_classify(s.get("gics", ""), level)
        if not sec:
            continue
        by_sec[sec] = by_sec.get(sec, 0.0) + (s.get("cap_eok") or 0.0)
    total = sum(by_sec.values())
    if total <= 0:
        return []
    rows = [{"name": n, "pct": round(v / total * 100, 2)} for n, v in by_sec.items()]
    rows.sort(key=lambda x: -x["pct"])
    return rows


def view_sectors_by_level(holdings_df, level, gics_code_map):
    """펀드/전사 holdings 를 level(big/mid/small)별로 합산해 [{name, pct}] 반환.

    비중 기준은 **순자산비** (펀드 NAV 대비 편입 비중, 현금 포함이라 합이 100% 미만).
    holdings_df 컬럼: '종목코드', '순자산비'. gics_code_map: 종목코드 → GICS 코드.
    전사(여러 펀드) 는 각 펀드 순자산비를 NAV 로 가중평균해야 하므로, _fund 별
    NAV 비중으로 스케일한다. 매핑 없는 종목은 '비분류' 로 별도 표시.
    """
    if holdings_df is None or len(holdings_df) == 0:
        return []
    df = holdings_df

    # 펀드별 NAV 가중치 (전사 합산 시 각 펀드 순자산비를 NAV 비율로 섞기 위함)
    # NAV_fund = Σ평가액 / (Σ순자산비/100). 단일 펀드면 가중치 1.
    fund_w = {}
    if "_fund" in df.columns:
        navs = {}
        for fn, g in df.groupby("_fund"):
            ev = float(g["평가액"].sum())
            nsb = float(g["순자산비"].sum())
            navs[fn] = ev / (nsb / 100.0) if nsb > 0 else ev
        tot = sum(navs.values())
        if tot > 0:
            fund_w = {fn: v / tot for fn, v in navs.items()}

    by_sec = {}
    for _, r in df.iterrows():
        code = str(r.get("종목코드") or "").zfill(6)
        w = float(r.get("순자산비") or 0)     # 이 펀드 NAV 대비 %
        if w <= 0:
            continue
        if fund_w:                            # 전사: 펀드 NAV 비율로 스케일
            w *= fund_w.get(r.get("_fund"), 0.0)
        gics = gics_code_map.get(code)
        sec = _wics_classify(gics, level) if gics else None
        if not sec:
            sec = "비분류"
        by_sec[sec] = by_sec.get(sec, 0.0) + w
    rows = [{"name": n, "pct": round(v, 2)} for n, v in by_sec.items()]
    rows.sort(key=lambda x: -x["pct"])
    return rows


def kospi_compare(view_sectors, kospi_rows):
    """뷰의 섹터 비중과 KOSPI 섹터 비중을 같은 라벨 축으로 정렬해 비교용 데이터로 변환.

    양쪽 모두 _normalize_sector 로 정규화한 키로 매칭하고, 화면 표시용 라벨은
    KOSPI 원본 이름 우선(있으면) → 펀드 원본 이름 → 정규화 키.
    """
    # 정규화 키 ↔ 원본 이름 매핑 + 비중
    kospi_map, kospi_label = {}, {}
    for r in kospi_rows:
        key = _normalize_sector(r["name"])
        kospi_map[key] = kospi_map.get(key, 0.0) + (r.get("pct") or 0.0)
        kospi_label.setdefault(key, r["name"])  # 첫 등장 원본 라벨

    view_norm, view_label = {}, {}
    for s in view_sectors:
        key = _normalize_sector(s.get("name"))
        view_norm[key] = view_norm.get(key, 0.0) + (s.get("pct") or 0.0)
        view_label.setdefault(key, s.get("name"))

    names_in_kospi = [k for k in kospi_map if kospi_map[k] > 0]
    extras = [k for k in view_norm if k not in names_in_kospi]

    # 펀드의 '코스닥' 류 → KOSPI 벤치마크 = 코스닥 전체 시총 비율
    if any("코스닥" in k for k in extras):
        kosdaq_pct = load_kosdaq_vs_kospi_pct()
        for k in extras:
            if "코스닥" in k:
                kospi_map[k] = kosdaq_pct

    keys = names_in_kospi + sorted(extras)
    labels = [kospi_label.get(k) or view_label.get(k) or k for k in keys]
    return {
        "labels": labels,
        "kospi": [round(kospi_map.get(k, 0.0), 2) for k in keys],
        "firm":  [round(view_norm.get(k, 0.0), 2) for k in keys],
    }


# 코스닥 전체 시총 / 코스피 전체 시총 비율 (펀드 보유 '코스닥종합' 항목의 벤치마크).
_KOSDAQ_PCT_CACHE = {"ts": 0, "value": None}


def load_kosdaq_vs_kospi_pct():
    """KOSDAQ 시총 / KOSPI 시총 × 100. 24시간 캐시.

    사용자 엑셀(data/코스피섹터.xlsx)에 시장구분=KOSDAQ 행이 있으면 그걸 우선,
    없으면 FDR로 KRX 시총 합계 계산.
    """
    import time as _time
    now = _time.time()
    if _KOSDAQ_PCT_CACHE["value"] is not None and (now - _KOSDAQ_PCT_CACHE["ts"]) < 86400:
        return _KOSDAQ_PCT_CACHE["value"]

    value = None
    if os.path.isdir(DATA_DIR):
        norm_map = {_nfc(name): name for name in os.listdir(DATA_DIR)}
        key = _nfc(KOSPI_SECTOR_FILE)
        if key in norm_map:
            try:
                df = pd.read_excel(os.path.join(DATA_DIR, norm_map[key]))
                cols = df.columns.tolist()
                cap_col = next((c for c in cols if "시가총액" in str(c) or "시총" in str(c)
                                or "marcap" in str(c).lower()), None)
                market_col = next((c for c in cols if "시장구분" in str(c) or "시장" in str(c)
                                   or "market" in str(c).lower()), None)
                if cap_col and market_col:
                    df[cap_col] = pd.to_numeric(df[cap_col], errors="coerce")
                    df = df.dropna(subset=[cap_col])
                    mkt = df[market_col].astype(str).str.upper()
                    kospi = float(df[mkt.str.contains("KOSPI", na=False)][cap_col].sum())
                    kosdaq = float(df[mkt.str.contains("KOSDAQ", na=False)][cap_col].sum())
                    if kospi > 0 and kosdaq > 0:
                        value = round(kosdaq / kospi * 100, 2)
            except Exception as e:
                print(f"[kosdaq-vs-kospi] 엑셀 읽기 실패: {e}")

    if value is None:
        try:
            import FinanceDataReader as fdr
            df = fdr.StockListing("KRX").dropna(subset=["Marcap"])
            kospi = float(df[df["Market"] == "KOSPI"]["Marcap"].sum())
            kosdaq = float(df[df["Market"] == "KOSDAQ"]["Marcap"].sum())
            if kospi > 0 and kosdaq > 0:
                value = round(kosdaq / kospi * 100, 2)
        except Exception as e:
            print(f"[kosdaq-vs-kospi] FDR 실패: {e}")

    if value is None:
        value = 20.0  # 모든 경로 실패 시 합리적 추정치

    _KOSDAQ_PCT_CACHE["ts"] = now
    _KOSDAQ_PCT_CACHE["value"] = value
    return value


# ---------------------------------------------------------------------------
# 신용잔고 / 투자자예탁금 (금융투자협회 종합통계, 공공데이터포털 오픈API)
# ---------------------------------------------------------------------------
_KOFIA_BASE = "https://apis.data.go.kr/1160100/service/GetKofiaStatisticsInfoService"
_CREDIT_CACHE = {"ts": 0, "data": None}
_CREDIT_TTL = 6 * 3600   # 6시간 캐시 (일별 데이터라 충분)


_FREESIS_URL = "https://freesis.kofia.or.kr/meta/getMetaDataList.do"


def _freesis_series(obj_nm, start_yyyymmdd, end_yyyymmdd):
    """FreeSIS getMetaDataList.do 직접 호출 → ds1 행 리스트 반환 (백만원).

    obj_nm 예: 'STATSCU0100000070BO'(신용공여) / 'STATSCU0100000060BO'(증시자금).
    행: {TMPV1: 'YYYYMMDD', TMPV2..: 값}. 웹사이트와 동일한 최신값(오픈API보다 하루 빠름).
    """
    import requests
    body = {"dmSearch": {"tmpV40": "1000000", "tmpV41": "1", "tmpV1": "D",
                         "tmpV45": start_yyyymmdd, "tmpV46": end_yyyymmdd, "OBJ_NM": obj_nm}}
    r = requests.post(_FREESIS_URL, json=body, timeout=20,
                      headers={"Content-Type": "application/json",
                               "Referer": "https://freesis.kofia.or.kr/stat/FreeSIS.do"})
    r.raise_for_status()
    return (r.json() or {}).get("ds1") or []


def load_credit_deposit(rows=550):
    """신용잔고(KOSPI/KOSDAQ) + 투자자예탁금 + KOSPI 지수 일별 시계열. 단위: 조원.

    1차: 금투협 FreeSIS 직접 조회 (웹사이트와 동일, 오픈API보다 ~1영업일 빠름)
      · 신용공여(STATSCU0100000070BO): TMPV3=KOSPI 신용융자, TMPV4=KOSDAQ 신용융자
      · 증시자금(STATSCU0100000060BO): TMPV2=투자자예탁금  (값 단위: 백만원)
    2차(폴백): 공공데이터포털 오픈API 2종.
    KOSPI 지수는 FinanceDataReader(KS11)로 보강.
    반환: {"as_of","series":[{d, kospi_credit, kosdaq_credit, deposit, kospi}], "src"} 또는 None.
    """
    import time as _time
    import datetime as _dt
    now = _time.time()
    if _CREDIT_CACHE["data"] and now - _CREDIT_CACHE["ts"] < _CREDIT_TTL:
        return _CREDIT_CACHE["data"]

    def _m(v):   # 백만원 → 조
        try:
            return round(float(v) / 1e6, 3)
        except (TypeError, ValueError):
            return None

    by_d = {}
    src = None

    # ---- 1차: FreeSIS 직접 ----
    try:
        end = _dt.date.today()
        start = end - _dt.timedelta(days=int(rows * 1.6))  # 영업일 감안 여유
        s, e = start.strftime("%Y%m%d"), end.strftime("%Y%m%d")
        credit = _freesis_series("STATSCU0100000070BO", s, e)
        deposit = _freesis_series("STATSCU0100000060BO", s, e)
        for row in credit:
            d = str(row.get("TMPV1") or "")
            if len(d) == 8:
                by_d.setdefault(d, {})["kospi_credit"] = _m(row.get("TMPV3"))
                by_d[d]["kosdaq_credit"] = _m(row.get("TMPV4"))
        for row in deposit:
            d = str(row.get("TMPV1") or "")
            if len(d) == 8:
                by_d.setdefault(d, {})["deposit"] = _m(row.get("TMPV2"))
        if by_d:
            src = "FreeSIS"
    except Exception as ex:
        print(f"[credit] FreeSIS 실패 → 오픈API 폴백: {ex}")

    # ---- 2차: 공공데이터포털 오픈API 폴백 ----
    if not by_d:
        key = os.environ.get("KOFIA_API_KEY")
        if not key:
            return None
        import requests

        def _fetch(op):
            try:
                r = requests.get(f"{_KOFIA_BASE}/{op}",
                    params={"serviceKey": key, "resultType": "json",
                            "numOfRows": rows, "pageNo": 1}, timeout=20)
                if not r.ok:
                    return []
                items = (((r.json() or {}).get("response") or {}).get("body") or {}).get("items") or {}
                it = items.get("item") or []
                return it if isinstance(it, list) else [it]
            except Exception as e:
                print(f"[credit] 오픈API {op} 예외: {e}")
                return []

        def _f(v):
            try:
                return round(float(v) / 1e12, 3)
            except (TypeError, ValueError):
                return None
        for row in _fetch("getGrantingOfCreditBalanceInfo"):
            d = str(row.get("basDt") or "")
            if len(d) == 8:
                by_d.setdefault(d, {})["kospi_credit"] = _f(row.get("crdTrFingScrs"))
                by_d[d]["kosdaq_credit"] = _f(row.get("crdTrFingKosdaq"))
        for row in _fetch("getSecuritiesMarketTotalCapitalInfo"):
            d = str(row.get("basDt") or "")
            if len(d) == 8:
                by_d.setdefault(d, {})["deposit"] = _f(row.get("invrDpsgAmt"))
        if by_d:
            src = "공공데이터포털 오픈API"

    if not by_d:
        return None

    # KOSPI 지수 (FDR) 보강
    kospi_map = {}
    try:
        import FinanceDataReader as fdr
        ds = sorted(by_d.keys())
        start_s = ds[0][:4] + "-" + ds[0][4:6] + "-" + ds[0][6:8]
        idx = fdr.DataReader("KS11", start_s)
        for dt, r in idx.iterrows():
            kospi_map[dt.strftime("%Y%m%d")] = round(float(r["Close"]), 2)
    except Exception as e:
        print(f"[credit] KOSPI 지수 보강 실패: {e}")

    series = []
    for d in sorted(by_d.keys()):
        e = by_d[d]
        series.append({
            "d": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
            "kospi_credit": e.get("kospi_credit"),
            "kosdaq_credit": e.get("kosdaq_credit"),
            "deposit": e.get("deposit"),
            "kospi": kospi_map.get(d),
        })
    out = {"as_of": series[-1]["d"] if series else None, "series": series, "src": src}
    _CREDIT_CACHE["ts"] = now
    _CREDIT_CACHE["data"] = out
    return out


# ---------------------------------------------------------------------------
# 한국추종 ETF 주간 펀드유입액 (EWY / FLKR / EEM) — 팀 공유 엑셀 (매주 갱신)
# ---------------------------------------------------------------------------
ETF_FLOW_SHARED_PATH = r"Z:\Team Investment\00. Screening\한국추종 펀드트래킹.xlsx"
ETF_FLOW_TICKERS = ["EWY", "DRAM"]   # 시트명 = 티커 (FLKR/EEM 은 금액 의미 낮아 제외)
_ETF_FLOW_CACHE = {"key": None, "data": None}


def load_etf_flows():
    """한국추종 ETF들의 주간 펀드유입액. 시트별(EWY/FLKR/EEM) C열=날짜, D열=유입액($백만).

    Z드라이브 원본 우선, 없으면 로컬 사본. 헤더 시작 행이 시트마다 달라도(3~4행)
    'C가 날짜 + D가 숫자'인 행만 추려 견고하게 파싱. mtime 기반 캐시.
    반환: {"as_of", "tickers":[...], "etfs": {TICKER: [{d, flow}]}} 또는 None.
    """
    src = ETF_FLOW_SHARED_PATH
    if not os.path.isfile(src):
        local = os.path.join(DATA_DIR, "한국추종 펀드트래킹.xlsx")
        src = local if os.path.isfile(local) else None
        if not src:
            return None
    try:
        mtime = os.path.getmtime(src)
    except OSError:
        return None
    if _ETF_FLOW_CACHE.get("key") == (src, mtime) and _ETF_FLOW_CACHE.get("data"):
        return _ETF_FLOW_CACHE["data"]

    try:
        xl = pd.ExcelFile(src)
    except Exception as e:
        print(f"[etf-flow] 열기 실패: {e}")
        return None

    etfs, as_of = {}, None
    for tk in ETF_FLOW_TICKERS:
        if tk not in xl.sheet_names:
            continue
        try:
            df = pd.read_excel(xl, sheet_name=tk, header=None)
        except Exception:
            continue
        series = []
        for _, row in df.iterrows():
            d = row[2] if len(row) > 2 else None    # C열 날짜
            v = row[3] if len(row) > 3 else None     # D열 유입액
            # 값 먼저 검증 (숫자 아니면 헤더/빈행)
            try:
                fv = round(float(v), 2)
            except (TypeError, ValueError):
                continue
            # 날짜 파싱 (NaT/빈값 스킵)
            try:
                dt = pd.to_datetime(d)
            except Exception:
                continue
            if pd.isna(dt):
                continue
            series.append({"d": dt.strftime("%Y-%m-%d"), "flow": fv})
        if series:
            series.sort(key=lambda x: x["d"])
            etfs[tk] = series
            as_of = max(as_of, series[-1]["d"]) if as_of else series[-1]["d"]

    if not etfs:
        return None
    out = {"as_of": as_of, "tickers": [t for t in ETF_FLOW_TICKERS if t in etfs], "etfs": etfs}
    _ETF_FLOW_CACHE["key"] = (src, mtime)
    _ETF_FLOW_CACHE["data"] = out
    return out


# ---------------------------------------------------------------------------
# 시장 스크리닝 데이터 (지수/환율/원자재/KOSPI 섹터)
# ---------------------------------------------------------------------------
MARKET_CACHE_SEC = 3600  # 1시간 캐시. 강제 새로고침 시 즉시 갱신.
_MARKET_CACHE = {"ts": 0, "data": None}

# 매크로 항목 이름 → 카테고리 분류.
# data/Market_Screening_*.xlsm 의 1.macro 시트에서 추출된 이름과 매칭한다.
MACRO_CATEGORIES = {
    "dom_idx":    ["KOSPI", "대형주", "중형주", "소형주", "KOSDAQ"],
    "global_idx": ["니케이225", "상해종합지수", "대만가권", "Dow", "NASDAQ",
                   "S&P500", "STOXX50", "FTSE100", "CAC40", "BOVESPA", "HOCHIMIN"],
    "fx":         ["KRW/USD", "JPY/USD", "Euro FX"],
    "comm":       ["WTI", "Gold", "Silver", "Copper", "Aluminium", "Nickel",
                   "Corn", "Wheat"],
    "etc":        ["BDI", "Bitcoin"],
}

# GICS 11 섹터 (대분류). G45 같은 6자리 세부 코드의 앞 2자리(45)로 매핑.
GICS_SECTORS = {
    "10": "에너지",      "15": "소재",       "20": "산업재",
    "25": "자유소비재",  "30": "필수소비재", "35": "헬스케어",
    "40": "금융",        "45": "IT",         "50": "통신서비스",
    "55": "유틸리티",    "60": "부동산",
}

# ===========================================================================
# WICS 표준 분류 — 대(2자리)/중(4자리)/소(6자리)
# ===========================================================================
# WICS 대분류 (G + 2자리). FnGuide WICS 한글 표준명. GICS_SECTORS 와 동일 골격.
WICS_BIG = {
    "G10": "에너지",       "G15": "소재",         "G20": "산업재",
    "G25": "자유소비재",   "G30": "필수소비재",   "G35": "헬스케어",
    "G40": "금융",         "G45": "IT",           "G50": "커뮤니케이션서비스",
    "G55": "유틸리티",
}

# WICS 중분류 (G + 4자리). FnGuide WICS Industry Group 한글 표준명.
# 5.market 시트에 실제 등장하는 prefix만 정리. 누락 prefix는 _wics_classify 가
# 대분류명 + ' (기타)' 로 fallback 한다.
WICS_MID = {
    "G1010": "에너지",
    "G1510": "소재",
    "G2010": "자본재",
    "G2020": "상업서비스와 공급품",
    "G2030": "운송",
    "G2510": "자동차와 부품",
    "G2520": "내구소비재와 의류",
    "G2530": "호텔·레저",
    "G2540": "미디어",
    "G2550": "유통",
    "G2560": "소매",
    "G3010": "음식료·담배 소매",
    "G3020": "식품·음료·담배",
    "G3030": "가정용품과 개인용품",
    "G3510": "헬스케어 장비와 서비스",
    "G3520": "제약·생물공학",
    "G4010": "은행",
    "G4020": "증권",
    "G4030": "보험",
    "G4040": "부동산",
    "G4510": "소프트웨어와 서비스",
    "G4520": "기술 하드웨어와 장비",
    "G4530": "반도체와 반도체장비",
    "G4535": "전기제품 (2차전지)",
    "G4540": "디스플레이",
    "G5010": "통신 서비스",
    "G5020": "미디어·엔터테인먼트",
    "G5510": "유틸리티",
}

# WICS 소분류 (G + 6자리) 매핑은 엑셀 2.Sector 시트에서 동적 추출 → _WICS_SMALL 캐시.
_WICS_SMALL = {}


def _wics_classify(gics_code, level="big"):
    """WICS 코드 → 단계별 분류명.

      level == 'big'   → G + 2자리 (10개 대분류)
      level == 'mid'   → G + 4자리 (~28개 중분류)
      level == 'small' → G + 6자리 (~50개 소분류, 엑셀 2.Sector 매핑)

    매핑이 없으면 (대분류명 + 코드) 로 fallback. None 코드면 None 반환.
    """
    if not gics_code or not isinstance(gics_code, str) or not gics_code.startswith("G"):
        return None
    if level == "big":
        return WICS_BIG.get(gics_code[:3])
    if level == "mid":
        name = WICS_MID.get(gics_code[:5])
        if name:
            return name
        big = WICS_BIG.get(gics_code[:3])
        return f"{big} (기타)" if big else None
    if level == "small":
        name = _WICS_SMALL.get(gics_code[:7])
        if name:
            return name
        # 소분류 매핑 없으면 중분류로 fallback (대표명 그대로 사용)
        return _wics_classify(gics_code, "mid")
    return None

# WICS 6자리 세부 코드 → 세분류명 (산업재 G20, IT G45 만 세분화).
# 그 외 섹터는 대분류 그대로 사용.
WICS_DETAIL = {
    # ===== 산업재 (G20) =====
    "G201030": "건설",
    "G201040": "전기장비",
    "G201050": "복합기업",
    "G201060": "기계",
    "G201070": "조선",
    "G201080": "상업서비스와 공급품",  # 사무용품·환경서비스 등
    "G201010": "상업서비스와 공급품",  # 우주항공과 국방(방산) — 사용자 요청대로 같이 묶음
    "G201020": "기타 산업재",          # 건축자재
    "G202010": "항공 및 해운",         # 항공화물운송과 물류
    "G202020": "항공 및 해운",         # 항공사
    "G202030": "항공 및 해운",         # 해운사
    # G202040(도로와철도), G202050(운송인프라) → 자동으로 "기타 산업재"

    # ===== IT (G45) =====
    "G453010": "반도체와 반도체장비",
    "G452030": "전기제품 (2차전지)",
    "G453520": "전기제품 (2차전지)",   # LG에너지솔루션 등 2차전지 별도 분류
    "G454010": "디스플레이",            # 디스플레이패널
    "G454020": "디스플레이",            # 디스플레이장비 및 부품
    # G45 의 다른 코드(G4510 소프트웨어, G4520 컴퓨터/통신장비, G4535 가전 등)는 "기타 IT"
}


# 세분류 → 대분류 역매핑 (big 모드에서 _MANUAL/_KRX 매핑 결과를 대분류로 통일).
DETAIL_TO_BIG = {
    "반도체와 반도체장비": "IT",
    "전기제품 (2차전지)": "IT",
    "디스플레이": "IT",
    "기타 IT": "IT",
    "기계": "산업재",
    "건설": "산업재",
    "전기장비": "산업재",
    "복합기업": "산업재",
    "조선": "산업재",
    "상업서비스와 공급품": "산업재",
    "항공 및 해운": "산업재",
    "기타 산업재": "산업재",
}


def _apply_sector_mode(sector_name):
    """현재 모드에 맞게 섹터명 변환. detail 모드면 그대로, big 모드면 세분류 → 대분류."""
    if not sector_name:
        return sector_name
    if _current_sector_mode() == "detail":
        return sector_name
    return DETAIL_TO_BIG.get(sector_name, sector_name)


def _current_sector_mode():
    """현재 요청의 섹터 모드 ('detail' or 'big'). flask.g 또는 기본값."""
    try:
        from flask import g, has_app_context
        if has_app_context():
            return getattr(g, "sector_mode", "detail")
    except Exception:
        pass
    return "detail"


def _gics_to_sector(gics_code):
    """WICS GICS 코드 → 섹터명.

    모드: 'detail' = 산업재/IT 세분류, 'big' = 11개 대분류 그대로.
    flask.g.sector_mode 로 결정 (쿠키 → before_request 에서 설정).
    """
    if not gics_code or not isinstance(gics_code, str) or not gics_code.startswith("G"):
        return None
    detail = _current_sector_mode() == "detail"
    head = gics_code[1:3]
    big = GICS_SECTORS.get(head)
    if not big:
        return None
    if not detail:
        return big  # 11분류만 — 산업재/IT 도 통합
    # 세분류 (산업재/IT 만)
    head6 = gics_code[:7]
    if head6 in WICS_DETAIL:
        return WICS_DETAIL[head6]
    if big == "산업재":
        return "기타 산업재"
    if big == "IT":
        return "기타 IT"
    return big


def _to_float(v, default=None):
    """엑셀 셀 값을 float으로 변환. None/문자열 등 처리."""
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def find_market_excel():
    """data/ 에서 가장 최신 Market_Screening_*.xlsm 파일 경로 반환. 없으면 None.

    정렬 기준: 파일명의 YYMMDD (예: Market_Screening_260626 → 260626) 우선,
    날짜를 추출할 수 없는 파일은 mtime fallback. zip 해제로 mtime이 모두 같아져도
    파일명 날짜로 올바른 최신 파일을 고르도록 한다.
    """
    if not os.path.isdir(DATA_DIR):
        return None
    candidates = []
    import re as _re
    for name in os.listdir(DATA_DIR):
        n = _nfc(name).lower()
        if n.startswith("market_screening") and (n.endswith(".xlsm") or n.endswith(".xlsx")):
            full = os.path.join(DATA_DIR, name)
            m = _re.search(r"(\d{6})", n)
            key = int(m.group(1)) if m else 0
            candidates.append((key, os.path.getmtime(full), full))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]


def _parse_consensus(wb):
    """4.consensus 시트에서 종목별 영업이익 컨센서스 변화율 파싱.
    반환: {종목코드: {fy1_op, fy1_chg, fy2_op, fy2_chg, growth, int1_chg, ...}}
    """
    out = {}
    sn = next((s for s in wb.sheetnames if "consensus" in s.lower()), None)
    if not sn:
        return out
    ws = wb[sn]
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if len(row) < 17:
            continue
        code = row[1]
        if not code or not isinstance(code, str) or not code.startswith("A"):
            continue
        out[code[1:].zfill(6)] = {
            "name": str(row[2]).strip() if row[2] else "",
            "cap_bil": _to_float(row[3]),
            "fy1_op": _to_float(row[4]),
            "fy1_op_prev": _to_float(row[5]),
            "fy1_chg": _to_float(row[6]),       # FY1 영업이익 변화율 (%)
            "fy1_contrib": _to_float(row[7]),   # FY1 기여도 (%)
            "fy2_op": _to_float(row[8]),
            "fy2_op_prev": _to_float(row[9]),
            "fy2_chg": _to_float(row[10]),
            "fy2_contrib": _to_float(row[11]),
            "growth": _to_float(row[12]),       # FY1→FY2 성장률
            "int1_op": _to_float(row[14]),
            "int1_op_prev": _to_float(row[15]),
            "int1_chg": _to_float(row[16]),     # 분기(INT1) 영업이익 변화율
        }
    return out


def _parse_flow(wb):
    """P 시트에서 종목별 수급 데이터(기관/외국인/사모펀드/연기금/금융투자) × 1D/5D/20D/60D 파싱.

    P 시트 R16 헤더 매핑 (0-indexed: row[N-1]):
      col 2 = 종목코드, col 3 = 종목명
      col 27-28: 1D 기관 / 1D 외국인
      col 29-30: 5D 기관 / 5D 외국인
      col 31-32: 20D 기관 / 20D 외국인
      col 33-34: 60D 기관 / 60D 외국인
      col 40-42: 1D 사모펀드 / 1D 연기금 / 1D 금융투자
      col 43-45: 5D 사모펀드 / 5D 연기금 / 5D 금융투자
      col 46-48: 20D 사모펀드 / 20D 연기금 / 20D 금융투자
      col 49-51: 60D 사모펀드 / 60D 연기금 / 60D 금융투자
    단위: 모두 백만원 (Local mn)
    """
    out = {}
    if "P" not in wb.sheetnames:
        return out
    ws = wb["P"]
    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=True):
        if len(row) < 51:
            continue
        code_raw = row[1]  # col 2
        if not code_raw or not isinstance(code_raw, str) or not code_raw.startswith("A"):
            continue
        code = code_raw[1:].zfill(6)
        out[code] = {
            "d1":  {"inst_mil": _to_float(row[26]), "frgn_mil": _to_float(row[27]),
                    "smfp_mil": _to_float(row[39]), "pens_mil": _to_float(row[40]),
                    "fini_mil": _to_float(row[41])},
            "d5":  {"inst_mil": _to_float(row[28]), "frgn_mil": _to_float(row[29]),
                    "smfp_mil": _to_float(row[42]), "pens_mil": _to_float(row[43]),
                    "fini_mil": _to_float(row[44])},
            "d20": {"inst_mil": _to_float(row[30]), "frgn_mil": _to_float(row[31]),
                    "smfp_mil": _to_float(row[45]), "pens_mil": _to_float(row[46]),
                    "fini_mil": _to_float(row[47])},
            "d60": {"inst_mil": _to_float(row[32]), "frgn_mil": _to_float(row[33]),
                    "smfp_mil": _to_float(row[48]), "pens_mil": _to_float(row[49]),
                    "fini_mil": _to_float(row[50])},
        }
    return out


def parse_market_excel(path):
    """Market_Screening 엑셀에서 매크로/종목 데이터 파싱.

    1.macro 시트 — 지수/환율/원자재 시계열 (1D, 5D, MTD, 1M, 3M, YTD, 1YR, 2YR)
    5.market 시트 — 개별 종목 (시총, GICS, 1D, 1Wk, 1M, 3M, 6M, 12M, YTD)
    4.consensus — 영업이익 컨센서스 변화율 (FY1, FY2, INT1)
    2.Sector   — 기관/외국인 순매수 (1D, 5D, 20D, 60D)
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)

    # WICS 소분류 매핑 자동 추출: 2.Sector 시트의 G+6자리 코드 → 한글명
    if "2.Sector" in wb.sheetnames:
        ws_sec = wb["2.Sector"]
        for row in ws_sec.iter_rows(min_row=1, max_row=ws_sec.max_row, values_only=True):
            if len(row) < 4:
                continue
            code, name = row[2], row[3]
            if (isinstance(code, str) and code.startswith("G") and len(code) == 7
                    and isinstance(name, str) and name.strip()):
                _WICS_SMALL[code] = name.strip()

    macro = []
    if "1.macro" in wb.sheetnames:
        ws = wb["1.macro"]
        for row in ws.iter_rows(min_row=13, max_row=50, values_only=True):
            if len(row) < 13:
                continue
            name = row[3]
            if not name or not isinstance(name, str):
                continue
            macro.append({
                "name": name.strip(),
                "close": _to_float(row[4]),
                "d1": _to_float(row[5]),
                "d5": _to_float(row[6]),
                "mtd": _to_float(row[7]),
                "m1": _to_float(row[8]),
                "m3": _to_float(row[9]),
                "ytd": _to_float(row[10]),
                "y1": _to_float(row[11]),
                "y2": _to_float(row[12]),
            })

    stocks = []
    seen_codes = set()  # 같은 종목이 여러 사용자 그룹에 중복 입력된 경우 제거
    if "5.market" in wb.sheetnames:
        ws = wb["5.market"]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if len(row) < 16:
                continue
            code = row[2]
            if not code or not isinstance(code, str) or not code.startswith("A"):
                continue
            gics = row[5]
            if not gics or not isinstance(gics, str) or not gics.startswith("G"):
                continue
            cap = _to_float(row[7], 0) or 0
            if cap <= 0:
                continue
            code_clean = code[1:].zfill(6)
            if code_clean in seen_codes:
                continue
            seen_codes.add(code_clean)
            stocks.append({
                "code": code_clean,   # A005930 → 005930
                "name": str(row[3]).strip() if row[3] else "",
                "gics": gics.strip(),
                "cap_eok": round(cap * 10, 0),  # 십억원 → 억원
                "price": int(round(_to_float(row[8], 0) or 0)),
                "d1": _to_float(row[9]),
                "wk1": _to_float(row[10]),
                "m1": _to_float(row[11]),
                "m3": _to_float(row[12]),
                "m6": _to_float(row[13]),
                "y1": _to_float(row[14]),
                "ytd": _to_float(row[15]),
            })

    # 갱신일자: 파일명의 YYMMDD 우선 (예: Market_Screening_260626 → 2026-06-26).
    # 사용자가 매일 새 파일을 받아오므로 파일명의 날짜가 가장 신뢰할 만하다.
    # 파일명에서 날짜를 못 뽑으면 1.macro B1 셀, 그것도 없으면 파일 mtime fallback.
    as_of = None
    import re as _re
    fname_match = _re.search(r"(\d{2})(\d{2})(\d{2})", os.path.basename(path))
    if fname_match:
        yy, mm, dd = fname_match.groups()
        as_of = f"20{yy}-{mm}-{dd}"
    if not as_of:
        try:
            if "1.macro" in wb.sheetnames:
                ws = wb["1.macro"]
                v = ws["B1"].value
                if hasattr(v, "strftime"):
                    as_of = v.strftime("%Y-%m-%d")
        except Exception:
            pass
    if not as_of:
        import datetime as _dt
        as_of = _dt.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")

    consensus = _parse_consensus(wb)
    flow = _parse_flow(wb)
    stocks_by_code = {s["code"]: s for s in stocks}
    p_gics = _parse_p_gics(wb)  # P 시트의 종목코드 → GICS 코드 (3,875종목)
    p_perf = _parse_p_perf(wb)  # P 시트의 종목코드 → 수익률 (알파 산점도 커버리지 확장)
    kodex200 = _parse_kodex200_index(wb)  # KODEX 200 지수 (알파 산점도 벤치마크)
    screening = _parse_screening_sheet(wb)  # 조건별 종목 스크리닝 보드 (같은 wb 재사용 → 추가 IO 0)

    return {
        "macro": macro,
        "stocks": stocks,
        "stocks_by_code": stocks_by_code,
        "consensus": consensus,
        "flow": flow,
        "p_gics": p_gics,
        "p_perf": p_perf,
        "kodex200": kodex200,
        "screening": screening,
        "as_of": as_of,
        "file": os.path.basename(path),
    }


def _parse_p_perf(wb):
    """P 시트의 종목별 수익률 매핑 반환. 종목코드(6자리) → {wk1, m1, m3}.

    P 시트 컬럼 위치 (row[N-1] 기준):
      col 2 (row[1]) = Ticker 'A005930'
      col 12 (row[11]) = 1D
      col 13 (row[12]) = 5D
      col 14 (row[13]) = 1M
      col 15 (row[14]) = 3M
    """
    if "P" not in wb.sheetnames:
        return {}
    ws = wb["P"]
    out = {}
    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=True):
        if len(row) < 15:
            continue
        code_raw = row[1]
        if not isinstance(code_raw, str) or not code_raw.startswith("A"):
            continue
        code = code_raw[1:].zfill(6)
        wk1 = _to_float(row[12])
        m1  = _to_float(row[13])
        m3  = _to_float(row[14])
        # 셋 다 없으면 스킵
        if wk1 is None and m1 is None and m3 is None:
            continue
        out[code] = {"wk1": wk1, "m1": m1, "m3": m3}
    return out


def _parse_kodex200_index(wb):
    """P 시트에서 KODEX 200 (A069500) 행을 찾아 지수 수익률 반환.

    반환: {"name": "KODEX 200", "wk1": <5D>, "m1": <1M>, "m3": <3M>, "d1": <1D>} 또는 None.
    수익률 컬럼: 종목 행과 동일 컬럼 위치 — col 12(1D), 13(5D), 14(1M), 15(3M).
    """
    if "P" not in wb.sheetnames:
        return None
    ws = wb["P"]
    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=True):
        if len(row) < 15:
            continue
        code_raw = row[1]
        if not isinstance(code_raw, str) or code_raw.strip() != "A069500":
            continue
        return {
            "name": "KODEX 200",
            "d1":  _to_float(row[11]),
            "wk1": _to_float(row[12]),
            "m1":  _to_float(row[13]),
            "m3":  _to_float(row[14]),
        }
    return None


def _parse_p_gics(wb):
    """P 시트에서 종목코드 → WICS 6자리 GICS 코드 매핑 (예: '005930' → 'G453010').

    세분류(산업재/IT)를 위해 6자리 소분류 코드 사용. 6자리 없으면 대분류(G45)로 폴백.
    """
    if "P" not in wb.sheetnames:
        return {}
    ws = wb["P"]
    out = {}
    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=True):
        if len(row) < 6:
            continue
        code_raw = row[1]    # col 2: Code (A005930)
        gics_big = row[4]    # col 5: WICS업종코드(대) — G45
        gics_sub = row[5]    # col 6: WICS업종코드(소) — G453010
        if not code_raw or not isinstance(code_raw, str) or not code_raw.startswith("A"):
            continue
        # 소분류 우선, 없으면 대분류
        chosen = gics_sub if (isinstance(gics_sub, str) and gics_sub.startswith("G")) else gics_big
        if not isinstance(chosen, str) or not chosen.startswith("G"):
            continue
        out[code_raw[1:].zfill(6)] = chosen.strip()
    return out


def _classify_macro(macro_rows):
    """이름 기준으로 매크로 행을 카테고리별로 분류."""
    name_to_cat = {}
    for cat, names in MACRO_CATEGORIES.items():
        for n in names:
            name_to_cat[n] = cat
    out = {cat: [] for cat in MACRO_CATEGORIES}
    for r in macro_rows:
        cat = name_to_cat.get(r["name"])
        if cat:
            out[cat].append(r)
    # 카테고리 내 순서는 MACRO_CATEGORIES 정의 순서 유지
    for cat, names in MACRO_CATEGORIES.items():
        order = {n: i for i, n in enumerate(names)}
        out[cat].sort(key=lambda r: order.get(r["name"], 999))
    return out


def load_screening_board():
    """조건별 종목 스크리닝 보드 — load_market_data 캐시에서 즉시 반환 (추가 IO 없음)."""
    m = load_market_data()
    return (m or {}).get("screening")


def _parse_screening_sheet(wb):
    """이미 열린 워크북의 'screening' 시트 파싱 → 스크리닝 보드 dict.

    row2=기준시각, row4/5=KOSPI/KOSDAQ 요약(일간수익률/거래대금/기관/외국인),
    row8=카테고리 헤더(col2~15), row9~=카테고리별 종목명.
    반환: {"as_of","summary":[...],"groups":[...]} 또는 None.
    """
    if "screening" not in wb.sheetnames:
        return None
    ws = wb["screening"]

    # read_only 워크북은 랜덤 셀 접근이 느리므로 iter_rows 로 한 번만 순회해 rows 리스트化.
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300), values_only=True):
        rows.append(row)

    def _g(r1, c1):  # 1-indexed → rows[r-1][c-1]
        ri, ci = r1 - 1, c1 - 1
        if 0 <= ri < len(rows) and rows[ri] and ci < len(rows[ri]):
            return rows[ri][ci]
        return None

    # 기준시각
    as_of = ""
    t = _g(2, 3)
    if hasattr(t, "strftime"):
        as_of = t.strftime("%Y-%m-%d %H:%M")
    elif t:
        as_of = str(t)

    # KOSPI/KOSDAQ 요약 (row4,5: col2 이름, col3 수익률, col4 거래대금(십억), col5 기관, col6 외국인 백만원)
    summary = []
    for r in (4, 5):
        nm = _g(r, 2)
        if not nm:
            continue
        summary.append({
            "name": str(nm),
            "ret": _to_float(_g(r, 3)),
            "val_bil": _to_float(_g(r, 4)),
            "inst_eok": (_to_float(_g(r, 5)) or 0) / 100.0,   # 백만원→억
            "frgn_eok": (_to_float(_g(r, 6)) or 0) / 100.0,
        })

    # 카테고리 (col2~15) → 그룹 3덩어리로 묶어 표시
    #  cls: 색상 힌트 (up=상승/신고/매수 계열, dn=하락/신저/매도 계열, neu=중립)
    CATS = [
        (2,  "up"),  (3,  "dn"),  (4,  "up"),  (5,  "dn"),  (6,  "neu"),
        (7,  "up"),  (8,  "up"),  (9,  "up"),  (10, "dn"),  (11, "dn"),  (12, "dn"),
        (13, "up"),  (14, "dn"),  (15, "neu"),
    ]
    cat_by_col = {}
    for c, cls in CATS:
        head = _g(8, c)
        head = str(head).replace("\n", " ").strip() if head else f"col{c}"
        stocks = []
        for ri in range(8, len(rows)):   # row9~ (0-indexed 8~)
            v = rows[ri][c - 1] if rows[ri] and (c - 1) < len(rows[ri]) else None
            if v is not None and str(v).strip():
                stocks.append(str(v).strip())
        cat_by_col[c] = {"name": head, "cls": cls, "stocks": stocks}

    groups = [
        {"title": "시총 5천억 이상 · 가격/거래", "cols": [2, 3, 4, 5, 6]},
        {"title": "투신·사모·외국인 수급 (시총대비)", "cols": [7, 8, 9, 10, 11, 12]},
        {"title": "밴드 · 무관심주", "cols": [13, 14, 15]},
    ]
    out_groups = []
    for g in groups:
        cats = [cat_by_col[c] for c in g["cols"] if c in cat_by_col]
        out_groups.append({"title": g["title"], "cats": cats})

    return {"as_of": as_of, "summary": summary, "groups": out_groups}


def build_sector_flow(m):
    """P 시트 수급을 WICS 대분류 섹터별로 집계 → 업종별 투자자 순매수 바차트용.

    반환: [{sector, cap_eok, d1:{inst,frgn,smfp,pens,fini}, d5, d20, d60,
            stocks:[{code,name,cap_eok, d1:{...},...}]}]  (금액 단위: 억원)
    투자자: inst=기관, frgn=외국인, smfp=사모펀드, pens=연기금, fini=금융투자
    """
    flow = m.get("flow") or {}
    sbc = m.get("stocks_by_code") or {}
    pg = m.get("p_gics") or {}
    invs = ["inst", "frgn", "smfp", "pens", "fini"]
    kmap = {"inst": "inst_mil", "frgn": "frgn_mil", "smfp": "smfp_mil",
            "pens": "pens_mil", "fini": "fini_mil"}
    periods = ["d1", "d5", "d20", "d60"]

    def _blank():
        return {p: {iv: 0.0 for iv in invs} for p in periods}

    by_sec = {}
    for code, f in flow.items():
        s = sbc.get(code)
        gics = (s.get("gics") if s else None) or pg.get(code)
        sec = _wics_classify(gics, "big") if gics else None
        if not sec:
            continue
        name = s.get("name") if s else code
        cap = (s.get("cap_eok") if s else 0) or 0
        e = by_sec.get(sec)
        if e is None:
            e = {"sector": sec, "cap_eok": 0.0, "stocks": [], **_blank()}
            by_sec[sec] = e
        e["cap_eok"] += cap
        st = {"code": code, "name": name, "cap_eok": round(cap, 0)}
        for p in periods:
            pf = f.get(p, {})
            st[p] = {}
            for iv in invs:
                eok = (pf.get(kmap[iv]) or 0) / 100.0   # 백만원 → 억원
                e[p][iv] += eok
                st[p][iv] = round(eok, 1)
        e["stocks"].append(st)

    out = []
    for sec, e in by_sec.items():
        for p in periods:
            for iv in invs:
                e[p][iv] = round(e[p][iv], 1)
        e["cap_eok"] = round(e["cap_eok"], 0)
        e["stocks"].sort(key=lambda x: -(x["cap_eok"] or 0))
        e["stocks"] = e["stocks"][:15]   # 주요종목 상위 15 (시총순)
        out.append(e)
    out.sort(key=lambda x: -x["cap_eok"])
    return out


def _group_sectors(stock_rows, level="big"):
    """종목들을 WICS 분류로 그룹핑 → 시총 가중 평균 변동률 + 상위 10 종목.

      level == 'big'   → WICS 대분류 (10개)
      level == 'mid'   → WICS 중분류 (~28개)
      level == 'small' → WICS 소분류 (~50개)

    각 sector 에 'code'(이 레벨 코드, 예 'G45'·'G4530'·'G453010') 와
    'parent_big' (대분류 코드, 예 'G45') 가 함께 들어간다. 클라이언트에서
    대분류 막대 클릭 → 자식 필터링에 사용.
    """
    by_sec = {}
    code_by_sec = {}   # 그룹명 → (이 레벨 코드, 대분류 코드)
    cut = {"big": 3, "mid": 5, "small": 7}.get(level, 3)
    for s in stock_rows:
        gics = s.get("gics", "") or ""
        sec_name = _wics_classify(gics, level) or "기타"
        by_sec.setdefault(sec_name, []).append(s)
        # 이 그룹의 대표 코드 = 첫 종목의 prefix
        if sec_name not in code_by_sec and isinstance(gics, str) and gics.startswith("G"):
            code_by_sec[sec_name] = (gics[:cut], gics[:3])

    sectors = []
    for sec, stocks in by_sec.items():
        if sec == "기타":
            continue
        total_cap = sum(s["cap_eok"] for s in stocks)
        if total_cap <= 0:
            continue

        def wavg(field):
            return sum((s["cap_eok"] * (s.get(field) or 0)) for s in stocks) / total_cap

        top10 = sorted(stocks, key=lambda x: -x["cap_eok"])[:10]
        code, parent_big = code_by_sec.get(sec, ("", ""))
        sectors.append({
            "name": sec,
            "code": code,
            "parent_big": parent_big,
            "cap_eok": round(total_cap, 0),
            "d1": round(wavg("d1"), 2),
            "wk1": round(wavg("wk1"), 2),
            "m1": round(wavg("m1"), 2),
            "m3": round(wavg("m3"), 2),
            "ytd": round(wavg("ytd"), 2),
            "stocks": top10,
            "stock_count": len(stocks),
        })
    sectors.sort(key=lambda x: -x["cap_eok"])
    return sectors


def _fetch_ticker_perf(ticker, name):
    """1개 ticker의 최근 종가 + 전일 대비 등락률 + 등락폭."""
    try:
        import FinanceDataReader as fdr
        import datetime as _dt
        start = (_dt.date.today() - _dt.timedelta(days=20)).strftime("%Y-%m-%d")
        df = fdr.DataReader(ticker, start)
        if df is None or len(df) < 2:
            return None
        last, prev = df.iloc[-1], df.iloc[-2]
        close = float(last["Close"])
        prev_close = float(prev["Close"])
        change = close - prev_close
        change_pct = change / prev_close * 100 if prev_close else 0.0
        return {
            "ticker": ticker, "name": name,
            "close": close,
            "change": change,
            "change_pct": round(change_pct, 2),
            "date": df.index[-1].strftime("%Y-%m-%d"),
        }
    except Exception as e:
        print(f"[market] {ticker} 실패: {e}")
        return None


def _kospi_sector_perf():
    """KOSPI 22업종 등락률(시총 가중) + 각 섹터 시총 상위 10종목.

    KRX listing 한 번 호출로 모든 종목의 등락률·시총을 받고,
    data/코스피섹터.xlsx 의 종목코드 → 업종 매핑으로 그룹핑한다.
    """
    try:
        import FinanceDataReader as fdr
        listing = fdr.StockListing("KRX")
    except Exception as e:
        print(f"[market] KRX listing 실패: {e}")
        return []
    listing = listing[listing["Market"] == "KOSPI"].copy()
    listing["Code"] = listing["Code"].astype(str).str.zfill(6)
    listing["Marcap"] = pd.to_numeric(listing["Marcap"], errors="coerce").fillna(0)
    listing["ChagesRatio"] = pd.to_numeric(listing["ChagesRatio"], errors="coerce").fillna(0)
    listing["Close"] = pd.to_numeric(listing["Close"], errors="coerce").fillna(0)

    if not os.path.isdir(DATA_DIR):
        return []
    norm_map = {_nfc(n): n for n in os.listdir(DATA_DIR)}
    key = _nfc(KOSPI_SECTOR_FILE)
    if key not in norm_map:
        return []
    try:
        sec_df = pd.read_excel(os.path.join(DATA_DIR, norm_map[key]))
    except Exception as e:
        print(f"[market] 코스피섹터.xlsx 읽기 실패: {e}")
        return []
    code_col = next((c for c in sec_df.columns if "종목코드" in str(c) or "code" in str(c).lower()), None)
    sec_col = next((c for c in sec_df.columns if "업종" in str(c) or "섹터" in str(c)), None)
    if not (code_col and sec_col):
        return []
    sec_df[code_col] = sec_df[code_col].astype(str).str.zfill(6)
    sec_map = dict(zip(sec_df[code_col], sec_df[sec_col]))

    listing["Sector"] = listing["Code"].map(sec_map).fillna("기타")
    listing["SectorNorm"] = listing["Sector"].apply(_normalize_sector)

    sectors = []
    for sec, g in listing.groupby("SectorNorm"):
        if sec == "기타":
            continue
        total_cap = float(g["Marcap"].sum())
        if total_cap <= 0:
            continue
        weighted = float((g["Marcap"] * g["ChagesRatio"]).sum() / total_cap)
        top10 = g.nlargest(10, "Marcap")
        stocks = [{
            "code": r["Code"], "name": r["Name"],
            "close": int(round(float(r["Close"]))),
            "change_pct": round(float(r["ChagesRatio"]), 2),
            "cap_eok": round(float(r["Marcap"]) / 1e8, 0),
        } for _, r in top10.iterrows()]
        sectors.append({
            "name": sec,
            "change_pct": round(weighted, 2),
            "cap_eok": round(total_cap / 1e8, 0),
            "stocks": stocks,
        })
    sectors.sort(key=lambda x: -x["cap_eok"])
    return sectors


def load_market_data(force_refresh=False):
    """시장 데이터 로드. data/Market_Screening_*.xlsm 우선, 없으면 None 반환.

    엑셀 파일 mtime 기반 캐시 — 새 파일 올리면 다음 호출 시 자동 갱신.
    """
    path = find_market_excel()
    if not path:
        # 엑셀 없으면 빈 데이터 (UI 가 안내 메시지 표시)
        return None

    import time as _time
    mtime = os.path.getmtime(path)
    cache_key = (path, mtime)
    if (not force_refresh and _MARKET_CACHE["data"]
            and _MARKET_CACHE.get("key") == cache_key):
        return _MARKET_CACHE["data"]

    try:
        parsed = parse_market_excel(path)
    except Exception as e:
        print(f"[market] 엑셀 파싱 실패: {e}")
        return None

    macro_by_cat = _classify_macro(parsed["macro"])

    # sectors 는 모드(detail/big)에 따라 달라지므로 캐시 데이터에 raw stocks 만 두고
    # 호출자(render_market) 가 매 요청마다 _group_sectors 다시 호출하도록 한다.
    data = {
        "as_of": parsed["as_of"],
        "file": parsed["file"],
        "fetched_at": _dt_now_str(),
        "dom_idx": macro_by_cat["dom_idx"],
        "global_idx": macro_by_cat["global_idx"],
        "fx": macro_by_cat["fx"],
        "comm": macro_by_cat["comm"],
        "etc": macro_by_cat["etc"],
        "raw_stocks": parsed["stocks"],
        "stock_count": len(parsed["stocks"]),
        "stocks_by_code": parsed["stocks_by_code"],
        "consensus": parsed["consensus"],
        "flow": parsed["flow"],
        "p_gics": parsed.get("p_gics", {}),
        "p_perf": parsed.get("p_perf", {}),
        "kodex200": parsed.get("kodex200"),
        "screening": parsed.get("screening"),
    }

    _MARKET_CACHE["key"] = cache_key
    _MARKET_CACHE["ts"] = _time.time()
    _MARKET_CACHE["data"] = data
    return data


def _dt_now_str():
    import datetime as _dt
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M")


# KOSPI 규모별 비중: FDR로 시총 상위 100/200/소형 분류해 캐시. 24시간 캐시.
_KOSPI_SIZE_CACHE = {"ts": 0, "data": None}


def load_kospi_size():
    """KOSPI 종목을 시총 상위 100(대형) / 101~300(중형) / 그 외(소형) 으로 분류해 시총 비중 반환.

    KRX 공식 분류 기준. data/_kospi_size_cache.json 으로 24시간 캐시한다.
    실패 시 합리적 기본값 사용.
    """
    import time as _time
    now = _time.time()
    if _KOSPI_SIZE_CACHE["data"] is not None and (now - _KOSPI_SIZE_CACHE["ts"]) < 86400:
        return _KOSPI_SIZE_CACHE["data"]

    cache_path = os.path.join(DATA_DIR, "_kospi_size_cache.json") if os.path.isdir(DATA_DIR) else None
    if cache_path and os.path.isfile(cache_path):
        try:
            if (now - os.path.getmtime(cache_path)) < 86400:
                with open(cache_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                _KOSPI_SIZE_CACHE["ts"] = now
                _KOSPI_SIZE_CACHE["data"] = data
                return data
        except Exception:
            pass

    try:
        import FinanceDataReader as fdr
        df = fdr.StockListing("KRX")
        kospi = df[df["Market"] == "KOSPI"].copy()
        kospi = kospi.dropna(subset=["Marcap"])
        kospi = kospi.sort_values("Marcap", ascending=False).reset_index(drop=True)
        large = float(kospi.iloc[:100]["Marcap"].sum())
        mid = float(kospi.iloc[100:300]["Marcap"].sum())
        small = float(kospi.iloc[300:]["Marcap"].sum())
        total = large + mid + small
        if total <= 0:
            raise RuntimeError("KOSPI 시총 합계 0")
        data = {
            "labels": ["대형주", "중형주", "소형주"],
            "values": [round(large / total * 100, 2),
                       round(mid / total * 100, 2),
                       round(small / total * 100, 2)],
        }
    except Exception as e:
        print(f"[kospi-size] FDR 실패, 기본값 사용: {e}")
        data = {"labels": ["대형주", "중형주", "소형주"], "values": [80.0, 15.0, 5.0]}

    _KOSPI_SIZE_CACHE["ts"] = now
    _KOSPI_SIZE_CACHE["data"] = data
    if cache_path:
        try:
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return data


# 전사 규모별 비중 캐시 (펀드 페이지 진입 시마다 5개 엑셀 IO를 피하기 위함).
_FIRM_SIZE_CACHE = {"key": None, "data": None}


def get_firm_size(funds):
    """전체 펀드의 종목을 규모구분별로 합산한 비중. holdings 파일 mtime 기반 캐싱."""
    if not funds:
        return {"labels": [], "values": []}
    try:
        key = tuple(sorted((fn, os.path.getmtime(funds[fn]["holdings"])) for fn in funds))
    except OSError:
        key = None
    if key and _FIRM_SIZE_CACHE["key"] == key and _FIRM_SIZE_CACHE["data"] is not None:
        return _FIRM_SIZE_CACHE["data"]

    size_order = ["대형주", "중형주", "소형주", "미분류"]
    totals = {}
    grand = 0.0
    for fn in funds:
        h = read_holdings(funds[fn]["holdings"])
        for k, v in h.groupby("규모구분")["평가액"].sum().items():
            totals[k] = totals.get(k, 0.0) + float(v)
            grand += float(v)
    data = {
        "labels": [s for s in size_order if s in totals],
        "values": [round(totals[s] / grand * 100, 2)
                   for s in size_order if s in totals],
    } if grand else {"labels": [], "values": []}
    _FIRM_SIZE_CACHE["key"] = key
    _FIRM_SIZE_CACHE["data"] = data
    return data


def match_risk(fund_name, risk_rows):
    """펀드 페이지 이름(파일 접두어)을 리스크 행과 매칭한다.

    1) 코드 완전일치  2) 이름/코드에 접두어가 들어있는지(대소문자·공백 무시)
    """
    if not risk_rows:
        return None
    fn = _norm(fund_name)
    for r in risk_rows:                       # 코드 완전일치 우선
        if fund_name.strip() == r["code"]:
            return r
    for r in risk_rows:                       # 부분일치
        if fn and (fn in _norm(r["name"]) or fn == _norm(r["code"])):
            return r
    return None


def build_risk(risk_row):
    """리스크 행을 화면용 그룹/타일 구조로 가공한다."""
    if not risk_row:
        return None

    def n(v, dec=2):
        try:
            return f"{float(v):,.{dec}f}"
        except (TypeError, ValueError):
            return "—"

    def eok(v):                                # 원 -> 억원
        try:
            return f"{float(v) / 1e8:,.1f}억"
        except (TypeError, ValueError):
            return "—"

    r = risk_row
    groups = [
        ("위험조정 성과", [
            ("Sharpe", n(r["sharpe"])),
            ("Jensen (α)", n(r["jensen"])),
            ("Treynor", n(r["treynor"])),
            ("IR", n(r["ir"])),
            ("RAROC", n(r["raroc"])),
        ]),
        ("시장 민감도", [
            ("베타 (BM대비)", n(r["beta_bm"])),
            ("R² (BM)", n(r["r2_bm"])),
            ("베타 (KOSPI대비)", n(r["beta_kospi"])),
            ("R² (KOSPI)", n(r["r2_kospi"])),
            ("가중평균 베타", n(r["wavg_beta"], 4)),
            ("지수가중 베타", n(r["idx_beta"], 4)),
        ]),
        ("변동성 · 추적오차", [
            ("표준편차 (펀드/BM)", f'{n(r["sd_fund"])} / {n(r["sd_bm"])}'),
            ("분산 (펀드/BM)", f'{n(r["var_fund"])} / {n(r["var_bm"])}'),
            ("TE (추적오차)", n(r["te"])),
            ("펀드 MD", n(r["md"])),
        ]),
        ("VaR", [
            ("델타노멀 VaR (95%)", eok(r["delta_var"])),
        ]),
    ]
    return {"name": r["name"], "code": r["code"], "groups": groups}


def _short_name(nm):
    """좌표 위에 표시할 짧은 펀드 이름."""
    base = str(nm).split("(")[0]
    for w in ["일반사모투자신탁", "일반사모", "투자신탁", "제1호", "신탁", "안다"]:
        base = base.replace(w, "")
    return base.strip() or str(nm)


def build_scatter(risk_rows, active_code=None):
    """5개 펀드를 좌표(산점도)에 찍기 위한 점 목록을 만든다."""
    def fv(x):
        try:
            return round(float(x), 4)
        except (TypeError, ValueError):
            return None

    pts = []
    for r in risk_rows:
        pts.append({
            "label": _short_name(r["name"]),
            "name": r["name"], "code": r["code"],
            "sd": fv(r["sd_fund"]), "sharpe": fv(r["sharpe"]),
            "beta": fv(r["beta_bm"]), "jensen": fv(r["jensen"]),
            "te": fv(r["te"]), "ir": fv(r["ir"]),
            "active": bool(active_code is not None and r["code"] == active_code),
        })
    return pts


# ---------------------------------------------------------------------------
# 2) 종목 읽기 / 한 펀드 가공 / 전사 집계
# ---------------------------------------------------------------------------
def read_holdings(path):
    """종목 엑셀을 읽어 정리한다. 구/신(AITAS) 두 포맷 모두 처리.

    GICS 대분류 매핑(`섹터` 컬럼) 추가 — Market_Screening 의 종목 GICS 코드 우선,
    없으면 펀드 엑셀 원본 업종명 fallback.

    포맷별 컬럼 대응 (canonical: 평가액 / 평가손익률 / 순자산비 / 편입비):
      구 포맷 — 평가액, 평가손익률, 순자산비(NAV<100%), 편입비(주식100%), 규모구분 有
      신 포맷 — 평가금액, 평가R(%), 편입비(NAV<100%), 구성비(주식100%), 규모구분 無
    """
    h = pd.read_excel(path)

    # 컬럼명 정리 — 정렬 화살표(▼▲▽△), 개행, 앞뒤 공백 제거 (AITAS 내보내기 잔재)
    h.columns = [str(c).replace("▼", "").replace("▲", "").replace("▽", "")
                 .replace("△", "").replace("\n", " ").strip() for c in h.columns]

    # 신 포맷(AITAS) 감지 — '구성비' 또는 '평가금액' 컬럼 존재.
    is_new = ("구성비" in h.columns) or ("평가금액" in h.columns)
    if is_new:
        rename = {}
        if "평가금액" in h.columns:  rename["평가금액"] = "평가액"
        if "평가R(%)" in h.columns:  rename["평가R(%)"] = "평가손익률"
        # 신 포맷의 '편입비'가 NAV 대비(순자산비 상당), '구성비'가 주식 100% 기준
        if "편입비" in h.columns:    rename["편입비"] = "순자산비"
        if "구성비" in h.columns:    rename["구성비"] = "편입비"
        h = h.rename(columns=rename)

    h = h.dropna(subset=["종목명"]).copy()
    if "규모구분" in h.columns:
        h["규모구분"] = h["규모구분"].fillna("미분류").replace("", "미분류")
    else:
        h["규모구분"] = "미분류"   # 신 포맷: 규모구분 컬럼 없음 → 아래서 시총 기반 추정
    h["업종명"] = h["업종명"].fillna("기타").replace("", "기타")
    for c in ["편입비", "순자산비", "시장비", "평가액", "평가손익률", "종가", "보유수량"]:
        if c in h.columns:
            h[c] = pd.to_numeric(h[c], errors="coerce").fillna(0)
    # 순자산비(NAV 대비 비중) 컬럼이 없으면 편입비로 fallback
    if "순자산비" not in h.columns:
        h["순자산비"] = h["편입비"] if "편입비" in h.columns else 0.0
    if "종목" in h.columns:
        h["종목코드"] = h["종목"].map(_fmt_code)
    else:
        h["종목코드"] = ""

    # 규모구분 없으면(신 포맷) Market_Screening 시총(억)으로 대형/중형/소형 추정
    if (h["규모구분"] == "미분류").all():
        m = load_market_data()
        capmap = {}
        if m:
            for cd, s in (m.get("stocks_by_code") or {}).items():
                capmap[cd] = s.get("cap_eok") or 0
        def _size(code):
            cap = capmap.get(code, 0)      # 억원
            if cap >= 50000:  return "대형주"   # ≥ 5조
            if cap >= 10000:  return "중형주"   # 1~5조
            if cap > 0:       return "소형주"
            return "미분류"
        h["규모구분"] = h["종목코드"].map(_size)

    # 종목코드 → GICS 대분류 (3단 fallback)
    #   1) Market_Screening 엑셀의 종목별 GICS (P 시트 + 5.market)
    #   2) 펀드 엑셀 업종명을 KRX→GICS 매핑으로 변환
    #   3) 그래도 매칭 안 되면 원본 업종명 그대로 (코스닥종합/기타 등)
    gmap = get_stock_gics_map()
    def _resolve(row):
        g = gmap.get(row["종목코드"])
        if g:
            return g
        nm = _normalize_sector(row["업종명"])
        if nm in _KRX_TO_GICS:
            return _apply_sector_mode(_KRX_TO_GICS[nm])
        return row["업종명"]
    h["섹터"] = h.apply(_resolve, axis=1)
    return h


def _fmt_code(x):
    """종목코드를 6자리 문자열로 정리 (예: 5930.0 -> 005930)."""
    try:
        return str(int(float(x))).zfill(6)
    except (TypeError, ValueError):
        s = str(x).strip()
        return s[:-2] if s.endswith(".0") else s


def _wavg_pnl(d):
    """평가액 가중평균 평가손익률."""
    v = d["평가액"].sum()
    return float((d["평가액"] * d["평가손익률"]).sum() / v) if v else 0.0


def process_fund(name, ret_path, hold_path, risk_row=None):
    # ----- 수익률 -----
    ret = pd.read_excel(ret_path)
    ret = ret.dropna(subset=["일자"]).copy()
    ret["일자"] = pd.to_datetime(ret["일자"])
    ret = ret.sort_values("일자")

    dates = ret["일자"].dt.strftime("%Y-%m-%d").tolist()
    fund_ret = ret["펀드수익률"].round(2).tolist()
    bm_ret = ret["BM수익률"].round(2).tolist()
    excess = ret["초과수익률"].round(2).tolist()

    as_of = ret["일자"].iloc[-1].strftime("%Y-%m-%d")
    stat_fund = round(float(ret["펀드수익률"].iloc[-1]), 2)
    stat_bm = round(float(ret["BM수익률"].iloc[-1]), 2)
    stat_excess = round(float(ret["초과수익률"].iloc[-1]), 2)

    # ----- 종목 -----
    hold = read_holdings(hold_path)
    hold = hold.sort_values("편입비", ascending=False)

    rows = []
    for _, r in hold.iterrows():
        rows.append({
            "종목명": r.get("종목명", ""),
            "종목코드": str(r.get("종목코드", "")),
            "편입비": round(float(r.get("편입비", 0) or 0), 2),
            "평가액억": round(float(r.get("평가액", 0) or 0) / 1e8, 1),
            "손익률": round(float(r.get("평가손익률", 0) or 0), 2),
            "종가": int(round(float(r.get("종가", 0) or 0))),
            "보유수량": int(round(float(r.get("보유수량", 0) or 0))),
            "업종명": r.get("업종명", ""),
            "규모구분": r.get("규모구분", ""),
        })

    # 펀드 섹터 비중 (순자산비 기준 = NAV 대비, GICS 대분류 — KOSPI 비교용 / 기존 레거시)
    sec_nsb = hold.groupby("섹터")["순자산비"].sum()
    fund_sectors = [{"name": nm, "pct": round(float(sec_nsb[nm]), 2)}
                    for nm in sec_nsb.index]
    kospi_cmp = kospi_compare(fund_sectors, load_kospi_sectors())

    # WICS 대/중/소 3종 비교 데이터 (새 토글용)
    gics_code_map = get_stock_gics_code_map()
    kospi_cmp_wics = {}
    for lv in ("big", "mid", "small"):
        view_secs = view_sectors_by_level(hold, lv, gics_code_map)
        kospi_cmp_wics[lv] = kospi_compare(view_secs, kospi_sectors_by_level(lv))

    # 알파 산점도 — 종목별 (5D×1M / 1M×3M) vs KOSPI 대형주 지수
    perf_map = get_stock_perf_map()
    bm = get_market_benchmark()
    alpha_codes = [(r.get("종목코드"), r.get("종목명"), r.get("편입비", 0))
                   for _, r in hold.iterrows()]
    alpha = build_alpha_scatters(alpha_codes, perf_map, bm)

    # 규모별 비중 (도넛 — 해당 펀드)
    size_order = ["대형주", "중형주", "소형주", "미분류"]
    siz = hold.groupby("규모구분")["편입비"].sum().round(2)
    siz = siz.reindex([s for s in size_order if s in siz.index])
    size = {"labels": siz.index.tolist(), "values": siz.values.tolist()}

    total_eval = round(float(hold["평가액"].sum()) / 1e8, 1)
    stock_count = len(hold)
    equity_ratio = round(float(hold["편입비"].sum()), 2)

    risk = build_risk(risk_row)
    official_name = risk["name"] if risk else None

    return {
        "name": name,
        "official_name": official_name,
        "as_of": as_of,
        "stat_fund": stat_fund,
        "stat_bm": stat_bm,
        "stat_excess": stat_excess,
        "total_eval": total_eval,
        "stock_count": stock_count,
        "equity_ratio": equity_ratio,
        "chart": {"dates": dates, "fund": fund_ret, "bm": bm_ret, "excess": excess},
        "kospi_cmp": kospi_cmp,
        "kospi_cmp_wics": kospi_cmp_wics,
        "alpha": alpha,
        "size": size,
        "rows": rows,
        "risk": risk,
    }


def _compute_firm_view(allh, view_funds):
    """전사 합산 집계를 종목/섹터/매트릭스 형태로 반환.

    allh: 합쳐진 종목 DF (_fund 컬럼 포함). view_funds: 이 뷰에 포함될 펀드 이름.
    종목·섹터 매트릭스의 펀드 컬럼 순서는 view_funds 그대로 사용한다.
    """
    if len(allh) == 0:
        return {"firm_total": 0.0, "stock_count": 0, "sector_count": 0,
                "firm_stocks": [], "firm_sectors": [],
                "matrix": {"funds": view_funds, "stock_rows": [], "sector_rows": []}}

    firm_total = float(allh["평가액"].sum())

    # 전사 순자산비(NAV 대비) 비중 = 각 펀드 순자산비를 펀드 NAV 비율로 가중합.
    #   NAV_fund = Σ평가액 / (Σ순자산비/100) → 펀드별 NAV 비율 fund_w 계산.
    fund_w = {}
    _navs = {}
    for fn, g in allh.groupby("_fund"):
        ev = float(g["평가액"].sum())
        nsb = float(g["순자산비"].sum())
        _navs[fn] = ev / (nsb / 100.0) if nsb > 0 else ev
    _navtot = sum(_navs.values())
    if _navtot > 0:
        fund_w = {fn: v / _navtot for fn, v in _navs.items()}
    # 종목코드/섹터별 전사 순자산비 (가중합)
    allh = allh.copy()
    allh["_nsb_firm"] = allh.apply(
        lambda r: float(r.get("순자산비") or 0) * fund_w.get(r.get("_fund"), 0.0), axis=1)

    pnl_stock = allh.groupby("종목코드").apply(_wavg_pnl)
    name_map = allh.groupby("종목코드")["종목명"].first().to_dict()
    sec_map = allh.groupby("종목코드")["섹터"].first().to_dict()
    val_stock = allh.groupby("종목코드")["평가액"].sum()
    wt_stock = allh.groupby("종목코드")["_nsb_firm"].sum()
    close_map = allh.groupby("종목코드")["종가"].max().to_dict()

    firm_stocks = []
    for code in val_stock.index:
        v = float(val_stock[code])
        firm_stocks.append({
            "code": code, "name": name_map.get(code, code),
            "sector": sec_map.get(code, "기타"),
            "value": round(v / 1e8, 1), "eok": round(v / 1e8, 1),
            "pct": round(float(wt_stock.get(code, 0.0)), 2),
            "pnl": round(float(pnl_stock[code]), 2)})
    firm_stocks.sort(key=lambda x: -x["value"])
    weight_map = {s["code"]: s["pct"] for s in firm_stocks}

    pnl_sec = allh.groupby("섹터").apply(_wavg_pnl)
    val_sec = allh.groupby("섹터")["평가액"].sum()
    wt_sec = allh.groupby("섹터")["_nsb_firm"].sum()
    firm_sectors = []
    for nm in val_sec.index:
        v = float(val_sec[nm])
        firm_sectors.append({
            "name": nm, "value": round(v / 1e8, 1), "eok": round(v / 1e8, 1),
            "pct": round(float(wt_sec.get(nm, 0.0)), 2),
            "pnl": round(float(pnl_sec[nm]), 2)})
    firm_sectors.sort(key=lambda x: -x["value"])

    # 종목 × 펀드 매트릭스 (순자산비 = NAV 대비 비중)
    stock_pivot = allh.pivot_table(index="종목코드", columns="_fund",
                                   values="순자산비", aggfunc="sum", fill_value=0)
    stock_rows = []
    for s in firm_stocks:
        code = s["code"]
        weights = {}
        for fn in view_funds:
            w = float(stock_pivot.loc[code, fn]) if (code in stock_pivot.index and fn in stock_pivot.columns) else 0.0
            weights[fn] = round(w, 2)
        stock_rows.append({
            "code": code, "name": s["name"],
            "close": int(round(float(close_map.get(code, 0)))),
            "weights": weights, "firm": weight_map.get(code, 0.0)})

    # 섹터 × 펀드 매트릭스 (GICS 대분류 기준, 각 펀드에서 해당 섹터 종목 순자산비 합산)
    sector_pivot = allh.pivot_table(index="섹터", columns="_fund",
                                    values="순자산비", aggfunc="sum", fill_value=0)
    # 종목 → stock_row dict (섹터 드릴다운용 빠른 조회)
    stock_row_by_code = {r["code"]: r for r in stock_rows}
    stocks_by_sector = {}
    for fs in firm_stocks:
        sr = stock_row_by_code.get(fs["code"])
        if sr:
            stocks_by_sector.setdefault(fs["sector"], []).append(sr)

    sector_rows = []
    for s in firm_sectors:
        nm = s["name"]
        weights = {}
        for fn in view_funds:
            w = float(sector_pivot.loc[nm, fn]) if (nm in sector_pivot.index and fn in sector_pivot.columns) else 0.0
            weights[fn] = round(w, 2)
        # 그 섹터에 속한 종목들 (전사비중 내림차순)
        children = sorted(stocks_by_sector.get(nm, []), key=lambda r: -r["firm"])
        sector_rows.append({
            "name": nm, "weights": weights, "firm": s["pct"],
            "stocks": children})

    return {
        "firm_total": round(firm_total / 1e8, 1),
        "stock_count": len(firm_stocks),
        "sector_count": len(firm_sectors),
        "firm_stocks": firm_stocks,
        "firm_sectors": firm_sectors,
        "matrix": {"funds": view_funds, "stock_rows": stock_rows, "sector_rows": sector_rows},
    }


def process_firm(funds):
    """전사 현황: 전체(all) + TMF 제외(ex_tmf) 두 가지 뷰를 함께 반환.

    TMF 펀드는 이름에 'tmf'가 포함된 펀드로 자동 판별한다 ([[_is_tmf]]).
    KOSPI 섹터 비중과의 비교 데이터도 함께 만들어 전달한다 ([[load_kospi_sectors]]).
    """
    fund_names = list(funds.keys())
    tmf_funds = [fn for fn in fund_names if _is_tmf(fn)]
    ex_tmf_funds = [fn for fn in fund_names if not _is_tmf(fn)]

    frames = []
    for fn in fund_names:
        h = read_holdings(funds[fn]["holdings"])
        h["_fund"] = fn
        frames.append(h)
    allh = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # 기준일 (펀드 수익률 파일들 중 최신 날짜)
    as_of = ""
    for fn in fund_names:
        try:
            r = pd.read_excel(funds[fn]["returns"], usecols=["일자"])
            d = pd.to_datetime(r["일자"]).dropna().max()
            as_of = max(as_of, d.strftime("%Y-%m-%d"))
        except Exception:
            pass

    all_view = _compute_firm_view(allh, fund_names)
    ex_tmf_view = (_compute_firm_view(allh[allh["_fund"].isin(ex_tmf_funds)].copy(), ex_tmf_funds)
                   if (frames and ex_tmf_funds) else
                   _compute_firm_view(pd.DataFrame(), ex_tmf_funds))

    kospi = load_kospi_sectors()
    # WICS 대/중/소 3종 비교 데이터 (KOSPI vs 전사 차트의 새 토글용)
    gics_code_map = get_stock_gics_code_map()
    allh_ex = allh[allh["_fund"].isin(ex_tmf_funds)] if (frames and ex_tmf_funds) else allh.iloc[0:0]
    kospi_cmp_wics_all, kospi_cmp_wics_ex = {}, {}
    for lv in ("big", "mid", "small"):
        kospi_lv = kospi_sectors_by_level(lv)
        kospi_cmp_wics_all[lv] = kospi_compare(view_sectors_by_level(allh, lv, gics_code_map), kospi_lv)
        kospi_cmp_wics_ex[lv]  = kospi_compare(view_sectors_by_level(allh_ex, lv, gics_code_map), kospi_lv)

    # 알파 산점도 — 전사 종목별 (all / ex_tmf 두 뷰)
    perf_map = get_stock_perf_map()
    bm = get_market_benchmark()
    def _codes_for(df):
        if df is None or len(df) == 0:
            return []
        # 종목코드별로 종목명·평가액 합산 → 대표 이름 사용
        g = df.groupby("종목코드").agg(
            **{"종목명": ("종목명", "first"), "평가액": ("평가액", "sum")}
        ).reset_index()
        total = float(g["평가액"].sum())
        if total <= 0:
            return []
        return [(r["종목코드"], r["종목명"], round(float(r["평가액"])/total*100, 2))
                for _, r in g.iterrows()]
    alpha_all = build_alpha_scatters(_codes_for(allh), perf_map, bm)
    alpha_ex = build_alpha_scatters(_codes_for(allh_ex), perf_map, bm)
    return {
        "as_of": as_of,
        "fund_names": fund_names,
        "tmf_funds": tmf_funds,
        "all": all_view,
        "ex_tmf": ex_tmf_view,
        "kospi_cmp_all": kospi_compare(all_view["firm_sectors"], kospi),
        "kospi_cmp_ex_tmf": kospi_compare(ex_tmf_view["firm_sectors"], kospi),
        "kospi_cmp_wics_all": kospi_cmp_wics_all,
        "kospi_cmp_wics_ex_tmf": kospi_cmp_wics_ex,
        "alpha_all": alpha_all,
        "alpha_ex_tmf": alpha_ex,
    }


# ---------------------------------------------------------------------------
# 트리맵(히트맵)을 서버에서 직접 계산해서 HTML로 그린다 (외부 라이브러리 불필요)
# ---------------------------------------------------------------------------
def _layout_row(sizes, x, y, dx, dy):
    covered = sum(sizes)
    width = covered / dy
    rects, yy = [], y
    for s in sizes:
        rects.append({"x": x, "y": yy, "dx": width, "dy": s / width})
        yy += s / width
    return rects


def _layout_col(sizes, x, y, dx, dy):
    covered = sum(sizes)
    height = covered / dx
    rects, xx = [], x
    for s in sizes:
        rects.append({"x": xx, "y": y, "dx": s / height, "dy": height})
        xx += s / height
    return rects


def _layout(sizes, x, y, dx, dy):
    return _layout_row(sizes, x, y, dx, dy) if dx >= dy else _layout_col(sizes, x, y, dx, dy)


def _leftover(sizes, x, y, dx, dy):
    covered = sum(sizes)
    if dx >= dy:
        w = covered / dy
        return x + w, y, dx - w, dy
    h = covered / dx
    return x, y + h, dx, dy - h


def _worst(sizes, x, y, dx, dy):
    rects = _layout(sizes, x, y, dx, dy)
    return max(max(r["dx"] / r["dy"], r["dy"] / r["dx"]) for r in rects if r["dx"] > 0 and r["dy"] > 0)


def _squarify(sizes, x, y, dx, dy):
    """정규화된 면적 목록을 사각형 좌표 목록으로 (squarified treemap)."""
    sizes = [float(s) for s in sizes if s > 0]
    if not sizes:
        return []
    if len(sizes) == 1:
        return _layout(sizes, x, y, dx, dy)
    i = 1
    while i < len(sizes) and _worst(sizes[:i], x, y, dx, dy) >= _worst(sizes[:i + 1], x, y, dx, dy):
        i += 1
    current, remaining = sizes[:i], sizes[i:]
    nx, ny, ndx, ndy = _leftover(current, x, y, dx, dy)
    return _layout(current, x, y, dx, dy) + _squarify(remaining, nx, ny, ndx, ndy)


def squarify(values, x, y, dx, dy):
    vals = [max(0.0, float(v)) for v in values]
    total = sum(vals)
    if total <= 0:
        return [{"x": x, "y": y, "dx": 0, "dy": 0} for _ in vals]
    area = dx * dy
    scaled = [v * area / total for v in vals]
    # 0인 항목은 squarify에서 빠지므로, 순서 유지 위해 직접 매핑
    nonzero = [s for s in scaled if s > 0]
    rects = _squarify(nonzero, x, y, dx, dy)
    out, j = [], 0
    for s in scaled:
        if s > 0:
            out.append(rects[j]); j += 1
        else:
            out.append({"x": x, "y": y, "dx": 0, "dy": 0})
    return out


def _pnl_color(p):
    """평가손익률 -> 색 (빨강=이익, 회색=중립, 파랑=손실 / 한국식)."""
    if p is None:
        return "rgb(132,147,163)"
    cap = 25.0
    t = max(-1.0, min(1.0, p / cap))
    g, up, dn = (124, 139, 156), (201, 58, 46), (21, 101, 192)
    to = up if t >= 0 else dn
    k = abs(t)
    c = tuple(round(g[i] + (to[i] - g[i]) * k) for i in range(3))
    return f"rgb({c[0]},{c[1]},{c[2]})"


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))


def treemap_flat_html(items, height=300):
    """평면 트리맵: 종목 타일 (크기=value, 색=pnl)."""
    items = [it for it in items if (it.get("value") or 0) > 0]
    items = sorted(items, key=lambda d: -d["value"])
    rects = squarify([it["value"] for it in items], 0, 0, 100, 100)
    tiles = []
    for it, r in zip(items, rects):
        if r["dx"] <= 0 or r["dy"] <= 0:
            continue
        pct = it.get("pct")
        pnl = it.get("pnl")
        tip = f'{_esc(it["name"])} · 비중 {pct}% · 평가손익률 {pnl}%'
        sub = f'{pct:.1f}%' if pct is not None else ""
        tiles.append(
            f'<div class="tmt" title="{tip}" style="left:{r["x"]:.3f}%;top:{r["y"]:.3f}%;'
            f'width:{r["dx"]:.3f}%;height:{r["dy"]:.3f}%;background:{_pnl_color(pnl)}">'
            f'<span class="tmn">{_esc(it["name"])}</span><span class="tmp">{sub}</span></div>')
    return f'<div class="tm" style="height:{height}px">{"".join(tiles)}</div>'


def treemap_grouped_html(items, height=300):
    """그룹 트리맵: 섹터로 묶고 그 안에 종목 타일."""
    items = [it for it in items if (it.get("value") or 0) > 0]
    secs = {}
    for it in items:
        secs.setdefault(it.get("sector", "기타"), []).append(it)
    sec_list = sorted(secs.items(), key=lambda kv: -sum(s["value"] for s in kv[1]))
    sec_rects = squarify([sum(s["value"] for s in stocks) for _, stocks in sec_list], 0, 0, 100, 100)

    out = []
    for (sname, stocks), sr in zip(sec_list, sec_rects):
        if sr["dx"] <= 0 or sr["dy"] <= 0:
            continue
        out.append(
            f'<div class="tmsec" style="left:{sr["x"]:.3f}%;top:{sr["y"]:.3f}%;'
            f'width:{sr["dx"]:.3f}%;height:{sr["dy"]:.3f}%">'
            f'<span class="tmcap">{_esc(sname)}</span>')
        stocks = sorted(stocks, key=lambda d: -d["value"])
        rects = squarify([s["value"] for s in stocks], 0, 0, 100, 100)
        for s, r in zip(stocks, rects):
            if r["dx"] <= 0 or r["dy"] <= 0:
                continue
            tip = f'{_esc(s["name"])} · 비중 {s.get("pct")}% · 평가손익률 {s.get("pnl")}%'
            out.append(
                f'<div class="tmt" title="{tip}" style="left:{r["x"]:.3f}%;top:{r["y"]:.3f}%;'
                f'width:{r["dx"]:.3f}%;height:{r["dy"]:.3f}%;background:{_pnl_color(s.get("pnl"))}">'
                f'<span class="tmn">{_esc(s["name"])}</span></div>')
        out.append("</div>")
    return f'<div class="tm" style="height:{height}px">{"".join(out)}</div>'


# ---------------------------------------------------------------------------
# 3) HTML 만들기
# ---------------------------------------------------------------------------
STYLE = """
  :root{ --bg:#eef1f6; --card:#fff; --ink:#16202e; --muted:#7a8696;
    --line:#e4e8ef; --navy:#16314f; --navy2:#234e7d;
    --green:#2e7d32; --blue:#2962ff; --red:#e2574c; --up:#d23b34; --down:#1565c0;
    --soft:#f6f8fb; --soft2:#f8fafc; --row-hover:#f7f9fc; --card-shadow:rgba(22,49,79,.08); }
  /* ===== 다크모드 ===== */
  body[data-theme="dark"]{
    --bg:#0f1822; --card:#1a2532; --ink:#e1e7ef; --muted:#9aa6b6;
    --line:#2a3645; --navy:#0a1422; --navy2:#5b94d8;
    --up:#ff5b4f; --down:#5ba0ff; --green:#7cc97c; --blue:#7eb7ff;
    --soft:#222e3d; --soft2:#1f2a37; --row-hover:#283649;
    --card-shadow:rgba(0,0,0,.4);
  }
  /* 헤더/탭 */
  body[data-theme="dark"] .topbar{ background:#0a1422; }
  body[data-theme="dark"] .tab{ color:#8093a8; }
  body[data-theme="dark"] .tab:hover{ background:rgba(255,255,255,.05); color:#cad4e0; }
  body[data-theme="dark"] .tab.active{ background:var(--bg); color:#e1e7ef; }
  body[data-theme="dark"] .searchbox{ background:rgba(255,255,255,.08); color:#fff; }
  body[data-theme="dark"] .searchbox::placeholder{ color:#7e92aa; }
  body[data-theme="dark"] .searchresults{ background:var(--card); box-shadow:0 8px 24px rgba(0,0,0,.5); }
  body[data-theme="dark"] .searchresults a{ color:#e1e7ef; }
  body[data-theme="dark"] .searchresults a:hover{ background:var(--row-hover); }
  body[data-theme="dark"] .searchresults a span{ color:var(--muted); }

  /* 카드/표/타일 — 어두운 카드 위에서 또 한 단 어두운 색으로 강조 */
  body[data-theme="dark"] .card{ background:var(--card); border-color:var(--line); }
  body[data-theme="dark"] .stat,
  body[data-theme="dark"] .tile,
  body[data-theme="dark"] .idx-card,
  body[data-theme="dark"] .sec-row,
  body[data-theme="dark"] .rs-card,
  body[data-theme="dark"] .rs-tp{ background:var(--soft); border-color:var(--line); }
  body[data-theme="dark"] table{ color:var(--ink); }
  body[data-theme="dark"] thead th,
  body[data-theme="dark"] .mkt-tbl thead th{
    background:var(--soft2); color:var(--muted); border-bottom-color:var(--line); }
  body[data-theme="dark"] tbody td,
  body[data-theme="dark"] .mkt-tbl tbody td{ border-bottom-color:#222e3d; }
  body[data-theme="dark"] tbody tr:hover,
  body[data-theme="dark"] .mkt-tbl tbody tr:hover{ background:var(--row-hover); }
  body[data-theme="dark"] td.muted,
  body[data-theme="dark"] .mkt-tbl td.muted{ color:var(--muted); }
  body[data-theme="dark"] td.tag{ color:#a5b3c5; }

  /* seg 토글 (펀드 페이지 등의 위험/베타/매트릭스 전환) */
  body[data-theme="dark"] .seg{ background:var(--soft2); border-color:var(--line); }
  body[data-theme="dark"] .seg button{ color:var(--muted); }
  body[data-theme="dark"] .seg button.on{ background:var(--card); color:#e1e7ef;
    box-shadow:0 1px 2px rgba(0,0,0,.5); }

  /* 매트릭스 sticky 헤더 색 */
  body[data-theme="dark"] .mtx thead th{ background:#0a1422; }

  /* 리서치 카드 — 코드 배지, 증권사 배지, EPS 배지 */
  body[data-theme="dark"] .rs-card .rs-stock .code{ color:var(--muted); }
  body[data-theme="dark"] .rs-broker{ background:var(--soft2); color:var(--muted); }
  body[data-theme="dark"] .rs-card .rs-eps.up{ background:rgba(255,91,79,.18); color:#ff8779; }
  body[data-theme="dark"] .rs-card .rs-eps.dn{ background:rgba(91,160,255,.18); color:#7eb7ff; }
  body[data-theme="dark"] .rs-empty{ background:var(--soft); border-color:var(--line); }
  body[data-theme="dark"] .rs-search{ background:var(--soft2); color:var(--ink); border-color:var(--line); }
  body[data-theme="dark"] .rs-date{ background:var(--soft2); color:var(--ink); border-color:var(--line); }
  body[data-theme="dark"] .rs-toolbar{ background:var(--card); border-color:var(--line); }

  /* 종목 페이지 종가 박스, 코드 배지 */
  body[data-theme="dark"] .stockhead .code{ background:var(--soft2); color:var(--navy2); }
  body[data-theme="dark"] .freqseg{ background:var(--soft2); border-color:var(--line); }
  body[data-theme="dark"] .freqseg button{ color:var(--muted); }
  body[data-theme="dark"] .freqseg button.on{ background:var(--card); color:#e1e7ef; }

  /* 시장 페이지 섹터 행 호버 */
  body[data-theme="dark"] .sec-head:hover{ background:var(--row-hover); }
  body[data-theme="dark"] .sec-head .cap{ color:var(--muted); }
  body[data-theme="dark"] .sec-head .bar{ background:#0a1422; }

  /* 종목 코드 배지 (전사 매트릭스 sticky 컬럼 등) */
  body[data-theme="dark"] .mtx tbody td:nth-child(1){ background:var(--card); }
  body[data-theme="dark"] .mtx tbody td.nm{ background:var(--card); border-right-color:var(--line); }
  /* 다크모드에서 매트릭스 전사비중 컬럼 — 카드보다 약간 더 어둡게 */
  body[data-theme="dark"] .mtx td.firm{ background:#0f1a26; color:#e1e7ef; }

  /* 코스닥/검색 결과 다크 */
  body[data-theme="dark"] .v.pos{ color:var(--up); }
  body[data-theme="dark"] .v.neg{ color:var(--down); }

  /* 토글 버튼 */
  .theme-toggle{ background:rgba(255,255,255,.13); color:#fff; border:0;
    padding:6px 10px; border-radius:7px; font-size:14px; cursor:pointer;
    font-family:inherit; line-height:1; margin-left:8px; }
  .theme-toggle:hover{ background:rgba(255,255,255,.22); }
  /* 섹터 분류 토글 — 그래프 카드 옆 */
  .sector-mode-sel{ padding:5px 10px; border:1px solid var(--line);
    border-radius:7px; font-size:12.5px; font-family:inherit;
    background:#fff; color:var(--ink); cursor:pointer; }
  body[data-theme="dark"] .sector-mode-sel{ background:var(--soft2);
    color:var(--ink); border-color:var(--line); }
  *{ box-sizing:border-box; }
  body{ margin:0; background:var(--bg); color:var(--ink);
    font-family:'Pretendard','Apple SD Gothic Neo','Malgun Gothic',sans-serif;
    font-feature-settings:"tnum"; }
  .topbar{ background:var(--navy); color:#fff; padding:0 28px; position:relative; }
  .topbar .brandrow{ display:flex; align-items:center; justify-content:space-between;
    padding:18px 0 4px; gap:16px; flex-wrap:wrap; }
  .topbar .brand{ display:flex; align-items:baseline; gap:12px; }
  .topbar .brand b{ font-size:20px; font-weight:800; letter-spacing:-.3px; }
  .topbar .brand span{ font-size:13px; color:#9fb6d0; }
  .searchbar{ position:relative; }
  .searchbox{ width:280px; padding:8px 32px 8px 32px; border-radius:8px; border:0;
    background:rgba(255,255,255,.13); color:#fff; font-size:13px; font-family:inherit;
    outline:none; }
  .searchbox::placeholder{ color:#9fb6d0; }
  .searchbox:focus{ background:rgba(255,255,255,.22); }
  .searchbar::before{ content:"🔎"; position:absolute; left:10px; top:8px; font-size:13px;
    pointer-events:none; opacity:.7; }
  .searchresults{ position:absolute; top:calc(100% + 4px); right:0; width:320px;
    background:#fff; color:var(--ink); border-radius:10px; box-shadow:0 8px 24px rgba(0,0,0,.18);
    display:none; max-height:380px; overflow:auto; z-index:50; padding:6px; }
  .searchresults a{ display:flex; justify-content:space-between; align-items:center;
    padding:9px 12px; color:var(--ink); text-decoration:none; border-radius:7px;
    font-size:13.5px; }
  .searchresults a:hover{ background:#f3f6fb; }
  .searchresults a b{ font-weight:700; }
  .searchresults a span{ color:var(--muted); font-size:12px; font-family:monospace; }
  .tabs{ display:flex; gap:4px; flex-wrap:wrap; }
  .tab{ color:#aebfd4; text-decoration:none; padding:11px 18px; font-size:14px;
    font-weight:600; border-radius:8px 8px 0 0; }
  .tab:hover{ color:#fff; background:rgba(255,255,255,.07); }
  .tab.active{ color:var(--navy); background:var(--bg); }
  .tab.firm{ color:#ffd98a; }
  .tab.firm.active{ color:var(--navy); background:var(--bg); }
  .tab.kosdaq{ color:#8ee0c4; }
  .tab.kosdaq.active{ color:var(--navy); background:var(--bg); }
  .tab.market{ color:#9cc8ff; }
  .tab.market.active{ color:var(--navy); background:var(--bg); }
  .tab.research{ color:#f4a8c1; }
  .tab.research.active{ color:var(--navy); background:var(--bg); }
  .wrap{ max-width:1280px; margin:0 auto; padding:24px 28px 60px; }

  .head{ display:flex; align-items:flex-end; justify-content:space-between;
    margin-bottom:18px; flex-wrap:wrap; gap:10px; }
  .head h1{ font-size:22px; margin:0; font-weight:800; letter-spacing:-.4px; }
  .head .asof{ color:var(--muted); font-size:13px; }
  .head .subname{ color:var(--navy2); font-size:13.5px; font-weight:600; margin-top:3px; }

  .stats{ display:grid; grid-template-columns:repeat(5,1fr); gap:14px; margin-bottom:20px; }
  .stats.s4{ grid-template-columns:repeat(4,1fr); }
  .stat{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:16px 18px; }
  .stat .k{ font-size:12px; color:var(--muted); font-weight:600; }
  .stat .v{ font-size:24px; font-weight:800; margin-top:6px; letter-spacing:-.5px; }
  .stat .v small{ font-size:14px; font-weight:700; color:var(--muted); }
  .v.pos{ color:var(--up); } .v.neg{ color:var(--down); }

  .card{ background:var(--card); border:1px solid var(--line); border-radius:14px;
    padding:20px 22px; margin-bottom:20px; }
  .card h2{ font-size:15px; margin:0 0 14px; font-weight:700; letter-spacing:-.2px;
    display:flex; align-items:center; gap:8px; }
  .card h2::before{ content:""; width:4px; height:15px; background:var(--navy2);
    border-radius:2px; display:inline-block; }
  .grid2{ display:grid; grid-template-columns:1fr 1fr; gap:20px; }

  table{ width:100%; border-collapse:collapse; font-size:13px; }
  thead th{ text-align:right; color:var(--muted); font-weight:600; font-size:12px;
    padding:8px 10px; border-bottom:2px solid var(--line); white-space:nowrap; }
  thead th:nth-child(2){ text-align:left; }
  tbody td{ padding:8px 10px; border-bottom:1px solid #f0f2f6; white-space:nowrap; }
  td.num{ text-align:right; } td.name{ font-weight:700; } td.strong{ font-weight:700; }
  td.muted{ color:var(--muted); } td.pos{ color:var(--up); font-weight:600; }
  td.neg{ color:var(--down); font-weight:600; }
  /* 일반 selector (td 외 span 등에서도 +/- 색상 적용) */
  .pos{ color:var(--up); }
  .neg{ color:var(--down); }
  .flat{ color:var(--muted); }
  td.tag{ color:#41556e; } td.tag.size{ font-weight:600; }
  tbody tr:hover{ background:#f7f9fc; }

  .legend{ display:flex; gap:18px; font-size:12.5px; color:var(--muted);
    margin-top:10px; justify-content:center; }
  .legend i{ width:11px; height:11px; border-radius:3px; display:inline-block;
    margin-right:6px; vertical-align:-1px; }

  .rgroup{ margin-bottom:16px; } .rgroup:last-child{ margin-bottom:0; }
  .rgroup .sub{ font-size:12px; color:var(--navy2); font-weight:700; margin:0 0 9px; }
  .rgrid{ display:grid; grid-template-columns:repeat(auto-fill,minmax(150px,1fr)); gap:10px; }
  .tile{ background:#f6f8fb; border:1px solid var(--line); border-radius:10px; padding:11px 13px; }
  .tile .lab{ font-size:11.5px; color:var(--muted); font-weight:600;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
  .tile .val{ font-size:18px; font-weight:800; margin-top:4px; letter-spacing:-.3px; }

  .seg{ display:inline-flex; gap:2px; background:#eef1f6; border:1px solid var(--line);
    border-radius:9px; padding:3px; margin-left:auto; }
  .seg button{ border:0; background:transparent; color:var(--muted); font:inherit;
    font-size:12.5px; font-weight:600; padding:6px 12px; border-radius:7px; cursor:pointer; }
  .seg button.on{ background:#fff; color:var(--navy); box-shadow:0 1px 2px rgba(0,0,0,.08); }
  .h2row{ display:flex; align-items:center; }

  /* 분류 기준 태그 — h2 옆에 붙는 작은 회색 배지 */
  .src-tag{ display:inline-block; margin-left:8px; padding:2px 8px;
    background:#eef1f6; color:var(--muted); font-size:11px; font-weight:600;
    letter-spacing:.2px; border-radius:5px; vertical-align:2px; }
  body[data-theme="dark"] .src-tag{ background:rgba(255,255,255,.08); }

  /* WICS 대/중/소 토글 행 — 차트 바로 아래 */
  .wics-toggle-row{ display:flex; align-items:center; gap:10px;
    margin-top:12px; padding-top:12px; border-top:1px solid var(--line); }
  .wics-toggle-row .wics-label{ font-size:13px; font-weight:700; color:var(--ink); }
  .wics-toggle-row .seg{ margin-left:0; }
  .wics-toggle-row .wics-count{ margin-left:auto; font-size:12px; color:var(--muted); }

  /* 조건별 종목 스크리닝 보드 */
  .scr-summary{ display:flex; gap:18px; flex-wrap:wrap; font-size:13px;
    padding:8px 12px; background:#f4f7fb; border:1px solid var(--line);
    border-radius:8px; margin-bottom:14px; }
  body[data-theme="dark"] .scr-summary{ background:var(--soft2); }
  .scr-summary .pos{ color:var(--up); } .scr-summary .neg{ color:var(--down); }
  .scr-group{ margin-bottom:16px; }
  .scr-group-title{ font-size:12px; font-weight:800; color:var(--muted);
    letter-spacing:.3px; margin-bottom:6px; }
  .scr-cols{ display:flex; gap:10px; overflow-x:auto; padding-bottom:4px; }
  .scr-col{ flex:1 0 150px; min-width:150px; background:var(--card);
    border:1px solid var(--line); border-radius:8px; overflow:hidden; }
  .scr-col-head{ font-size:11.5px; font-weight:700; padding:7px 9px;
    line-height:1.3; display:flex; align-items:flex-start; gap:4px;
    border-bottom:1px solid var(--line); background:#eef1f6; color:var(--ink); }
  .scr-col-head.up{ background:#fdecea; color:#c0392b; }
  .scr-col-head.dn{ background:#e7f0fd; color:#1565c0; }
  .scr-col-head.neu{ background:#fff6e0; color:#8a6d00; }
  body[data-theme="dark"] .scr-col-head.up{ background:rgba(201,58,46,.20); color:#f4a79f; }
  body[data-theme="dark"] .scr-col-head.dn{ background:rgba(21,101,192,.24); color:#9cc2f0; }
  body[data-theme="dark"] .scr-col-head.neu{ background:rgba(212,170,20,.18); color:#e6cd7a; }
  .scr-col-head .scr-cnt{ margin-left:auto; font-size:10.5px; opacity:.7; font-weight:800; }
  .scr-col-body{ max-height:280px; overflow-y:auto; padding:4px 0; }
  .scr-item{ display:block; padding:4px 9px; font-size:12px; color:var(--ink);
    text-decoration:none; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
  a.scr-item:hover{ background:#eef3f9; color:var(--navy); }
  body[data-theme="dark"] a.scr-item:hover{ background:rgba(255,255,255,.06); }
  .scr-item.nolink{ color:var(--muted); }

  /* 업종별 수급 투자자 토글 행 */
  .flow-inv-row{ display:flex; align-items:center; gap:10px;
    margin:12px 0; flex-wrap:wrap; }
  .flow-inv-row .wics-label{ font-size:13px; font-weight:700; color:var(--ink); }
  .flow-inv-row .seg{ margin-left:0; }
  .flow-inv-row .wics-count{ margin-left:auto; font-size:12px; color:var(--muted); }

  /* 대분류 차트 위 안내 문구 */
  .chart-sub{ font-size:12px; color:var(--muted); margin:4px 0 8px; }
  #sectorBarsBig{ cursor:pointer; }

  .tmbox{ height:380px; position:relative; }
  .tmbox.sm{ height:300px; }
  .tmbox.lg{ height:560px; }
  .topctl{ margin-bottom:14px; display:flex; justify-content:flex-end; }
  .cmpbox{ height:360px; position:relative; }
  .size3{ display:grid; grid-template-columns:repeat(3,1fr); gap:18px; }
  .sizebox{ text-align:center; }
  .sizebox .cap{ font-size:12px; color:var(--muted); font-weight:700; margin-bottom:6px; }
  @media (max-width:860px){ .size3{ grid-template-columns:1fr; } }

  /* 알파 산점도 (시장 대비 종목 성과) */
  .alpha-sub{ font-size:12.5px; color:var(--muted); margin-bottom:12px; }
  .alpha-sub b{ color:var(--ink); }
  .alpha-grid{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }
  .alpha-box{ background:#fafbfd; border:1px solid var(--line); border-radius:10px;
    padding:10px 12px; }
  body[data-theme="dark"] .alpha-box{ background:var(--soft2); }
  .alpha-box .alpha-cap{ font-size:12px; font-weight:700; color:var(--ink);
    margin-bottom:6px; text-align:center; }
  .alpha-box canvas{ height:340px !important; }
  @media (max-width:860px){ .alpha-grid{ grid-template-columns:1fr; } }

  /* 삼중 도넛 — 펀드/전사/KOSPI 동심원 */
  .nested-wrap{ display:grid; grid-template-columns:1fr 280px; gap:24px; align-items:center; }
  .nested-chart{ position:relative; min-height:340px; }
  .nested-chart canvas{ max-height:380px; }
  .nested-side{ font-size:13px; }
  .nested-side .ns-title{ font-size:11px; font-weight:800; color:var(--muted);
    margin-bottom:6px; letter-spacing:.3px; text-transform:uppercase; }
  .nested-side .ns-ringkey,.nested-side .ns-colorkey{
    background:#f7f9fc; border:1px solid var(--line);
    border-radius:10px; padding:10px 12px; margin-bottom:10px; }
  body[data-theme="dark"] .nested-side .ns-ringkey,
  body[data-theme="dark"] .nested-side .ns-colorkey{ background:var(--soft2); }
  .nested-side .ns-row{ display:flex; align-items:center; gap:8px;
    padding:3px 0; font-size:12.5px; }
  .nested-side .ns-dot{ width:14px; height:14px; border-radius:50%; flex-shrink:0;
    border:2px solid var(--ink); background:transparent; }
  .nested-side .ns-dot.ring-fund{ border-color:#16314f; border-width:3px; }
  .nested-side .ns-dot.ring-firm{ border-color:#5b91c4; border-width:3px; }
  .nested-side .ns-dot.ring-kospi{ border-color:#c9a227; border-width:3px; }
  .nested-side .ns-sw{ width:14px; height:14px; border-radius:3px; flex-shrink:0; }
  .ns-table{ font-size:12px; }
  .ns-table table{ width:100%; border-collapse:collapse; }
  .ns-table th,.ns-table td{ padding:5px 6px; border-bottom:1px solid var(--line); }
  .ns-table th{ font-size:11px; color:var(--muted); font-weight:700; text-align:right; }
  .ns-table th:first-child,.ns-table td:first-child{ text-align:left; color:var(--muted); }
  .ns-table td{ text-align:right; font-variant-numeric:tabular-nums; }
  @media (max-width:860px){ .nested-wrap{ grid-template-columns:1fr; } }

  /* ===== 시장 스크리닝 페이지 ===== */
  .mkt-tbl{ width:100%; border-collapse:collapse; font-size:13px; }
  .mkt-tbl thead th{ text-align:right; color:var(--muted); font-weight:600;
    font-size:11.5px; padding:7px 10px; border-bottom:1.5px solid var(--line);
    background:#f8fafc; white-space:nowrap; }
  .mkt-tbl thead th:first-child{ text-align:left; }
  .mkt-tbl tbody td{ padding:7px 10px; border-bottom:1px solid #f0f2f6;
    white-space:nowrap; font-feature-settings:"tnum"; }
  .mkt-tbl td.name{ font-weight:600; text-align:left; }
  .mkt-tbl td.num{ text-align:right; }
  .mkt-tbl td.strong{ font-weight:700; }
  .mkt-tbl td.pos{ color:var(--up); font-weight:600; }
  .mkt-tbl td.neg{ color:var(--down); font-weight:600; }
  .mkt-tbl td.muted{ color:var(--muted); }
  .mkt-tbl tbody tr:hover{ background:#f7f9fc; }
  .secchart{ position:relative; height:420px; }
  .idx-grid{ display:grid; grid-template-columns:repeat(auto-fill,minmax(160px,1fr)); gap:12px; }
  .idx-card{ background:#f6f8fb; border:1px solid var(--line); border-radius:11px;
    padding:13px 14px; display:flex; flex-direction:column; gap:4px; }
  .idx-card .nm{ font-size:12.5px; color:var(--muted); font-weight:700;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
  .idx-card .px{ font-size:18px; font-weight:800; letter-spacing:-.3px; }
  .idx-card .chg{ font-size:12.5px; font-weight:700; }
  .idx-card .chg.up{ color:var(--up); }
  .idx-card .chg.dn{ color:var(--down); }
  .idx-card .chg.flat{ color:var(--muted); }
  .idx-card .dt{ font-size:11px; color:var(--muted); }

  .sec-list{ display:flex; flex-direction:column; gap:8px; }
  .sec-row{ background:#f6f8fb; border:1px solid var(--line); border-radius:10px;
    overflow:hidden; }
  .sec-head{ display:flex; align-items:center; gap:12px; padding:11px 14px;
    cursor:pointer; user-select:none; }
  .sec-head:hover{ background:#eef2f7; }
  .sec-head .arrow{ width:14px; font-size:11px; color:var(--muted); transition:transform .15s; }
  .sec-row.open .sec-head .arrow{ transform:rotate(90deg); }
  .sec-head .nm{ font-weight:700; flex:1; font-size:14px; }
  .sec-head .cap{ color:var(--muted); font-size:12px; }
  .sec-head .chg{ font-weight:700; font-size:14px; min-width:70px; text-align:right; }
  .sec-head .chg.up{ color:var(--up); } .sec-head .chg.dn{ color:var(--down); }
  .sec-head .bar{ flex-basis:120px; flex-shrink:0; height:8px; background:#e8edf3;
    border-radius:4px; position:relative; overflow:hidden; }
  .sec-head .bar i{ position:absolute; top:0; height:100%; }
  .sec-head .bar i.up{ background:var(--up); left:50%; }
  .sec-head .bar i.dn{ background:var(--down); right:50%; }
  .sec-body{ display:none; padding:0 14px 12px; }
  .sec-row.open .sec-body{ display:block; }

  /* 블룸버그 카드 — 중분류(sub) nested 아코디언 */
  .bl-sub-list{ display:flex; flex-direction:column; gap:6px; padding-top:8px; }
  .bl-sub{ background:#fff; border:1px solid var(--line); border-radius:8px; }
  body[data-theme="dark"] .bl-sub{ background:var(--soft2); }
  .bl-sub-head{ display:flex; align-items:center; gap:10px; padding:9px 12px;
    cursor:pointer; font-size:13px; }
  .bl-sub-head:hover{ background:#f4f7fb; }
  body[data-theme="dark"] .bl-sub-head:hover{ background:rgba(255,255,255,.04); }
  .bl-sub-head .arrow{ width:12px; font-size:10px; color:var(--muted);
    transition:transform .15s; }
  .bl-sub.open .bl-sub-head .arrow{ transform:rotate(90deg); }
  .bl-sub-head .nm{ font-weight:600; flex:1; }
  .bl-sub-head .chg{ font-weight:700; font-size:13px; min-width:80px; text-align:right; }
  .bl-sub-head .chg.up{ color:var(--up); } .bl-sub-head .chg.dn{ color:var(--down); }
  .bl-sub-head .cap{ color:var(--muted); font-size:11.5px; }
  .bl-sub-body{ display:none; padding:0 12px 12px; overflow-x:auto; }
  .bl-sub.open .bl-sub-body{ display:block; }
  /* 중분류가 하나뿐인 경우 → 자동 펼침 */
  .bl-sub-list.one-sub .bl-sub .bl-sub-body{ display:block; }
  .bl-sub-list.one-sub .bl-sub .arrow{ transform:rotate(90deg); }

  /* 블룸버그 종목 표 — 국가/PER 강조 */
  .bl-tbl td.ctry{ width:24px; text-align:center; font-size:14px;
    padding-left:6px; padding-right:6px; }
  /* 한국 종목 행 — 은은한 파랑 배경 + 좌측 세로줄 */
  .bl-tbl tr.kr-row{ background:#f0f6fd; }
  .bl-tbl tr.kr-row td:first-child{ box-shadow:inset 3px 0 0 #1976d2; }
  body[data-theme="dark"] .bl-tbl tr.kr-row{ background:rgba(25,118,210,.10); }
  /* PER 컬럼 — 헤더/셀 모두 강조 (연노랑 배경, 굵은 폰트, 좌측 세로선) */
  .bl-tbl th.per, .bl-tbl td.per{ background:#fffaea; font-weight:700;
    color:#7a5a00; }
  .bl-tbl th.per{ background:#e8d982; color:#2a2000; }
  .bl-tbl th.per:first-of-type, .bl-tbl td.per:first-of-type{
    border-left:2px solid #d4bf50; }
  body[data-theme="dark"] .bl-tbl th.per, body[data-theme="dark"] .bl-tbl td.per{
    background:rgba(212,191,80,.18); color:#eadf9a; }
  body[data-theme="dark"] .bl-tbl th.per{ background:rgba(212,191,80,.35); color:#fff8d0; }
  .sec-body table{ font-size:12.5px; }
  .sec-body thead th{ font-size:11px; padding:6px 8px; }
  .sec-body tbody td{ padding:6px 8px; }

  .refresh-btn{ background:var(--navy2); color:#fff; border:0; border-radius:7px;
    padding:6px 12px; font-size:12.5px; font-weight:600; cursor:pointer;
    font-family:inherit; margin-left:10px; }
  .refresh-btn:hover{ background:var(--navy); }

  @media (max-width:600px){
    .idx-grid{ grid-template-columns:repeat(2,1fr); gap:8px; }
    .idx-card{ padding:10px 11px; }
    .idx-card .px{ font-size:15px; }
    .sec-head{ padding:10px 11px; gap:8px; flex-wrap:wrap; }
    .sec-head .bar{ flex-basis:100%; order:99; }
    .sec-head .cap{ display:none; }
  }
  .tm{ position:relative; width:100%; border-radius:8px; overflow:hidden; background:#0e1b2a; }
  .tmt{ position:absolute; overflow:hidden; border:1px solid rgba(255,255,255,.9);
    display:flex; flex-direction:column; justify-content:center; align-items:center;
    color:#fff; text-align:center; padding:1px; }
  .tmt .tmn{ font-size:11px; font-weight:700; line-height:1.05; max-width:100%;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
    text-shadow:0 1px 2px rgba(0,0,0,.45); padding:0 2px; }
  .tmt .tmp{ font-size:10px; opacity:.92; text-shadow:0 1px 2px rgba(0,0,0,.45); }
  .tmsec{ position:absolute; overflow:hidden; border:2px solid #fff; }
  .tmcap{ position:absolute; left:4px; top:2px; z-index:2; pointer-events:none;
    font-size:10.5px; font-weight:800; color:#fff; text-shadow:0 1px 3px rgba(0,0,0,.7); }
  .tmlegend{ display:flex; gap:14px; justify-content:center; font-size:12px;
    color:var(--muted); margin-top:10px; align-items:center; }
  .tmlegend i{ width:14px; height:11px; border-radius:2px; display:inline-block; vertical-align:-1px; }

  .mtxbox{ max-height:600px; overflow:auto; border:1px solid var(--line); border-radius:10px; }
  .mtx{ border-collapse:separate; border-spacing:0; font-size:12.5px; width:100%; }
  .mtx th, .mtx td{ padding:7px 10px; white-space:nowrap; border-bottom:1px solid #eef1f6; }
  .mtx thead th{ position:sticky; top:0; background:var(--navy); color:#fff;
    font-weight:600; font-size:12px; z-index:2; text-align:center; }
  .mtx thead th.l{ text-align:left; }
  .mtx td.c{ text-align:center; color:var(--muted); }
  .mtx td.nm{ text-align:left; font-weight:700; }
  .mtx td.cl{ text-align:right; }
  .mtx td.wt{ text-align:center; font-weight:600; }
  .mtx td.firm{ text-align:center; font-weight:700; background:var(--soft, #f6f8fb); }

  /* 섹터별 보기 — 드릴다운 (섹터 클릭 → 종목 펼침) */
  .mtx tr.sec-h{ cursor:pointer; }
  .mtx tr.sec-h:hover{ background:#f4f7fb; }
  body[data-theme="dark"] .mtx tr.sec-h:hover{ background:rgba(255,255,255,.04); }
  .mtx tr.sec-h .arr{ display:inline-block; width:14px; color:var(--muted);
    font-size:10px; transition:transform .15s; margin-right:4px; }
  .mtx tr.sec-h.open{ background:#eef3f9; }
  body[data-theme="dark"] .mtx tr.sec-h.open{ background:rgba(201,162,39,.10); }
  .mtx tr.sec-h .cnt{ display:inline-block; margin-left:8px; padding:1px 7px;
    font-size:11px; font-weight:700; color:var(--muted); background:#eef1f6;
    border-radius:9px; }
  body[data-theme="dark"] .mtx tr.sec-h .cnt{ background:rgba(255,255,255,.08); }
  .mtx tr.sec-sub{ background:#fafbfd; font-size:12px; }
  body[data-theme="dark"] .mtx tr.sec-sub{ background:rgba(255,255,255,.02); }
  .mtx tr.sec-sub td.sub-idx{ color:#aab3bf; font-size:11px; padding-left:24px; }
  .mtx tr.sec-sub td.sub-nm{ font-weight:500; padding-left:28px; color:var(--ink); }
  .mtx tr.sec-sub td.sub-nm .code{ font-family:monospace; font-size:11px;
    color:var(--muted); margin-right:4px; }
  .mtx tr.sec-sub td.firm{ font-weight:600; }
  .mtx tr.sec-sub a:hover{ text-decoration:underline; }

  @media (max-width:860px){ .stats,.stats.s4{ grid-template-columns:repeat(2,1fr); }
    .grid2{ grid-template-columns:1fr; } }

  /* ===== 모바일 (≤600px) ===== */
  @media (max-width:600px){
    .wrap{ padding:14px 12px 40px; }
    .topbar{ padding:0 12px; }
    .topbar .brandrow{ flex-direction:column; align-items:stretch; gap:8px;
      padding:12px 0 4px; }
    .topbar .brand{ justify-content:center; }
    .topbar .brand b{ font-size:16px; }
    .topbar .brand span{ display:none; }
    .searchbar{ width:100%; }
    .searchbox{ width:100%; box-sizing:border-box; }
    .searchresults{ width:100%; right:auto; left:0; }
    .tabs{ overflow-x:auto; flex-wrap:nowrap; padding-bottom:2px; }
    .tab{ padding:9px 12px; font-size:13px; white-space:nowrap; }
    .head h1{ font-size:18px; }
    .head .asof{ font-size:12px; }
    .stats, .stats.s4{ grid-template-columns:1fr 1fr; gap:10px; }
    .stat{ padding:12px 13px; }
    .stat .v{ font-size:19px; }
    .stat .k{ font-size:11px; }
    .card{ padding:14px 14px; margin-bottom:14px; border-radius:11px; }
    .card h2{ font-size:14px; margin-bottom:11px; }
    .h2row{ flex-wrap:wrap; gap:8px; }
    .seg{ margin-left:0; }
    .seg button{ padding:5px 10px; font-size:12px; }
    .tmbox{ height:280px; }
    .tmbox.sm{ height:240px; }
    .tmbox.lg{ height:400px; }
    .size3{ grid-template-columns:1fr; gap:12px; }
    .rgrid{ grid-template-columns:repeat(2,1fr); gap:8px; }
    .tile{ padding:9px 10px; }
    .tile .val{ font-size:15px; }
    table{ font-size:12px; }
    thead th{ padding:6px 6px; font-size:11px; }
    tbody td{ padding:6px 6px; }
    .mtx{ font-size:11px; }
    .mtx th, .mtx td{ padding:5px 7px; }
    /* 매트릭스에서 첫 컬럼(#)·종목명/섹터명을 좌측 sticky 로 */
    .mtx thead th:nth-child(1), .mtx thead th:nth-child(2), .mtx thead th:nth-child(3){
      position:sticky; left:0; z-index:3; }
    .mtx tbody td:nth-child(1){ position:sticky; left:0; background:#fff; z-index:1; }
    .mtx tbody td.nm{ position:sticky; left:24px; background:#fff; z-index:1;
      border-right:1px solid var(--line); }
    .topctl{ justify-content:center; }
  }
"""

COMMON_JS = """
// 섹터 분류 토글 — 쿠키 기반, 모든 페이지에 일괄 적용
(function(){
  const sel = document.getElementById('sectorMode');
  if (!sel) return;
  // 쿠키에서 현재 값 읽어 select 동기화
  const m = document.cookie.match(/sector_mode=([^;]+)/);
  sel.value = m ? m[1] : 'detail';
  sel.addEventListener('change', () => {
    document.cookie = `sector_mode=${sel.value}; path=/; max-age=31536000`;
    location.reload();
  });
})();

// 알파 산점도 — 시장 벤치마크 대비 종목별 (X, Y) 성과 표시.
// data = {bm:{name, wk1, m1, m3}, points:[{code, name, weight, wk1, m1, m3}]}
// xKey/yKey ∈ {'wk1','m1','m3'}. axisLabel: 축 라벨 매핑.
window.drawAlphaScatter = function(canvasId, data, xKey, yKey) {
  if (!data || !data.points) return null;
  const bm = data.bm || {};
  const bmx = bm[xKey], bmy = bm[yKey];
  const AXIS_LABEL = { wk1: '5일 수익률 (%)', m1: '1개월 수익률 (%)', m3: '3개월 수익률 (%)' };
  // 종목 포인트 (편입비를 반경으로) — outperform 색상 (벤치마크 양방향 초과: 빨강, 반대: 파랑)
  const stocks = data.points.filter(p => p[xKey] != null && p[yKey] != null);
  const pts = stocks.map(p => {
    const outX = (bmx != null) ? (p[xKey] > bmx) : true;
    const outY = (bmy != null) ? (p[yKey] > bmy) : true;
    const color = (outX && outY) ? '#d23b34'
                : (!outX && !outY) ? '#1565c0'
                : '#9ca3af';
    const r = Math.max(3, Math.min(14, Math.sqrt((p.weight || 0.5)) * 2.5));
    return { x: p[xKey], y: p[yKey], r, backgroundColor: color, _meta: p };
  });
  // 벤치마크 참조선 (annotation 없이) — dataset 2개 추가로 십자선
  const allX = pts.map(pt => pt.x).concat(bmx != null ? [bmx] : []);
  const allY = pts.map(pt => pt.y).concat(bmy != null ? [bmy] : []);
  const xMin = Math.min(...allX), xMax = Math.max(...allX);
  const yMin = Math.min(...allY), yMax = Math.max(...allY);
  const xPad = Math.max(2, (xMax - xMin) * 0.08);
  const yPad = Math.max(2, (yMax - yMin) * 0.08);
  const datasets = [{
    type: 'scatter',
    label: '종목',
    data: pts,
    pointRadius: ctx => ctx.raw?.r || 4,
    pointHoverRadius: ctx => (ctx.raw?.r || 4) + 2,
    backgroundColor: ctx => ctx.raw?.backgroundColor || '#9ca3af',
    borderColor: 'rgba(0,0,0,0.15)', borderWidth: 0.5,
  }];
  if (bmx != null && bmy != null) {
    // 벤치마크 크로스헤어 (수평선 + 수직선)
    datasets.push({
      type: 'line', label: 'bm-h', showLine: true, borderColor: '#8894a6',
      borderWidth: 1, borderDash: [4, 4], pointRadius: 0, fill: false,
      data: [{x: xMin - xPad, y: bmy}, {x: xMax + xPad, y: bmy}]
    });
    datasets.push({
      type: 'line', label: 'bm-v', showLine: true, borderColor: '#8894a6',
      borderWidth: 1, borderDash: [4, 4], pointRadius: 0, fill: false,
      data: [{x: bmx, y: yMin - yPad}, {x: bmx, y: yMax + yPad}]
    });
    // 벤치마크 점 (마름모)
    datasets.push({
      type: 'scatter', label: '벤치마크',
      data: [{x: bmx, y: bmy}],
      pointStyle: 'rectRot', pointRadius: 9, pointHoverRadius: 11,
      backgroundColor: '#111827', borderColor: '#111827',
    });
  }
  return new Chart(document.getElementById(canvasId), {
    data: { datasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: (c) => {
              if (c.dataset.label === '종목') {
                const p = c.raw._meta;
                return `${p.name} (${p.code}) · ${xKey.toUpperCase()}=${p[xKey]}% / ${yKey.toUpperCase()}=${p[yKey]}% · 편입 ${p.weight}%`;
              }
              if (c.dataset.label === '벤치마크') {
                return `${bm.name}: X=${bmx?.toFixed(2)}% / Y=${bmy?.toFixed(2)}%`;
              }
              return '';
            },
            title: () => ''
          },
          filter: (item) => item.dataset.label === '종목' || item.dataset.label === '벤치마크'
        }
      },
      scales: {
        x: { title: { display: true, text: AXIS_LABEL[xKey] },
             grid: { color: '#eef1f6' },
             ticks: { callback: v => v + '%' } },
        y: { title: { display: true, text: AXIS_LABEL[yKey] },
             grid: { color: '#eef1f6' },
             ticks: { callback: v => v + '%' } }
      }
    }
  });
};

// 다크모드 — 페이지 로드 시 저장된 테마 복원
(function(){
  const saved = localStorage.getItem('anda-theme') || 'light';
  if (saved === 'dark') document.body.setAttribute('data-theme', 'dark');
})();

function _applyChartTheme(){
  const isDark = document.body.getAttribute('data-theme') === 'dark';
  Chart.defaults.color = isDark ? '#9aa6b6' : '#7a8696';
  Chart.defaults.borderColor = isDark ? '#2a3645' : '#eef1f6';
}

(function(){
  const btn = document.getElementById('themeToggle');
  function syncIcon(){
    const isDark = document.body.getAttribute('data-theme') === 'dark';
    if (btn) btn.textContent = isDark ? '☀️' : '🌙';
  }
  syncIcon();
  if (btn) {
    btn.addEventListener('click', () => {
      const next = document.body.getAttribute('data-theme') === 'dark' ? 'light' : 'dark';
      if (next === 'dark') document.body.setAttribute('data-theme', 'dark');
      else document.body.removeAttribute('data-theme');
      localStorage.setItem('anda-theme', next);
      syncIcon();
      _applyChartTheme();
      // 페이지별 JS 가 theme-changed 이벤트로 매트릭스/차트 등 즉시 다시 그릴 수 있음
      window.dispatchEvent(new Event('theme-changed'));
    });
  }
})();

Chart.defaults.font.family = "Pretendard, sans-serif";
_applyChartTheme();
const PIE = ['#16314f','#234e7d','#3d6ea5','#5b91c4','#86b3d9','#b0d0e8','#c9a227','#d98c5f','#9a6fb0','#6a9a6a','#b5546a'];

function doughnut(id, obj){
  new Chart(document.getElementById(id), {
    type:'doughnut',
    data:{ labels:obj.labels, datasets:[{ data:obj.values, backgroundColor:PIE, borderColor:'#fff', borderWidth:2 }] },
    options:{ responsive:true, maintainAspectRatio:true, cutout:'58%',
      plugins:{ legend:{ position:'right', labels:{boxWidth:11, font:{size:12}, padding:8} },
        tooltip:{ callbacks:{ label:(c)=> c.label+': '+c.parsed.toFixed(2)+'%' } } } }
  });
}
"""

FUND_JS = """
const C = __CHART__;
const SIZES = __SIZES__;
const KCMP = __KCMP__;
const SCATTER = __SCATTER__;

new Chart(document.getElementById('retChart'), {
  data:{ labels:C.dates, datasets:[
    { type:'bar', label:'초과수익률', data:C.excess, yAxisID:'yR',
      backgroundColor:'rgba(226,87,76,.5)', borderWidth:0, order:3, barPercentage:1, categoryPercentage:1 },
    { type:'line', label:'펀드수익률', data:C.fund, yAxisID:'yL',
      borderColor:'#2e7d32', backgroundColor:'#2e7d32', borderWidth:1.6, pointRadius:0, tension:.15, order:1 },
    { type:'line', label:'BM수익률', data:C.bm, yAxisID:'yL',
      borderColor:'#2962ff', backgroundColor:'#2962ff', borderWidth:1.6, pointRadius:0, tension:.15, order:2 },
  ]},
  options:{ responsive:true, maintainAspectRatio:true, interaction:{mode:'index',intersect:false},
    plugins:{ legend:{display:false}, tooltip:{ callbacks:{ label:(c)=> c.dataset.label+': '+c.parsed.y.toFixed(2)+'%' } } },
    scales:{ x:{ grid:{display:false}, ticks:{maxTicksLimit:16, autoSkip:true, maxRotation:0} },
      yL:{ position:'left', ticks:{callback:v=>v+'%'}, grid:{color:'#eef1f6'} },
      yR:{ position:'right', suggestedMax:0, ticks:{callback:v=>v+'%'}, grid:{display:false} } }
  }
});

// 삼중 도넛: 안쪽=펀드, 중간=전사, 바깥=KOSPI
// 통일된 라벨 순서 + 라벨별 색 매핑으로 동심원 비교가 정확하도록 한다.
(function nestedSizeDoughnut(){
  const LABEL_ORDER = ['대형주', '중형주', '소형주', '미분류'];
  const COLOR_MAP = {
    '대형주':  '#16314f',
    '중형주':  '#5b91c4',
    '소형주':  '#c9a227',
    '미분류':  '#c3cad4'
  };

  // {labels, values} → LABEL_ORDER 순서의 values 배열
  function normalize(obj) {
    if (!obj || !obj.labels) return LABEL_ORDER.map(() => 0);
    const m = {};
    obj.labels.forEach((l, i) => { m[l] = (obj.values || [])[i] || 0; });
    return LABEL_ORDER.map(l => m[l] || 0);
  }

  const fundVals = normalize(SIZES.fund);
  const firmVals = normalize(SIZES.firm);
  const kospiVals = normalize(SIZES.kospi);
  const bg = LABEL_ORDER.map(l => COLOR_MAP[l]);

  const ctx = document.getElementById('sizeNested');
  if (!ctx) return;

  new Chart(ctx, {
    type: 'doughnut',
    data: {
      labels: LABEL_ORDER,
      datasets: [
        // Chart.js 다중 dataset 도넛: 배열 순서 = 안쪽 → 바깥
        { label: '펀드',  data: fundVals,  backgroundColor: bg, borderColor:'#fff', borderWidth:2 },
        { label: '전사',  data: firmVals,  backgroundColor: bg, borderColor:'#fff', borderWidth:2 },
        { label: 'KOSPI', data: kospiVals, backgroundColor: bg, borderColor:'#fff', borderWidth:2 }
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: true,
      cutout: '35%',
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            title: (items) => {
              if (!items.length) return '';
              return items[0].dataset.label + ' — ' + items[0].label;
            },
            label: (c) => c.parsed.toFixed(2) + '%'
          }
        }
      }
    }
  });

  // 우측 상세 표 — 같은 데이터를 숫자로 한 번 더
  const tbl = document.getElementById('nsTable');
  if (tbl) {
    const fmt = v => v.toFixed(1);
    const rows = LABEL_ORDER.map((lab, i) =>
      `<tr><td><i class="ns-sw" style="background:${COLOR_MAP[lab]};display:inline-block;` +
      `width:10px;height:10px;border-radius:2px;margin-right:5px;vertical-align:-1px"></i>${lab}</td>` +
      `<td>${fmt(fundVals[i])}</td><td>${fmt(firmVals[i])}</td><td>${fmt(kospiVals[i])}</td></tr>`
    ).join('');
    tbl.innerHTML =
      '<table><thead><tr><th>구분</th><th>펀드</th><th>전사</th><th>KOSPI</th></tr></thead>' +
      `<tbody>${rows}</tbody></table>`;
  }
})();

// KOSPI vs 펀드 — WICS 대/중/소 + 1위 제외 토글. KCMP_WICS = {big, mid, small}.
const KCMP_WICS = __KCMP_WICS__;
let _cmpLevel = 'big';
let _cmpFocus = 'all';  // 'all' | 'ex1'
let _cmpChart = null;
// KOSPI/펀드 둘 중 더 큰 값을 가진 라벨 1개를 제외하고 라벨/값 배열 반환
function _filterTop(data){
  if (_cmpFocus !== 'ex1' || !data.labels.length) return data;
  let mi = 0, mv = -Infinity;
  for (let i = 0; i < data.labels.length; i++) {
    const v = Math.max(data.kospi[i] || 0, data.firm[i] || 0);
    if (v > mv) { mv = v; mi = i; }
  }
  return {
    labels: data.labels.filter((_, i) => i !== mi),
    kospi:  data.kospi.filter((_, i) => i !== mi),
    firm:   data.firm.filter((_, i) => i !== mi),
    excluded: data.labels[mi]
  };
}
function drawCmp(){
  const raw = (KCMP_WICS && KCMP_WICS[_cmpLevel]) || KCMP;
  const data = _filterTop(raw);
  if (_cmpChart) { _cmpChart.destroy(); _cmpChart = null; }
  _cmpChart = new Chart(document.getElementById('sectorCmp'), {
    type:'bar',
    data:{ labels: data.labels, datasets:[
      { label:'KOSPI', data: data.kospi, backgroundColor:'#16314f', borderWidth:0 },
      { label:'펀드',  data: data.firm,  backgroundColor:'#c9a227', borderWidth:0 },
    ]},
    options:{ responsive:true, maintainAspectRatio:true,
      plugins:{ legend:{display:false},
        title: data.excluded ? { display:true, text:`(제외됨: ${data.excluded})`,
          color:'#7c8b9c', font:{size:11,weight:'normal'}, padding:{top:0,bottom:6} } : { display:false },
        tooltip:{ callbacks:{ label:c => c.dataset.label+': '+c.parsed.y.toFixed(2)+'%' } } },
      scales:{
        x:{ grid:{display:false}, ticks:{ font:{size:11}, maxRotation:60, minRotation:0,
            autoSkip: false } },
        y:{ grid:{color:'#eef1f6'}, ticks:{ callback:v=>v+'%' } } } }
  });
}
document.querySelectorAll('#cmpLevelSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#cmpLevelSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    _cmpLevel = b.dataset.lv;
    drawCmp();
  });
});
document.querySelectorAll('#cmpFocusSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#cmpFocusSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    _cmpFocus = b.dataset.f;
    drawCmp();
  });
});
drawCmp();

if (SCATTER.length) {
  const PRESETS = {
    rr: {x:'sd', y:'sharpe', xl:'표준편차 SD (%)', yl:'Sharpe (위험조정 성과)', yref:0},
    ba: {x:'beta', y:'jensen', xl:'베타 (BM대비)', yl:'Jensen α', xref:1, yref:0},
    ar: {x:'te', y:'ir', xl:'TE (활성위험)', yl:'IR (정보비율)', yref:0},
  };
  const NAVY='#16314f', GRAY='#b9c4d2';
  const labelPlugin = { id:'labels', afterDatasetsDraw(chart){
      const ctx = chart.ctx, meta = chart.getDatasetMeta(0);
      ctx.save(); ctx.font='600 11px Pretendard, sans-serif';
      meta.data.forEach((pt,i)=>{ const d = chart.data.datasets[0].data[i];
        if(d.x==null || d.y==null) return;
        ctx.fillStyle = d.active ? NAVY : '#8290a3'; ctx.fillText(d.label, pt.x+9, pt.y+4); });
      ctx.restore(); } };
  const refPlugin = { id:'refs', beforeDatasetsDraw(chart){
      const p = chart.$preset, ctx = chart.ctx, a = chart.chartArea, s = chart.scales;
      ctx.save(); ctx.strokeStyle='#d7dee8'; ctx.setLineDash([4,4]); ctx.lineWidth=1;
      if(p.yref!=null){ const y=s.y.getPixelForValue(p.yref);
        ctx.beginPath(); ctx.moveTo(a.left,y); ctx.lineTo(a.right,y); ctx.stroke(); }
      if(p.xref!=null){ const x=s.x.getPixelForValue(p.xref);
        ctx.beginPath(); ctx.moveTo(x,a.top); ctx.lineTo(x,a.bottom); ctx.stroke(); }
      ctx.restore(); } };
  const pts = p => SCATTER.map(d=>({x:d[p.x], y:d[p.y], label:d.label, name:d.name, active:d.active}));
  let preset = PRESETS.rr;
  const sc = new Chart(document.getElementById('scatter'), {
    type:'scatter',
    data:{ datasets:[{ data:pts(preset),
      pointBackgroundColor:c=>c.raw&&c.raw.active?NAVY:GRAY, pointBorderColor:'#fff', pointBorderWidth:1.5,
      pointRadius:c=>c.raw&&c.raw.active?8:5, pointHoverRadius:c=>c.raw&&c.raw.active?10:7 }] },
    options:{ responsive:true, maintainAspectRatio:true, layout:{padding:{right:64, top:6}},
      plugins:{ legend:{display:false},
        tooltip:{ callbacks:{ title:i=>i[0].raw.name,
          label:c=>preset.xl+': '+c.parsed.x+'  ·  '+preset.yl+': '+c.parsed.y }}},
      scales:{
        x:{ title:{display:true, text:preset.xl, color:'#7a8696', font:{size:12}},
            grid:{color:'#eef1f6'}, grace:'12%' },
        y:{ title:{display:true, text:preset.yl, color:'#7a8696', font:{size:12}},
            grid:{color:'#eef1f6'}, grace:'25%' } } },
    plugins:[refPlugin, labelPlugin]
  });
  sc.$preset = preset;
  document.querySelectorAll('#seg button').forEach(b=>{
    b.addEventListener('click', ()=>{
      document.querySelectorAll('#seg button').forEach(x=>x.classList.remove('on'));
      b.classList.add('on'); preset = PRESETS[b.dataset.p]; sc.$preset = preset;
      sc.data.datasets[0].data = pts(preset);
      sc.options.scales.x.title.text = preset.xl; sc.options.scales.y.title.text = preset.yl;
      sc.update(); }); });
}
"""

FIRM_JS = """
const FD = __FD__;
let _view = 'all';
let _mode = 'stock';

function wBg(w){
  const k = Math.min(1.0, w / 18.0);
  const isDark = document.body.getAttribute('data-theme') === 'dark';
  if (isDark) {
    // 다크 배경에선 골드(번호석) 톤으로 — 진할수록 더 선명
    const a = (0.22 + 0.65 * k).toFixed(2);
    const fg = k > 0.35 ? '#0e1620' : '#f2c94c';
    return `background:rgba(201,162,39,${a});color:${fg}`;
  }
  const a = (0.10 + 0.62 * k).toFixed(2);
  const fg = k > 0.45 ? '#fff' : '#16202e';
  return `background:rgba(22,49,79,${a});color:${fg}`;
}
function f2(x){ return Number(x).toLocaleString('ko-KR',{minimumFractionDigits:2,maximumFractionDigits:2}); }
function f1(x){ return Number(x).toLocaleString('ko-KR',{minimumFractionDigits:1,maximumFractionDigits:1}); }
function f0(x){ return Number(x).toLocaleString('ko-KR'); }

function renderStats(){
  const v = FD[_view];
  const suffix = (_view === 'ex_tmf' && FD.tmf_funds.length)
    ? ` (TMF 제외: ${FD.tmf_funds.join(', ')})` : '';
  document.getElementById('firmSub').textContent =
    `${v.matrix.funds.length}개 펀드 합산 · 종목 ${v.stock_count}개${suffix}`;
  document.getElementById('kFirmTotal').innerHTML = `${f1(v.firm_total)}<small>억</small>`;
  document.getElementById('kFundCount').innerHTML = `${v.matrix.funds.length}<small>개</small>`;
  document.getElementById('kStockCount').innerHTML = `${v.stock_count}<small>종목</small>`;
}

function renderMatrix(){
  const m = FD[_view].matrix;
  const funds = m.funds;
  const isStock = (_mode === 'stock');
  let head = '<tr>';
  head += isStock
    ? '<th>#</th><th>코드</th><th class="l">종목명</th><th>종가</th>'
    : '<th>#</th><th class="l">섹터</th>';
  for (const fn of funds) head += `<th>${fn}</th>`;
  head += '<th>전사비중</th></tr>';
  document.getElementById('mtxHead').innerHTML = head;
  const rows = isStock ? m.stock_rows : m.sector_rows;
  const colSpan = (isStock ? 4 : 2) + funds.length + 1;  // 모든 컬럼 수 (자식 row span 용)
  let body = '';
  rows.forEach((r, i) => {
    let cells = '';
    for (const fn of funds) {
      const w = r.weights[fn] || 0;
      cells += w
        ? `<td class="wt" style="${wBg(w)}">${f2(w)}</td>`
        : '<td class="wt" style="color:#cfd6df">·</td>';
    }
    if (isStock) {
      body +=
        `<tr><td class="c">${i+1}</td><td class="c">${r.code}</td>` +
        `<td class="nm">${r.name}</td><td class="cl">${f0(r.close)}</td>` +
        `${cells}<td class="firm">${f2(r.firm)}</td></tr>`;
    } else {
      const n = (r.stocks || []).length;
      body +=
        `<tr class="sec-h" data-sec="${i}">` +
        `<td class="c">${i+1}</td>` +
        `<td class="nm"><span class="arr">▶</span>${r.name}` +
        `<span class="cnt">${n}</span></td>` +
        `${cells}<td class="firm">${f2(r.firm)}</td></tr>`;
      // 자식 종목 행 (펼침 상태에서만 display)
      (r.stocks || []).forEach((sr, j) => {
        let childCells = '';
        for (const fn of funds) {
          const w = sr.weights[fn] || 0;
          childCells += w
            ? `<td class="wt" style="${wBg(w)}">${f2(w)}</td>`
            : '<td class="wt" style="color:#cfd6df">·</td>';
        }
        body +=
          `<tr class="sec-sub" data-parent="${i}" style="display:none">` +
          `<td class="c sub-idx">${j+1}</td>` +
          `<td class="nm sub-nm">` +
          `<a href="/stock/${sr.code}" style="color:inherit;text-decoration:none">` +
          `<span class="code">${sr.code}</span> ${sr.name}</a></td>` +
          `${childCells}<td class="firm">${f2(sr.firm)}</td></tr>`;
      });
    }
  });
  document.getElementById('mtxBody').innerHTML = body;

  // 섹터 행 클릭 → 자식 펼치기/접기
  document.querySelectorAll('#mtxBody .sec-h').forEach(row => {
    row.addEventListener('click', () => {
      const idx = row.dataset.sec;
      const arrow = row.querySelector('.arr');
      const isOpen = row.classList.toggle('open');
      arrow.textContent = isOpen ? '▼' : '▶';
      document.querySelectorAll(`#mtxBody .sec-sub[data-parent="${idx}"]`)
        .forEach(s => { s.style.display = isOpen ? '' : 'none'; });
    });
  });
}


let _cmpChart = null;
let _cmpLevel = 'big';   // 'big' | 'mid' | 'small'
let _cmpFocus = 'all';   // 'all' | 'ex1'
function _filterTop(data){
  if (_cmpFocus !== 'ex1' || !data.labels.length) return data;
  let mi = 0, mv = -Infinity;
  for (let i = 0; i < data.labels.length; i++) {
    const v = Math.max(data.kospi[i] || 0, data.firm[i] || 0);
    if (v > mv) { mv = v; mi = i; }
  }
  return {
    labels: data.labels.filter((_, i) => i !== mi),
    kospi:  data.kospi.filter((_, i) => i !== mi),
    firm:   data.firm.filter((_, i) => i !== mi),
    excluded: data.labels[mi]
  };
}
function renderCmp(){
  const wicsMap = (_view === 'ex_tmf') ? FD.kospi_cmp_wics_ex_tmf : FD.kospi_cmp_wics_all;
  const legacy = (_view === 'ex_tmf') ? FD.kospi_cmp_ex_tmf : FD.kospi_cmp_all;
  const raw = (wicsMap && wicsMap[_cmpLevel]) || legacy;
  const data = _filterTop(raw);
  if (_cmpChart) { _cmpChart.destroy(); _cmpChart = null; }
  _cmpChart = new Chart(document.getElementById('sectorCmp'), {
    type: 'bar',
    data: { labels: data.labels, datasets: [
      { label: 'KOSPI', data: data.kospi, backgroundColor:'#16314f', borderWidth:0 },
      { label: '전사', data: data.firm, backgroundColor:'#c9a227', borderWidth:0 },
    ]},
    options: { responsive:true, maintainAspectRatio:true,
      plugins:{ legend:{display:false},
        title: data.excluded ? { display:true, text:`(제외됨: ${data.excluded})`,
          color:'#7c8b9c', font:{size:11,weight:'normal'}, padding:{top:0,bottom:6} } : { display:false },
        tooltip:{ callbacks:{ label:c=>c.dataset.label+': '+c.parsed.y.toFixed(2)+'%' } } },
      scales:{
        x:{ grid:{display:false}, ticks:{ font:{size:11}, maxRotation:60, minRotation:0,
            autoSkip:false } },
        y:{ grid:{color:'#eef1f6'}, ticks:{ callback:v=>v+'%' } } } }
  });
}

function rerender(){ renderStats(); renderMatrix(); renderCmp(); }

document.querySelectorAll('#viewSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#viewSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _view = b.dataset.v; rerender();
  });
});
document.querySelectorAll('#mtxSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#mtxSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _mode = b.dataset.m; renderMatrix();
  });
});
// WICS 대/중/소 토글 — KOSPI vs 전사 차트 다시 그리기
document.querySelectorAll('#cmpLevelSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#cmpLevelSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _cmpLevel = b.dataset.lv; renderCmp();
  });
});
// 전체 / 1위 제외 토글
document.querySelectorAll('#cmpFocusSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#cmpFocusSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _cmpFocus = b.dataset.f; renderCmp();
  });
});

// 다크모드 토글 시 매트릭스 + 비교 차트 즉시 다시 그리기
window.addEventListener('theme-changed', () => { renderMatrix(); renderCmp(); });

rerender();
"""

SEARCH_JS = """
(function(){
  const inp = document.getElementById('stockSearch');
  const box = document.getElementById('searchResults');
  if(!inp || !box) return;
  function render(q){
    if(!q){ box.style.display='none'; return; }
    const ql = q.toLowerCase();
    const hits = (STOCKS||[]).filter(s =>
      s.name.toLowerCase().includes(ql) || s.code.includes(q)
    ).slice(0, 10);
    if(!hits.length){
      box.innerHTML = '<a style="cursor:default;color:#7a8696"><b>일치하는 종목 없음</b><span>엑셀에 등록된 10개 종목만 검색 가능</span></a>';
      box.style.display='block';
      return;
    }
    box.innerHTML = hits.map(s =>
      `<a href="/stock/${s.code}"><b>${s.name}</b><span>${s.code}</span></a>`
    ).join('');
    box.style.display='block';
  }
  inp.addEventListener('input', e=>render(e.target.value.trim()));
  inp.addEventListener('focus', e=>{ if(e.target.value.trim()) render(e.target.value.trim()); });
  inp.addEventListener('blur', ()=>setTimeout(()=>{ box.style.display='none'; }, 180));
  inp.addEventListener('keydown', e=>{
    if(e.key==='Enter'){
      const ql = inp.value.trim().toLowerCase();
      const hit = (STOCKS||[]).find(s => s.name.toLowerCase()===ql || s.code===inp.value.trim());
      if(hit) location.href = '/stock/'+hit.code;
    } else if(e.key==='Escape'){ box.style.display='none'; }
  });
})();
"""

KOSDAQ_JS = """
const SECTORS = __SECTORS__;
if (SECTORS && SECTORS.length) {
  new Chart(document.getElementById('sectorChart'), {
    type:'doughnut',
    data:{ labels: SECTORS.map(s=>s.name),
      datasets:[{ data: SECTORS.map(s=>s.pct), backgroundColor: PIE,
                  borderColor:'#fff', borderWidth:2 }] },
    options:{ responsive:true, maintainAspectRatio:true, cutout:'58%',
      plugins:{ legend:{ position:'right', labels:{boxWidth:11, font:{size:12}, padding:8} },
        tooltip:{ callbacks:{ label:(c)=> c.label+': '+c.parsed.toFixed(2)+'%' } } } }
  });
}
"""


def _wbg(w):
    """편입비(%) -> 히트맵 칸 배경 스타일."""
    k = min(1.0, w / 18.0)
    a = 0.10 + 0.62 * k
    fg = "#ffffff" if k > 0.45 else "#16202e"
    return f"background:rgba(22,49,79,{a:.2f});color:{fg}"


def _nav(active, fund_names):
    mc = "tab market active" if active == "__market__" else "tab market"
    html = f'<a class="{mc}" href="/market">시장</a>'
    rc = "tab research active" if active == "__research__" else "tab research"
    html += f'<a class="{rc}" href="/research">리서치</a>'
    fc = "tab firm active" if active == "__firm__" else "tab firm"
    html += f'<a class="{fc}" href="/firm">전사 현황</a>'
    ac = "tab active" if active == "__alpha__" else "tab"
    html += f'<a class="{ac}" href="/alpha">알파 산점도</a>'
    for fn in fund_names:
        c = "tab active" if fn == active else "tab"
        html += f'<a class="{c}" href="/fund/{fn}">{fn}</a>'
    return html


def page_shell(title, active, fund_names, body, scripts):
    nav = _nav(active, fund_names)
    stocks = [{"code": v["code"], "name": v["name"]} for v in get_stock_master().values()]
    stocks_json = json.dumps(stocks, ensure_ascii=False)

    # scripts 인자에 <style>...</style> 가 섞여 있으면 head 로 분리
    # (script 태그 안에 style 태그가 들어가면 JS 파싱이 중단됨)
    import re as _re
    style_extra = ""
    js_part = scripts
    m = _re.match(r'^(<style>.*?</style>)(.*)$', scripts, _re.DOTALL)
    if m:
        style_extra = m.group(1)
        js_part = m.group(2)

    return (
        "<!doctype html><html lang='ko'><head><meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width, initial-scale=1'>"
        f"<title>{title}</title>"
        "<link rel='stylesheet' href='https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css'>"
        "<script src='https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js'></script>"
        "<script src='https://unpkg.com/lightweight-charts@4.1.3/dist/lightweight-charts.standalone.production.js'></script>"
        f"<style>{STYLE}</style>{style_extra}</head><body>"
        "<div class='topbar'>"
        "<div class='brandrow'>"
        "<div class='brand'><b>ANDA 펀드 대시보드</b>"
        "<span>안다자산운용 · 펀드 운용 현황</span></div>"
        "<div class='searchbar'>"
        "<input class='searchbox' id='stockSearch' placeholder='종목명 또는 코드 검색' autocomplete='off'>"
        "<div class='searchresults' id='searchResults'></div>"
        "</div>"
        "<button id='themeToggle' class='theme-toggle' title='다크모드 토글'>🌙</button>"
        "</div>"
        f"<div class='tabs'>{nav}</div></div>"
        f"<div class='wrap'>{body}</div>"
        f"<script>const STOCKS = {stocks_json};{COMMON_JS}{SEARCH_JS}{js_part}</script></body></html>"
    )


def render_page(active, fund_names, data):
    table_rows = ""
    for i, r in enumerate(data["rows"], 1):
        pcls = "pos" if r["손익률"] >= 0 else "neg"
        sign = "+" if r["손익률"] >= 0 else ""
        table_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="name">{r["종목명"]}</td>'
            f'<td class="num muted">{r["종목코드"]}</td>'
            f'<td class="num strong">{r["편입비"]:.2f}</td>'
            f'<td class="num">{r["평가액억"]:,.1f}</td>'
            f'<td class="num {pcls}">{sign}{r["손익률"]:.2f}</td>'
            f'<td class="num">{r["종가"]:,}</td>'
            f'<td class="num muted">{r["보유수량"]:,}</td>'
            f'<td class="tag">{r["업종명"]}</td>'
            f'<td class="tag size">{r["규모구분"]}</td></tr>')

    excess_cls = "pos" if data["stat_excess"] >= 0 else "neg"
    excess_sign = "+" if data["stat_excess"] >= 0 else ""

    risk_card = ""
    if data.get("risk"):
        rg = ""
        for sub, items in data["risk"]["groups"]:
            tiles = "".join(
                f'<div class="tile"><div class="lab">{lab}</div>'
                f'<div class="val">{val}</div></div>' for lab, val in items)
            rg += f'<div class="rgroup"><p class="sub">{sub}</p><div class="rgrid">{tiles}</div></div>'
        risk_card = f'<div class="card"><h2>리스크 지표</h2>{rg}</div>'

    scatter_card = ""
    if data.get("risk_scatter"):
        scatter_card = (
            '<div class="card"><div class="h2row"><h2>리스크 좌표 (펀드 비교)</h2>'
            '<div class="seg" id="seg"><button data-p="rr" class="on">위험·성과</button>'
            '<button data-p="ba">베타·알파</button><button data-p="ar">활성위험</button></div></div>'
            '<canvas id="scatter" height="150"></canvas></div>')

    subtitle = ""
    if data.get("official_name"):
        subtitle = f'<div class="subname">{data["official_name"]}</div>'

    body = (
        f'<div class="head"><div><h1>{active}</h1>{subtitle}</div>'
        f'<div class="asof">기준일 {data["as_of"]}</div></div>'
        '<div class="stats">'
        f'<div class="stat"><div class="k">펀드수익률</div><div class="v">{data["stat_fund"]:,.2f}<small>%</small></div></div>'
        f'<div class="stat"><div class="k">BM수익률</div><div class="v">{data["stat_bm"]:,.2f}<small>%</small></div></div>'
        f'<div class="stat"><div class="k">초과수익률</div><div class="v {excess_cls}">{excess_sign}{data["stat_excess"]:,.2f}<small>%</small></div></div>'
        f'<div class="stat"><div class="k">총 평가액</div><div class="v">{data["total_eval"]:,.1f}<small>억</small></div></div>'
        f'<div class="stat"><div class="k">보유종목 / 주식편입</div><div class="v">{data["stock_count"]}<small>종목 · {data["equity_ratio"]:.1f}%</small></div></div>'
        '</div>'
        '<div class="card"><h2>누적 수익률 추이</h2><canvas id="retChart" height="105"></canvas>'
        '<div class="legend"><span><i style="background:var(--red)"></i>초과수익률(%, 우)</span>'
        '<span><i style="background:var(--green)"></i>펀드수익률(%, 좌)</span>'
        '<span><i style="background:var(--blue)"></i>BM수익률(%, 좌)</span></div></div>'
        f'{scatter_card}{risk_card}'
        '<div class="card"><div class="h2row"><h2>KOSPI 섹터 비중 vs 펀드 섹터 비중 '
        '<span class="src-tag">WICS 기준</span></h2>'
        '<div class="seg" id="cmpLevelSeg" style="margin-left:auto">'
        '<button data-lv="big" class="on">대분류</button>'
        '<button data-lv="mid">중분류</button>'
        '<button data-lv="small">소분류</button>'
        '</div>'
        '<div class="seg" id="cmpFocusSeg" style="margin-left:8px">'
        '<button data-f="all" class="on">전체</button>'
        '<button data-f="ex1">1위 제외</button>'
        '</div></div>'
        '<canvas id="sectorCmp" height="120"></canvas>'
        '<div class="tmlegend"><i style="background:#16314f"></i>KOSPI'
        f'<i style="background:#c9a227"></i>{active}<span>· 단위 % · WICS 기준</span></div></div>'
        '<div class="card"><h2>규모별 비중 (삼중 도넛 비교)</h2>'
        '<div class="nested-wrap">'
        '<div class="nested-chart"><canvas id="sizeNested"></canvas></div>'
        '<div class="nested-side">'
        '<div class="ns-ringkey">'
        '<div class="ns-title">링 (안→밖)</div>'
        f'<div class="ns-row"><i class="ns-dot ring-fund"></i>안쪽 · {active}</div>'
        '<div class="ns-row"><i class="ns-dot ring-firm"></i>중간 · 전사</div>'
        '<div class="ns-row"><i class="ns-dot ring-kospi"></i>바깥 · KOSPI</div>'
        '</div>'
        '<div class="ns-colorkey">'
        '<div class="ns-title">색상</div>'
        '<div class="ns-row"><i class="ns-sw" style="background:#16314f"></i>대형주</div>'
        '<div class="ns-row"><i class="ns-sw" style="background:#5b91c4"></i>중형주</div>'
        '<div class="ns-row"><i class="ns-sw" style="background:#c9a227"></i>소형주</div>'
        '<div class="ns-row"><i class="ns-sw" style="background:#c3cad4"></i>미분류</div>'
        '</div>'
        '<div class="ns-table" id="nsTable"></div>'
        '</div>'
        '</div></div>'
        f'<div class="card"><h2>보유종목 리스트 ({data["stock_count"]}종목) '
        '<span class="src-tag">업종은 펀드 엑셀 원본, 규모는 KRX 기준</span></h2>'
        '<div style="overflow-x:auto;"><table><thead><tr>'
        '<th>#</th><th>종목명</th><th>코드</th><th>편입비(%)</th><th>평가액(억)</th>'
        '<th>평가손익률(%)</th><th>종가(원)</th><th>보유수량</th><th>업종</th><th>규모</th>'
        f'</tr></thead><tbody>{table_rows}</tbody></table></div></div>')

    sizes_payload = {
        "fund": data["size"],
        "firm": data.get("firm_size") or {"labels": [], "values": []},
        "kospi": data.get("kospi_size") or {"labels": [], "values": []},
    }
    scripts = (FUND_JS
               .replace("__CHART__", json.dumps(data["chart"], ensure_ascii=False))
               .replace("__SIZES__", json.dumps(sizes_payload, ensure_ascii=False))
               .replace("__KCMP__", json.dumps(data["kospi_cmp"], ensure_ascii=False))
               .replace("__KCMP_WICS__", json.dumps(data.get("kospi_cmp_wics") or {}, ensure_ascii=False))
               .replace("__SCATTER__", json.dumps(data.get("risk_scatter") or [], ensure_ascii=False)))

    return page_shell(f"ANDA 대시보드 · {active}", active, fund_names, body, scripts)


def render_firm(fund_names, fd):
    has_tmf = bool(fd.get("tmf_funds"))
    # 토글 UI는 TMF 펀드가 1개 이상 있을 때만 노출
    view_toggle = ""
    if has_tmf:
        view_toggle = (
            '<div class="seg" id="viewSeg">'
            '<button data-v="all" class="on">전체</button>'
            '<button data-v="ex_tmf">TMF 제외</button>'
            '</div>')

    body = (
        '<div class="head"><div><h1>전사 현황</h1>'
        '<div class="subname" id="firmSub"></div></div>'
        f'<div class="asof">기준일 {fd["as_of"]}</div></div>'

        # 뷰 토글 (TMF 포함/제외)
        f'<div class="topctl">{view_toggle}</div>'

        # 통계 카드 (JS 로 동적 업데이트)
        '<div class="stats s4">'
        '<div class="stat"><div class="k">전사 총 평가액</div><div class="v" id="kFirmTotal">—</div></div>'
        '<div class="stat"><div class="k">펀드 수</div><div class="v" id="kFundCount">—</div></div>'
        '<div class="stat"><div class="k">보유종목 수(중복제외)</div><div class="v" id="kStockCount">—</div></div>'
        f'<div class="stat"><div class="k">기준일</div><div class="v" style="font-size:18px">{fd["as_of"]}</div></div>'
        '</div>'

        # 매트릭스 (종목/섹터 토글)
        '<div class="card"><div class="h2row"><h2>비중 매트릭스 '
        '<span class="src-tag">순자산비 · WICS 대분류</span></h2>'
        '<div class="seg" id="mtxSeg" style="margin-left:auto">'
        '<button data-m="stock" class="on">종목별</button>'
        '<button data-m="sector">섹터별</button>'
        '</div></div>'
        '<div class="mtxbox"><table class="mtx" id="mtxTbl"><thead id="mtxHead"></thead><tbody id="mtxBody"></tbody></table></div>'
        '<div class="tmlegend"><span>칸 색이 진할수록 순자산비(NAV 대비) 높음 · 단위 %</span></div></div>'

        # KOSPI 섹터 비중 vs 전사 섹터 비중
        '<div class="card"><div class="h2row"><h2>KOSPI 섹터 비중 vs 전사 섹터 비중 '
        '<span class="src-tag">WICS 기준</span></h2>'
        '<div class="seg" id="cmpLevelSeg" style="margin-left:auto">'
        '<button data-lv="big" class="on">대분류</button>'
        '<button data-lv="mid">중분류</button>'
        '<button data-lv="small">소분류</button>'
        '</div>'
        '<div class="seg" id="cmpFocusSeg" style="margin-left:8px">'
        '<button data-f="all" class="on">전체</button>'
        '<button data-f="ex1">1위 제외</button>'
        '</div></div>'
        '<canvas id="sectorCmp" height="120"></canvas>'
        '<div class="tmlegend"><i style="background:#16314f"></i>KOSPI'
        '<i style="background:#c9a227"></i>전사<span>· 단위 % · WICS 기준</span></div></div>')

    payload = {
        "all": fd["all"], "ex_tmf": fd["ex_tmf"],
        "has_tmf": has_tmf, "tmf_funds": fd["tmf_funds"],
        "kospi_cmp_all": fd["kospi_cmp_all"], "kospi_cmp_ex_tmf": fd["kospi_cmp_ex_tmf"],
        "kospi_cmp_wics_all": fd.get("kospi_cmp_wics_all") or {},
        "kospi_cmp_wics_ex_tmf": fd.get("kospi_cmp_wics_ex_tmf") or {},
        "alpha_all": fd.get("alpha_all") or {},
        "alpha_ex_tmf": fd.get("alpha_ex_tmf") or {},
    }
    scripts = FIRM_JS.replace("__FD__", json.dumps(payload, ensure_ascii=False))
    return page_shell("ANDA 대시보드 · 전사 현황", "__firm__", fund_names, body, scripts)


ALPHA_CSS = """
  /* 상단 컨트롤 바 — 그룹별 박스로 명확히 구분 */
  .a-topctl{ display:flex; flex-wrap:wrap; gap:12px; align-items:stretch;
    padding:14px 16px; background:var(--card); border:1px solid var(--line);
    border-radius:12px; margin-bottom:14px; }
  .a-group{ display:flex; flex-direction:column; gap:6px;
    padding:9px 12px 10px; background:#f4f7fb; border:1px solid var(--line);
    border-radius:10px; }
  body[data-theme="dark"] .a-group{ background:rgba(255,255,255,.04); }
  .a-group-lbl{ font-size:10.5px; font-weight:800; color:var(--muted);
    letter-spacing:.6px; text-transform:uppercase; }
  .a-group .seg{ margin:0; background:#fff; border-color:var(--line); }
  body[data-theme="dark"] .a-group .seg{ background:var(--soft2); }
  .a-group select{ padding:6px 10px; border:1px solid var(--line);
    border-radius:8px; font-size:13px; font-family:inherit; background:#fff;
    color:var(--ink); cursor:pointer; min-width:120px; }
  body[data-theme="dark"] .a-group select{ background:var(--soft2); color:var(--ink); }

  /* 부제 텍스트 — 컨트롤 바 아래 별도 줄 */
  .a-sub{ font-size:12.5px; color:var(--muted); margin:0 4px 14px;
    padding:6px 12px; }
  .a-sub b{ color:var(--ink); font-weight:700; }

  /* 그래프 카드 — 세로 스택, 크게 확대 */
  .alpha-card{ padding:18px 20px; }
  .alpha-grid.big{ display:grid; grid-template-columns:1fr; gap:22px; }
  .alpha-grid.big .alpha-box{ padding:14px 16px 18px; }
  .alpha-grid.big .alpha-box .alpha-cap{ font-size:14px; font-weight:700;
    margin-bottom:10px; text-align:center; color:var(--ink); }
  .alpha-grid.big .alpha-box canvas{ height:600px !important; }
  @media (max-width:900px){
    .alpha-grid.big .alpha-box canvas{ height:480px !important; }
  }
"""


ALPHA_JS = """
const AP = __AP__;
let _mode = 'firm';       // 'firm' | 'fund'
let _firmView = 'all';    // 'all' | 'ex_tmf'
let _fundName = null;     // 'ckvf' 등
let _emphasize = 'ours';  // 'ours' | 'k200'
let _search = '';         // 검색어 (종목명·코드 부분 매칭)
let _chart1 = null, _chart2 = null;

// 강조 대상은 항상 진한 파랑, 배경은 항상 옅은 회색 — 두 강조 모드가 대칭
const COLOR_STRONG = '#1976d2';   // 강조 (진한 파랑, 시인성 높음)
const COLOR_MUTED  = '#cfd6df';   // 배경 (옅은 회색)
const COLOR_HIT    = '#f43f5e';   // 검색 매칭 (핫핑크 — 눈에 확 띄게)

function currentOurs(){
  if (_mode === 'firm') return (_firmView === 'ex_tmf') ? AP.firm_ex_tmf : AP.firm_all;
  return AP.funds[_fundName] || {bm: AP.bm, points: []};
}

function drawTwo() {
  if (_chart1) { _chart1.destroy(); _chart1 = null; }
  if (_chart2) { _chart2.destroy(); _chart2 = null; }
  const ours = currentOurs();
  const k200 = AP.kospi200;
  const strongOurs = (_emphasize === 'ours');
  const oursColor = strongOurs ? COLOR_STRONG : COLOR_MUTED;
  const k200Color = strongOurs ? COLOR_MUTED  : COLOR_STRONG;

  _chart1 = drawAlphaWithOverlay('alphaChart1', ours, k200,
                                 oursColor, k200Color, strongOurs, 'm3', 'm1');
  _chart2 = drawAlphaWithOverlay('alphaChart2', ours, k200,
                                 oursColor, k200Color, strongOurs, 'm1', 'wk1');

  // subtitle 텍스트 갱신
  const oursN = (ours.points || []).filter(p => p.m1 != null && p.m3 != null).length;
  const k200N = (k200.points || []).filter(p => p.m1 != null && p.m3 != null).length;
  const bmName = (ours.bm || AP.bm || {}).name || '—';
  let searchInfo = '';
  const q2 = (_search || '').trim().toLowerCase();
  if (q2) {
    const isHit = (p) => (p.name||'').toLowerCase().includes(q2) || (p.code||'').toLowerCase().includes(q2);
    const oh = (ours.points || []).filter(isHit).length;
    const kh = (k200.points || []).filter(isHit).length;
    searchInfo = ' · <span style="color:#f43f5e;font-weight:700">검색 매칭 우리 ' + oh + ' · K200 ' + kh + '</span>';
  }
  document.getElementById('aSub').innerHTML =
    '벤치마크 <b>' + bmName + '</b> · 우리 종목 <b>' + oursN + '</b> · KOSPI 200 <b>' +
    k200N + '</b>개 표시 (전체 ' + AP.kospi200_count + ' 중 매핑 가능한 것) · 점 크기 = 편입비' +
    searchInfo;
}

// 우리 종목 + KOSPI 200 오버레이 산점도.
// 좌표 = 종목 수익률 − 벤치마크 수익률 (즉 KOSPI 200 대비 초과수익률/알파).
// 벤치마크 자체는 항상 원점 (0,0), 크로스헤어는 x=0/y=0.
function drawAlphaWithOverlay(canvasId, ours, k200, oursColor, k200Color, oursOnTop, xKey, yKey) {
  const AXIS_LABEL = {
    wk1: '5일 초과수익률 (%p, KOSPI 200 대비)',
    m1:  '1개월 초과수익률 (%p, KOSPI 200 대비)',
    m3:  '3개월 초과수익률 (%p, KOSPI 200 대비)'
  };
  const bm = ours.bm || k200.bm || {};
  const bmx = bm[xKey], bmy = bm[yKey];
  // 벤치마크 값이 없으면 (0으로 처리 불가) 원 수익률 그대로 표시
  const bx = (bmx == null) ? 0 : bmx;
  const by = (bmy == null) ? 0 : bmy;

  function toPoints(list, radius_scale) {
    return (list.points || [])
      .filter(p => p[xKey] != null && p[yKey] != null)
      .map(p => ({
        x: +(p[xKey] - bx).toFixed(2),
        y: +(p[yKey] - by).toFixed(2),
        r: Math.max(3, Math.min(14, Math.sqrt(Math.max(p.weight || 0.5, 0.05)) * radius_scale)),
        _meta: p
      }));
  }
  const oursPts = toPoints(ours, 3.2);
  const k200Pts = toPoints(k200, 2.0);

  // 축 범위 — 점들 + (0,0) 포함 대칭 padding
  const allX = oursPts.concat(k200Pts).map(p => p.x).concat([0]);
  const allY = oursPts.concat(k200Pts).map(p => p.y).concat([0]);
  const xMin = Math.min(...allX), xMax = Math.max(...allX);
  const yMin = Math.min(...allY), yMax = Math.max(...allY);
  const xPad = Math.max(2, (xMax - xMin) * 0.06);
  const yPad = Math.max(2, (yMax - yMin) * 0.06);

  // 검색어 매칭 판정 (종목명 또는 코드 부분 매칭, 대소문자 무시)
  const q = (_search || '').trim().toLowerCase();
  function _hit(p) {
    if (!q) return false;
    const nm = (p.name || '').toLowerCase();
    const cd = (p.code || '').toLowerCase();
    return nm.includes(q) || cd.includes(q);
  }
  const oursDs = {
    type: 'scatter', label: '우리 종목', data: oursPts,
    pointRadius: ctx => {
      const r = ctx.raw?.r || 4;
      return _hit(ctx.raw?._meta) ? r + 3 : r;    // 매칭이면 더 큼
    },
    pointHoverRadius: ctx => (ctx.raw?.r || 4) + 3,
    hitRadius: 12,
    backgroundColor: ctx => _hit(ctx.raw?._meta) ? COLOR_HIT : oursColor,
    borderColor: ctx => _hit(ctx.raw?._meta) ? '#7f1d1d' : 'rgba(0,0,0,0.25)',
    borderWidth: ctx => _hit(ctx.raw?._meta) ? 2 : 0.8,
    order: oursOnTop ? 0 : 1,
  };
  const k200Ds = {
    type: 'scatter', label: 'KOSPI 200', data: k200Pts,
    pointRadius: ctx => {
      const r = ctx.raw?.r || 3;
      return _hit(ctx.raw?._meta) ? r + 3 : r;
    },
    pointHoverRadius: ctx => (ctx.raw?.r || 3) + 3,
    hitRadius: 10,
    backgroundColor: ctx => _hit(ctx.raw?._meta) ? COLOR_HIT : k200Color,
    borderColor: ctx => _hit(ctx.raw?._meta) ? '#7f1d1d' : 'rgba(0,0,0,0.15)',
    borderWidth: ctx => _hit(ctx.raw?._meta) ? 2 : 0.4,
    order: oursOnTop ? 1 : 0,
  };
  const datasets = oursOnTop ? [k200Ds, oursDs] : [oursDs, k200Ds];

  // 벤치마크 크로스헤어 (x=0, y=0) + 원점 마커
  datasets.push({ type:'line', label:'bm-h', showLine:true, borderColor:'#8894a6',
    borderWidth:1.2, borderDash:[4,4], pointRadius:0, fill:false, order:5,
    data: [{x: xMin - xPad, y: 0}, {x: xMax + xPad, y: 0}] });
  datasets.push({ type:'line', label:'bm-v', showLine:true, borderColor:'#8894a6',
    borderWidth:1.2, borderDash:[4,4], pointRadius:0, fill:false, order:5,
    data: [{x: 0, y: yMin - yPad}, {x: 0, y: yMax + yPad}] });
  datasets.push({ type:'scatter', label:'벤치마크', data:[{x: 0, y: 0}],
    pointStyle:'rectRot', pointRadius:9, pointHoverRadius:11,
    backgroundColor:'#111827', borderColor:'#111827', order: -1 });

  return new Chart(document.getElementById(canvasId), {
    data: { datasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      // 커서가 점에 정확히 얹혀 있지 않아도 근처면 툴팁 표시
      interaction: { mode: 'nearest', intersect: false, axis: 'xy' },
      plugins: {
        legend: { display: false },
        tooltip: {
          mode: 'nearest', intersect: false, axis: 'xy',
          animation: false,
          callbacks: {
          title: () => '',
          label: (c) => {
            if (c.dataset.label === '우리 종목' || c.dataset.label === 'KOSPI 200') {
              const p = c.raw._meta;
              const dx = (p[xKey] - bx).toFixed(2);
              const dy = (p[yKey] - by).toFixed(2);
              const sx = dx > 0 ? '+' + dx : dx;
              const sy = dy > 0 ? '+' + dy : dy;
              return c.dataset.label + ' · ' + p.name + ' (' + p.code + ') · 초과 X=' +
                     sx + '%p / Y=' + sy + '%p · 원 수익률 ' +
                     xKey.toUpperCase() + '=' + p[xKey] + '% / ' +
                     yKey.toUpperCase() + '=' + p[yKey] + '% · 편입 ' + p.weight + '%';
            }
            if (c.dataset.label === '벤치마크') {
              return (bm.name || '벤치마크') + ' (기준 0%p) · 원 수익률 X=' +
                     (bmx != null ? bmx.toFixed(2) : '?') + '% / Y=' +
                     (bmy != null ? bmy.toFixed(2) : '?') + '%';
            }
            return '';
          }
        }, filter: item => ['우리 종목','KOSPI 200','벤치마크'].includes(item.dataset.label) }
      },
      scales: {
        x: { title:{display:true, text: AXIS_LABEL[xKey]}, grid:{color:'#eef1f6'},
             ticks:{callback: v => (v > 0 ? '+' : '') + v + '%p'} },
        y: { title:{display:true, text: AXIS_LABEL[yKey]}, grid:{color:'#eef1f6'},
             ticks:{callback: v => (v > 0 ? '+' : '') + v + '%p'} }
      }
    }
  });
}

// 상단 컨트롤 이벤트
document.getElementById('aModeSeg').querySelectorAll('button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#aModeSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _mode = b.dataset.m;
    document.getElementById('aFirmCtl').style.display = (_mode === 'firm') ? '' : 'none';
    document.getElementById('aFundCtl').style.display = (_mode === 'fund') ? '' : 'none';
    if (_mode === 'fund' && !_fundName) {
      _fundName = AP.fund_names[0];
      document.getElementById('aFundSel').value = _fundName;
    }
    drawTwo();
  });
});
document.getElementById('aFirmSeg').querySelectorAll('button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#aFirmSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _firmView = b.dataset.v; drawTwo();
  });
});
document.getElementById('aFundSel').addEventListener('change', e => {
  _fundName = e.target.value; drawTwo();
});
document.getElementById('aEmpSeg').querySelectorAll('button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#aEmpSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _emphasize = b.dataset.e; drawTwo();
  });
});
// 종목 검색 — 타이핑 시 즉시 반영. 디바운스로 렌더 부하 줄임.
let _searchTimer = null;
const searchInput = document.getElementById('aSearch');
if (searchInput) {
  searchInput.addEventListener('input', e => {
    _search = e.target.value;
    if (_searchTimer) clearTimeout(_searchTimer);
    _searchTimer = setTimeout(drawTwo, 150);
  });
}

// 초기 렌더
drawTwo();
"""


def render_alpha(fund_names, ap):
    tmf_note = ""
    if ap.get("tmf_funds"):
        tmf_note = f' (TMF: {", ".join(ap["tmf_funds"])})'
    has_tmf = bool(ap.get("tmf_funds"))
    firm_toggle = (
        '<div class="seg" id="aFirmSeg">'
        '<button data-v="all" class="on">전체</button>'
        + ('<button data-v="ex_tmf">TMF 제외</button>' if has_tmf else '')
        + '</div>')

    fund_opts = "".join(
        f'<option value="{fn}">{fn}</option>' for fn in ap["fund_names"])

    body = (
        '<div class="head"><div><h1>알파 산점도</h1>'
        '<div class="subname">우리 종목 vs KOSPI 200 · 벤치마크 대비 성과 비교</div></div>'
        f'<div class="asof">벤치마크 {(ap.get("bm") or {}).get("name","-")} '
        f'· KOSPI 200 매핑 {ap.get("kospi200_matched",0)}/{ap.get("kospi200_count",0)}</div></div>'

        # 컨트롤 바 — 그룹별로 박스로 분리
        '<div class="a-topctl">'
        '<div class="a-group">'
        '<div class="a-group-lbl">모드</div>'
        '<div class="seg" id="aModeSeg">'
        '<button data-m="firm" class="on">전사</button>'
        '<button data-m="fund">펀드별</button>'
        '</div></div>'

        # 전사 뷰 (TMF 포함/제외) — 전사 모드일 때만 표시
        '<div class="a-group" id="aFirmCtl">'
        '<div class="a-group-lbl">뷰</div>'
        f'{firm_toggle}'
        '</div>'

        # 펀드 셀렉트 — 펀드 모드일 때만 표시
        '<div class="a-group" id="aFundCtl" style="display:none">'
        '<div class="a-group-lbl">펀드</div>'
        f'<select id="aFundSel">{fund_opts}</select>'
        '</div>'

        # 강조 대상
        '<div class="a-group">'
        '<div class="a-group-lbl">강조</div>'
        '<div class="seg" id="aEmpSeg">'
        '<button data-e="ours" class="on">우리 종목</button>'
        '<button data-e="k200">KOSPI 200</button>'
        '</div></div>'
        # 종목 검색
        '<div class="a-group">'
        '<div class="a-group-lbl">종목 검색</div>'
        '<input type="text" id="aSearch" placeholder="종목명 또는 코드"'
        ' style="padding:6px 10px;border:1px solid var(--line);border-radius:8px;'
        'font-size:13px;font-family:inherit;background:#fff;color:var(--ink);width:150px">'
        '</div>'
        '</div>'

        '<div class="a-sub" id="aSub">—</div>'

        # 두 산점도 — 크게
        '<div class="card alpha-card"><div class="alpha-grid big">'
        '<div class="alpha-box"><div class="alpha-cap">Y: 1개월 수익률 · X: 3개월 수익률</div>'
        '<canvas id="alphaChart1"></canvas></div>'
        '<div class="alpha-box"><div class="alpha-cap">Y: 5일 수익률 · X: 1개월 수익률</div>'
        '<canvas id="alphaChart2"></canvas></div>'
        '</div></div>')

    scripts = ("<style>" + ALPHA_CSS + "</style>" +
               ALPHA_JS.replace("__AP__", json.dumps(ap, ensure_ascii=False)))
    return page_shell("ANDA 대시보드 · 알파 산점도", "__alpha__", fund_names, body, scripts)


STOCK_CSS = """
  .stockhead{ display:flex; align-items:flex-end; justify-content:space-between;
    margin-bottom:18px; flex-wrap:wrap; gap:8px; }
  .stockhead .nameline{ display:flex; align-items:baseline; gap:10px; flex-wrap:wrap; }
  .stockhead h1{ font-size:24px; margin:0; font-weight:800; letter-spacing:-.4px; }
  .stockhead .code{ font-family:monospace; background:#eef1f6; padding:3px 8px;
    border-radius:6px; font-size:12px; color:var(--navy2); font-weight:600; }
  .stockhead .market{ font-size:12px; color:var(--muted); }
  .priceline{ display:flex; align-items:baseline; gap:14px; margin-top:6px; flex-wrap:wrap; }
  .priceline .px{ font-size:34px; font-weight:800; letter-spacing:-.6px; }
  .priceline .chg{ font-size:15px; font-weight:700; }
  .priceline .chg.up{ color:var(--up); } .priceline .chg.dn{ color:var(--down); }
  .stockhead .asof{ color:var(--muted); font-size:12.5px; margin-top:2px; }
  #chartbox{ width:100%; height:460px; position:relative; }
  .freqseg{ display:inline-flex; gap:2px; background:#eef1f6; border:1px solid var(--line);
    border-radius:9px; padding:3px; }
  .freqseg button{ border:0; background:transparent; color:var(--muted); font:inherit;
    font-size:12.5px; font-weight:600; padding:6px 14px; border-radius:7px; cursor:pointer; }
  .freqseg button.on{ background:#fff; color:var(--navy); box-shadow:0 1px 2px rgba(0,0,0,.08); }
  .lblrow{ display:flex; gap:14px; font-size:12px; color:var(--muted); margin-top:10px;
    flex-wrap:wrap; justify-content:center; }
  .lblrow i{ display:inline-block; width:14px; height:3px; vertical-align:middle;
    margin-right:5px; border-radius:2px; }

  /* 기술지표 표 */
  .tech-tbl{ width:100%; border-collapse:collapse; font-size:13.5px; }
  .tech-tbl td{ padding:10px 12px; border-bottom:1px solid var(--line); vertical-align:middle; }
  .tech-tbl td.k{ color:var(--muted); font-weight:600; width:40%; font-size:12.5px; }
  .tech-tbl td.v{ font-weight:700; font-variant-numeric:tabular-nums; }
  .tech-tbl td.v .muted{ font-weight:500; }
  .tech-tbl .pos{ color:var(--up); } .tech-tbl .neg{ color:var(--down); }
  .tech-badge{ display:inline-block; padding:2px 9px; border-radius:6px; font-size:12px;
    font-weight:700; margin-right:6px; background:#eef1f6; color:var(--muted); }
  .tech-badge.up{ background:#fdecea; color:#c93a2e; }
  .tech-badge.dn{ background:#e7f0fd; color:#1565c0; }
  .tech-badge.flat{ background:#eef1f6; color:#5a6b7e; }
  body[data-theme="dark"] .tech-badge{ background:rgba(255,255,255,.08); }
  body[data-theme="dark"] .tech-badge.up{ background:rgba(201,58,46,.22); color:#f4a79f; }
  body[data-theme="dark"] .tech-badge.dn{ background:rgba(21,101,192,.24); color:#9cc2f0; }
  /* 52주 위치 바 */
  .pos-bar{ position:relative; display:inline-block; width:120px; height:6px;
    background:linear-gradient(90deg,#1565c0,#cfd6df,#d23b34); border-radius:3px;
    margin:0 10px; vertical-align:middle; }
  .pos-bar i{ position:absolute; top:-3px; width:3px; height:12px; background:#16202e;
    border-radius:1px; transform:translateX(-50%); }
  body[data-theme="dark"] .pos-bar i{ background:#e8edf3; }
"""

STOCK_JS = """
const STOCK = __STOCK__;
let _ohlcv = null;
let _chart = null;

function aggregate(daily, freq){
  if(freq === 'D') return daily.slice();
  // 주봉/월봉 집계
  const out = [];
  let cur = null;
  for(const d of daily){
    const dt = new Date(d.time);
    let key;
    if(freq === 'W'){
      // ISO week 시작 (월요일)
      const t = new Date(dt); const day = (t.getUTCDay()+6)%7;
      t.setUTCDate(t.getUTCDate()-day);
      key = t.toISOString().slice(0,10);
    } else { // M
      key = dt.toISOString().slice(0,7) + '-01';
    }
    if(!cur || cur.time !== key){
      if(cur) out.push(cur);
      cur = { time: key, open: d.open, high: d.high, low: d.low, close: d.close, volume: d.volume };
    } else {
      cur.high = Math.max(cur.high, d.high);
      cur.low = Math.min(cur.low, d.low);
      cur.close = d.close;
      cur.volume += d.volume;
    }
  }
  if(cur) out.push(cur);
  return out;
}

function sma(bars, n){
  const out = [];
  for(let i=0;i<bars.length;i++){
    if(i < n-1) continue;
    let s = 0;
    for(let j=i-n+1;j<=i;j++) s += bars[j].close;
    out.push({ time: bars[i].time, value: +(s/n).toFixed(2) });
  }
  return out;
}

function nfmt(v){ return v.toLocaleString('ko-KR'); }

function buildChart(freq){
  const el = document.getElementById('chartbox');
  el.innerHTML = '';
  const dark = document.body.getAttribute('data-theme') === 'dark';
  const chart = LightweightCharts.createChart(el, {
    layout: { background:{type:'solid', color: dark ? '#0f1720' : '#ffffff'},
      textColor: dark ? '#c3cdd9' : '#16202e', fontFamily:'Pretendard, sans-serif', fontSize:11 },
    grid: { vertLines:{ color: dark?'#1c2735':'#eef1f6' }, horzLines:{ color: dark?'#1c2735':'#eef1f6' } },
    rightPriceScale:{ borderColor: dark?'#26313f':'#e4e8ef' },
    timeScale:{ borderColor: dark?'#26313f':'#e4e8ef', timeVisible:false, secondsVisible:false },
    crosshair:{ mode: LightweightCharts.CrosshairMode.Normal },
    autoSize: true,
  });
  const bars = aggregate(_ohlcv, freq);
  const candle = chart.addCandlestickSeries({
    upColor:'#d23b34', downColor:'#1565c0',
    borderUpColor:'#d23b34', borderDownColor:'#1565c0',
    wickUpColor:'#d23b34', wickDownColor:'#1565c0',
  });
  candle.setData(bars);
  // 가격(캔들) 영역을 위쪽 72% 로 제한 → 아래 거래량과 겹치지 않음
  candle.priceScale().applyOptions({ scaleMargins:{ top:0.06, bottom:0.28 } });

  // 이동평균선 (5,10,20,60,120,224)
  const periods = [[5,'#9a59b5'],[10,'#3498db'],[20,'#16a085'],[60,'#27ae60'],[120,'#e67e22'],[224,'#7f8c8d']];
  for(const [p, color] of periods){
    if(bars.length < p) continue;
    const s = chart.addLineSeries({ color, lineWidth:1, priceLineVisible:false, lastValueVisible:false });
    s.setData(sma(bars, p));
  }

  // 거래량 — 별도 오버레이 스케일, 하단 20% 영역에만 배치 (캔들과 분리)
  const vol = chart.addHistogramSeries({
    priceFormat:{ type:'volume' },
    priceScaleId:'vol',
    lastValueVisible:false, priceLineVisible:false,
  });
  vol.priceScale().applyOptions({ scaleMargins:{ top:0.82, bottom:0 } });
  vol.setData(bars.map(b => ({
    time: b.time, value: b.volume,
    color: b.close >= b.open ? 'rgba(210,59,52,.35)' : 'rgba(21,101,192,.35)',
  })));

  chart.timeScale().fitContent();
  return chart;
}

// ---- 기술지표 계산 (일봉 기준) ----
function _maAt(closes, end, p){
  if(end - p + 1 < 0) return null;
  let s = 0; for(let i=end-p+1;i<=end;i++) s += closes[i];
  return s/p;
}
function _maSeries(closes, p){
  const out = new Array(closes.length).fill(null);
  for(let i=p-1;i<closes.length;i++){ let s=0; for(let j=i-p+1;j<=i;j++) s+=closes[j]; out[i]=s/p; }
  return out;
}
function _rsi(closes, n){
  if(closes.length < n+1) return null;
  let gain=0, loss=0;
  for(let i=closes.length-n;i<closes.length;i++){
    const ch = closes[i]-closes[i-1];
    if(ch>=0) gain+=ch; else loss-=ch;
  }
  if(gain+loss===0) return 50;
  const rs = (gain/n)/((loss/n)||1e-9);
  return 100 - 100/(1+rs);
}
function computeTech(bars){
  const n = bars.length;
  if(n < 2) return null;
  const closes = bars.map(b=>b.close);
  const last = bars[n-1];
  // 52주(≈250거래일) 고저
  const win = bars.slice(Math.max(0, n-250));
  const hi52 = Math.max(...win.map(b=>b.high));
  const lo52 = Math.min(...win.map(b=>b.low));
  const fromHi = (last.close/hi52-1)*100;
  const fromLo = (last.close/lo52-1)*100;
  const isNewHi = last.high >= hi52 - 1e-9;
  const isNewLo = last.low <= lo52 + 1e-9;
  // 이동평균 + 이격도
  const ma5=_maAt(closes,n-1,5), ma20=_maAt(closes,n-1,20), ma60=_maAt(closes,n-1,60), ma120=_maAt(closes,n-1,120);
  const disp20 = ma20 ? last.close/ma20*100 : null;
  const disp60 = ma60 ? last.close/ma60*100 : null;
  // 배열
  let arrange='혼조', arrangeCls='flat';
  if(ma5&&ma20&&ma60){
    if(ma5>ma20 && ma20>ma60){ arrange='정배열'; arrangeCls='up'; }
    else if(ma5<ma20 && ma20<ma60){ arrange='역배열'; arrangeCls='dn'; }
  }
  // 골든/데드크로스 (MA20 vs MA60) — 최근 발생 시점
  let cross=null, crossDays=null, crossCls='flat';
  const s20=_maSeries(closes,20), s60=_maSeries(closes,60);
  let prevSign=null;
  for(let i=0;i<n;i++){
    if(s20[i]==null||s60[i]==null) continue;
    const sign = s20[i]>s60[i] ? 1 : (s20[i]<s60[i] ? -1 : 0);
    if(prevSign!==null && sign!==0 && sign!==prevSign){
      cross = sign>0 ? '골든크로스' : '데드크로스';
      crossCls = sign>0 ? 'up' : 'dn';
      crossDays = n-1-i;
    }
    if(sign!==0) prevSign = sign;
  }
  // RSI(14)
  const rsi = _rsi(closes,14);
  return {lastClose:last.close,hi52,lo52,fromHi,fromLo,isNewHi,isNewLo,ma5,ma20,ma60,ma120,
    disp20,disp60,arrange,arrangeCls,cross,crossDays,crossCls,rsi};
}
function _sv(v, d){ return v==null ? '—' : (v>0?'+':'')+v.toFixed(d); }
function _cls(v){ return v==null?'':(v>0?'pos':(v<0?'neg':'')); }
function renderTech(t){
  const box = document.getElementById('techTable');
  if(!box) return;
  if(!t){ box.innerHTML = '<div class="muted" style="padding:16px">데이터 부족</div>'; return; }
  const badge = (txt, cls) => `<span class="tech-badge ${cls}">${txt}</span>`;
  // 52주 위치 바
  const rng = t.hi52 - t.lo52;
  const pos = rng>0 ? ((t.lastClose - t.lo52)/rng*100) : 50;
  const rsiCls = t.rsi==null?'':(t.rsi>=70?'neg':(t.rsi<=30?'pos':''));
  const rsiTag = t.rsi==null?'':(t.rsi>=70?' 과매수':(t.rsi<=30?' 과매도':''));
  let crossHtml = '—';
  if(t.cross){ crossHtml = badge(t.cross + (t.crossDays===0?' (오늘)':` (${t.crossDays}일 전)`), t.crossCls); }
  box.innerHTML =
    '<table class="tech-tbl"><tbody>' +
    `<tr><td class="k">52주 신고/신저</td><td class="v">` +
      (t.isNewHi?badge('신고가',' up'):(t.isNewLo?badge('신저가',' dn'):'—')) +
      `<span class="pos-bar"><i style="left:${Math.max(0,Math.min(100,pos)).toFixed(1)}%"></i></span>` +
      `<span class="muted" style="font-size:11px">저 ${nfmt(Math.round(t.lo52))} ~ 고 ${nfmt(Math.round(t.hi52))}</span></td></tr>` +
    `<tr><td class="k">고점 대비 / 저점 대비</td><td class="v"><span class="${_cls(t.fromHi)}">${_sv(t.fromHi,1)}%</span>` +
      ` <span class="muted">/</span> <span class="${_cls(t.fromLo)}">${_sv(t.fromLo,1)}%</span></td></tr>` +
    `<tr><td class="k">이격도 (20일 / 60일)</td><td class="v">` +
      `${t.disp20!=null?t.disp20.toFixed(1):'—'} <span class="muted">/</span> ${t.disp60!=null?t.disp60.toFixed(1):'—'}` +
      `<span class="muted" style="font-size:11px"> (100=이평선)</span></td></tr>` +
    `<tr><td class="k">이평선 배열</td><td class="v">${badge(t.arrange, t.arrangeCls)}` +
      `<span class="muted" style="font-size:11px"> MA5·20·60 기준</span></td></tr>` +
    `<tr><td class="k">최근 크로스 (20·60)</td><td class="v">${crossHtml}</td></tr>` +
    `<tr><td class="k">RSI (14일)</td><td class="v"><span class="${rsiCls}">${t.rsi==null?'—':t.rsi.toFixed(0)}${rsiTag}</span></td></tr>` +
    '</tbody></table>';
}

async function loadAndRender(freq){
  if(!_ohlcv){
    const r = await fetch('/api/ohlcv/'+STOCK.code);
    const j = await r.json();
    if(!j.ok){ document.getElementById('chartbox').innerHTML = '<div style="padding:40px;text-align:center;color:#c0392b">차트 데이터를 가져오지 못했습니다: '+(j.error||'')+'</div>'; return; }
    _ohlcv = j.bars;
    // 헤더의 현재가/등락은 OHLCV 마지막 행으로 갱신
    if(_ohlcv.length >= 2){
      const last = _ohlcv[_ohlcv.length-1], prev = _ohlcv[_ohlcv.length-2];
      const px = document.getElementById('hdrPrice');
      const chg = document.getElementById('hdrChg');
      const asof = document.getElementById('hdrAsof');
      if(px) px.textContent = nfmt(last.close) + '원';
      if(chg){
        const d = last.close - prev.close;
        const r = d/prev.close*100;
        const sign = d>=0?'+':'';
        chg.textContent = `${sign}${nfmt(d)} (${sign}${r.toFixed(2)}%)`;
        chg.className = 'chg ' + (d>=0?'up':'dn');
      }
      if(asof) asof.textContent = last.time + ' 종가';
    }
  }
  if(_chart){ _chart.remove(); _chart = null; }
  _chart = buildChart(freq);

  // 기술지표 (일봉 기준, 최초 1회)
  if(!_techDone){
    _techDone = true;
    renderTech(computeTech(_ohlcv));
  }
}
let _techDone = false;

document.querySelectorAll('#freqseg button').forEach(b=>{
  b.addEventListener('click', ()=>{
    document.querySelectorAll('#freqseg button').forEach(x=>x.classList.remove('on'));
    b.classList.add('on');
    loadAndRender(b.dataset.freq);
  });
});

loadAndRender('D');
"""


def render_stock(fund_names, stock):
    """종목 상세 페이지: 헤더 + 캔들 차트 + 컨센서스/수급."""
    name = stock["name"]
    code = stock["code"]
    market = stock.get("market", "")
    price = stock.get("price")
    cap_mil = stock.get("cap_mil")

    px_init = f"{int(round(float(price))):,}원" if price else "—"
    cap_str = ""
    if cap_mil:
        try:
            cap_str = f"시가총액 {float(cap_mil)/100:,.0f}억"
        except (TypeError, ValueError):
            pass

    # Market_Screening 엑셀에서 추가 데이터 가져오기
    m = load_market_data()
    sinfo = (m.get("stocks_by_code") or {}).get(code) if m else None
    cons = (m.get("consensus") or {}).get(code) if m else None
    flow = (m.get("flow") or {}).get(code) if m else None

    # 시계열 수익률 카드
    perf_card = ""
    if sinfo:
        def perf_cell(v):
            if v is None:
                return '<td class="num muted">—</td>'
            cls = "pos" if v > 0 else ("neg" if v < 0 else "muted")
            sign = "+" if v > 0 else ""
            return f'<td class="num {cls}">{sign}{v:.2f}</td>'
        perf_card = (
            '<div class="card"><h2>기간별 수익률</h2>'
            '<table class="mkt-tbl"><thead><tr>'
            '<th>1D</th><th>1W</th><th>1M</th><th>3M</th><th>6M</th><th>YTD</th><th>1Y</th>'
            '</tr></thead><tbody><tr>'
            f'{perf_cell(sinfo["d1"])}{perf_cell(sinfo["wk1"])}{perf_cell(sinfo["m1"])}'
            f'{perf_cell(sinfo["m3"])}{perf_cell(sinfo["m6"])}{perf_cell(sinfo["ytd"])}'
            f'{perf_cell(sinfo["y1"])}'
            '</tr></tbody></table></div>')

    # 영업이익 컨센서스 카드
    cons_card = ""
    if cons:
        def n2(v, suffix="%"):
            if v is None:
                return "—"
            cls = "pos" if v > 0 else ("neg" if v < 0 else "muted")
            sign = "+" if v > 0 else ""
            return f'<span class="{cls}" style="font-weight:700">{sign}{v:.2f}{suffix}</span>'
        def n_op(v):
            if v is None:
                return "—"
            return f"{v/1000:,.1f}조" if abs(v) >= 1000 else f"{v:,.0f}억"
        cons_card = (
            '<div class="card"><h2>영업이익 컨센서스 (전월 대비)</h2>'
            '<table class="mkt-tbl"><thead><tr>'
            '<th style="text-align:left">기간</th>'
            '<th>이번달 컨센서스</th><th>1달 전</th><th>변화율</th>'
            '</tr></thead><tbody>'
            f'<tr><td class="name">FY1</td><td class="num">{n_op(cons["fy1_op"])}</td>'
            f'<td class="num muted">{n_op(cons["fy1_op_prev"])}</td>'
            f'<td class="num">{n2(cons["fy1_chg"])}</td></tr>'
            f'<tr><td class="name">FY2</td><td class="num">{n_op(cons["fy2_op"])}</td>'
            f'<td class="num muted">{n_op(cons["fy2_op_prev"])}</td>'
            f'<td class="num">{n2(cons["fy2_chg"])}</td></tr>'
            f'<tr><td class="name">분기 (INT1)</td><td class="num">{n_op(cons["int1_op"])}</td>'
            f'<td class="num muted">{n_op(cons["int1_op_prev"])}</td>'
            f'<td class="num">{n2(cons["int1_chg"])}</td></tr>'
            '</tbody></table>'
            f'<div class="muted" style="font-size:12px;margin-top:8px">'
            f'FY1→FY2 성장률: {n2(cons.get("growth"))}</div>'
            '</div>')

    # 수급 카드 (기관/외국인/사모/연기금/금투 × 1D/5D/20D/60D)
    flow_card = ""
    if flow:
        def amt(v):
            """백만원 → 억원 변환 + 부호 색상."""
            if v is None:
                return '<span class="muted">—</span>'
            eok = v / 100.0
            if abs(eok) < 0.01:
                return '<span class="muted">0</span>'
            cls = "pos" if eok > 0 else "neg"
            sign = "+" if eok > 0 else ""
            return f'<span class="{cls}" style="font-weight:600">{sign}{eok:,.1f}</span>'

        periods = [("1일", "d1"), ("5일", "d5"), ("20일", "d20"), ("60일", "d60")]
        flow_rows = ""
        for label, key in periods:
            f = flow.get(key, {})
            flow_rows += (
                f'<tr><td class="name">{label}</td>'
                f'<td class="num">{amt(f.get("inst_mil"))}</td>'
                f'<td class="num">{amt(f.get("frgn_mil"))}</td>'
                f'<td class="num">{amt(f.get("smfp_mil"))}</td>'
                f'<td class="num">{amt(f.get("pens_mil"))}</td>'
                f'<td class="num">{amt(f.get("fini_mil"))}</td></tr>')

        flow_card = (
            '<div class="card"><h2>투자자별 순매수 (단위: 억원)</h2>'
            '<table class="mkt-tbl"><thead><tr>'
            '<th style="text-align:left">기간</th>'
            '<th>기관 (전체)</th><th>외국인</th>'
            '<th>사모펀드</th><th>연기금</th><th>금융투자</th>'
            '</tr></thead><tbody>' + flow_rows + '</tbody></table>'
            '<div class="muted" style="font-size:12px;margin-top:8px">'
            '기관(전체)에는 사모/연기금/금융투자 + 보험/은행/기타금융 등이 포함됩니다. '
            '5/20/60일은 누적값.'
            '</div></div>')

    body = (
        f'<style>{STOCK_CSS}</style>'
        '<a href="javascript:history.length>1?history.back():(location.href=\'/\')" '
        'style="color:var(--muted);font-size:13px;text-decoration:none">← 뒤로</a>'
        '<div class="stockhead">'
        '<div>'
        '<div class="nameline">'
        f'<h1>{name}</h1>'
        f'<span class="code">{code}</span>'
        f'<span class="market">{market}</span>'
        '</div>'
        '<div class="priceline">'
        f'<span class="px" id="hdrPrice">{px_init}</span>'
        '<span class="chg" id="hdrChg">—</span>'
        '</div>'
        f'<div class="asof" id="hdrAsof">{cap_str}</div>'
        '</div>'
        '</div>'

        f'{perf_card}'

        '<div class="card"><div class="h2row"><h2>주가 차트</h2>'
        '<div class="freqseg" id="freqseg" style="margin-left:auto">'
        '<button class="on" data-freq="D">일봉</button>'
        '<button data-freq="W">주봉</button>'
        '<button data-freq="M">월봉</button>'
        '</div></div>'
        '<div id="chartbox"></div>'
        '<div class="lblrow">'
        '<span><i style="background:#9a59b5"></i>MA5</span>'
        '<span><i style="background:#3498db"></i>MA10</span>'
        '<span><i style="background:#16a085"></i>MA20</span>'
        '<span><i style="background:#27ae60"></i>MA60</span>'
        '<span><i style="background:#e67e22"></i>MA120</span>'
        '<span><i style="background:#7f8c8d"></i>MA224</span>'
        '</div></div>'

        '<div class="card"><h2>기술적 지표 '
        '<span class="src-tag">일봉 · KIS</span></h2>'
        '<div id="techTable"><div class="muted" style="padding:16px">계산 중…</div></div>'
        '<div class="muted" style="font-size:11.5px;margin-top:8px;line-height:1.6">'
        '· <b>이격도</b>: 현재가 ÷ 이동평균 ×100 (100 초과=이평선 위, 과열 신호) · '
        '<b>정배열</b>: 단기·중기·장기 이평선이 위→아래 순 (상승추세) / <b>역배열</b>은 반대 · '
        '<b>골든크로스</b>: 20일선이 60일선 상향 돌파 / <b>데드크로스</b>는 하향 돌파</div>'
        '</div>'

        f'{cons_card}{flow_card}'
    )

    scripts = STOCK_JS.replace("__STOCK__", json.dumps({"code": code, "name": name}, ensure_ascii=False))
    return page_shell(f"ANDA · {name}", "__stock__", fund_names, body, scripts)


def render_kosdaq70(fund_names, k):
    """코스닥 70 페이지 렌더링."""

    # 엑셀의 시총·순이익이 모두 '천원' 단위로 저장돼 있어 ÷1e5 로 억원 환산
    def fmt_eok(v):
        try:
            return f"{float(v) / 1e5:,.0f}"
        except (TypeError, ValueError):
            return "—"

    def fmt_ni(v):
        try:
            return f"{float(v) / 1e5:,.0f}"
        except (TypeError, ValueError):
            return "—"

    # 섹터 표
    sec_rows = ""
    for s in k["sectors"]:
        sec_rows += (
            f'<tr><td class="name">{s["name"]}</td>'
            f'<td class="num">{s["cap_eok"]:,.1f}</td>'
            f'<td class="num strong">{s["pct"]:.2f}</td></tr>')

    # 코스닥 70 후보 표
    k70_rows = ""
    for i, r in enumerate(k["k70"], 1):
        k70_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="num muted">{r.get("code","")}</td>'
            f'<td class="name">{r.get("name","")}</td>'
            f'<td class="tag">{r.get("sector","")}</td>'
            f'<td class="num">{fmt_eok(r.get("cap"))}</td></tr>')

    # 코스닥 150 표
    k150_rows = ""
    for i, r in enumerate(k["k150"], 1):
        ni = r.get("ni")
        ni_cls = "neg" if (ni is not None and ni < 0) else ""
        k150_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="num muted">{r.get("code","")}</td>'
            f'<td class="name">{r.get("name","")}</td>'
            f'<td class="tag">{r.get("sector","")}</td>'
            f'<td class="num {ni_cls}">{fmt_ni(ni)}</td>'
            f'<td class="num">{fmt_eok(r.get("cap"))}</td></tr>')

    # 신규편입 후보 (이름·업종만)
    new_rows = ""
    for i, r in enumerate(k["new"], 1):
        new_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="name">{r.get("name","")}</td>'
            f'<td class="tag">{r.get("sector","")}</td></tr>')

    # 퇴출 후보
    out_rows = ""
    for i, r in enumerate(k["out"], 1):
        ni = r.get("ni")
        ni_cls = "neg" if (ni is not None and ni < 0) else ""
        out_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="name">{r.get("name","")}</td>'
            f'<td class="tag">{r.get("sector","")}</td>'
            f'<td class="num {ni_cls}">{fmt_ni(ni)}</td>'
            f'<td class="num">{fmt_eok(r.get("cap"))}</td></tr>')

    # 흑자 퇴출 후보
    pos_rows = ""
    for i, r in enumerate(k["out_pos"], 1):
        pos_rows += (
            f'<tr><td class="num muted">{i}</td>'
            f'<td class="name">{r.get("name","")}</td>'
            f'<td class="tag">{r.get("sector","")}</td>'
            f'<td class="num pos">{fmt_ni(r.get("ni"))}</td>'
            f'<td class="num">{fmt_eok(r.get("cap"))}</td></tr>')

    body = (
        '<div class="head"><div><h1>코스닥 70 / 150 모니터</h1>'
        '<div class="subname">코스닥150 구성 · 코스닥70 후보 · 신규편입/퇴출 후보</div></div>'
        f'<div class="asof">기준일 {k["as_of"]}</div></div>'

        '<div class="stats">'
        f'<div class="stat"><div class="k">코스닥 150 종목</div><div class="v">{len(k["k150"])}<small>종목</small></div></div>'
        f'<div class="stat"><div class="k">코스닥 70 후보</div><div class="v">{len(k["k70"])}<small>종목</small></div></div>'
        f'<div class="stat"><div class="k">신규편입 후보</div><div class="v">{len(k["new"])}<small>종목</small></div></div>'
        f'<div class="stat"><div class="k">퇴출 후보</div><div class="v">{len(k["out"])}<small>종목</small></div></div>'
        f'<div class="stat"><div class="k">흑자 퇴출 후보</div><div class="v pos">{len(k["out_pos"])}<small>종목</small></div></div>'
        '</div>'

        # 섹터별 시총 비중
        '<div class="card"><h2>섹터별 시총 비중 (코스닥 150 기준)</h2>'
        '<div class="grid2">'
        '<div><canvas id="sectorChart" height="240"></canvas></div>'
        '<div style="max-height:340px;overflow:auto;">'
        '<table><thead><tr><th style="text-align:left">섹터</th>'
        '<th>시가총액(억)</th><th>비중(%)</th></tr></thead>'
        f'<tbody>{sec_rows}</tbody></table></div>'
        '</div></div>'

        # 신규편입 / 퇴출 후보
        '<div class="grid2">'
        f'<div class="card"><h2>신규편입 후보 ({len(k["new"])}종목)</h2>'
        '<div style="max-height:420px;overflow:auto;"><table><thead><tr>'
        '<th>#</th><th style="text-align:left">종목명</th><th style="text-align:left">업종</th>'
        f'</tr></thead><tbody>{new_rows}</tbody></table></div></div>'

        f'<div class="card"><h2>퇴출 후보 ({len(k["out"])}종목)</h2>'
        '<div style="max-height:420px;overflow:auto;"><table><thead><tr>'
        '<th>#</th><th style="text-align:left">종목명</th><th style="text-align:left">업종</th>'
        '<th>26E 순이익(억)</th><th>시총(억)</th>'
        f'</tr></thead><tbody>{out_rows}</tbody></table></div></div>'
        '</div>'

        # 흑자 퇴출 후보 (강조)
        f'<div class="card"><h2>퇴출 후보 중 26E 순이익 흑자 예상 ({len(k["out_pos"])}종목)</h2>'
        '<div style="max-height:420px;overflow:auto;"><table><thead><tr>'
        '<th>#</th><th style="text-align:left">종목명</th><th style="text-align:left">업종</th>'
        '<th>26E 순이익(억)</th><th>시총(억)</th>'
        f'</tr></thead><tbody>{pos_rows}</tbody></table></div></div>'

        # 코스닥 70 후보
        f'<div class="card"><h2>코스닥 70 후보 ({len(k["k70"])}종목)</h2>'
        '<div style="max-height:560px;overflow:auto;"><table><thead><tr>'
        '<th>#</th><th>코드</th><th style="text-align:left">종목명</th>'
        '<th style="text-align:left">업종</th><th>시가총액(억)</th>'
        f'</tr></thead><tbody>{k70_rows}</tbody></table></div></div>'

        # 코스닥 150
        f'<div class="card"><h2>코스닥 150 전체 ({len(k["k150"])}종목)</h2>'
        '<div style="max-height:560px;overflow:auto;"><table><thead><tr>'
        '<th>#</th><th>코드</th><th style="text-align:left">종목명</th>'
        '<th style="text-align:left">업종</th>'
        '<th>26E 순이익(억)</th><th>시가총액(억)</th>'
        f'</tr></thead><tbody>{k150_rows}</tbody></table></div></div>'
    )

    scripts = KOSDAQ_JS.replace("__SECTORS__", json.dumps(k["sectors"], ensure_ascii=False))
    return page_shell("ANDA 대시보드 · 코스닥 70", "__kosdaq__", fund_names, body, scripts)


# ---------------------------------------------------------------------------
# 시장 스크리닝 페이지 렌더링
# ---------------------------------------------------------------------------
MARKET_JS = """
const SEC_BIG = __SECBIG__;
const SEC_MID = __SECMID__;
const SEC_SMALL = __SECSMALL__;
const SEC_FLOW = __SECFLOW__;
const CREDIT_DEP = __CREDITDEP__;
const ETF_FLOWS = __ETFFLOWS__;
const _ETF_DESC = {'EWY':'iShares MSCI Korea','DRAM':'DRAM 관련 자금','FLKR':'Franklin FTSE Korea','EEM':'iShares MSCI EM (한국 ~11% 편입)'};
let _etfTk = Object.keys(ETF_FLOWS||{})[0] || null;
let _secFreq = 'd1';        // 'd1' | 'wk1' | 'm1' | 'm3' | 'ytd'
let _childLevel = 'mid';    // 'mid' | 'small'
let _selectedBig = null;    // 선택된 대분류 코드 (예 'G45')
let _bigChart = null;
let _childChart = null;
let _flowInv = 'inst';      // 기관/외국인/사모/연기금/금융투자
let _flowPer = 'd20';       // d1/d5/d20/d60
let _flowSel = null;        // 선택된 섹터명
let _flowChart = null, _flowChildChart = null;

document.querySelectorAll('.sec-head').forEach(h => {
  h.addEventListener('click', () => h.parentElement.classList.toggle('open'));
});
// 블룸버그 중분류(sub) 아코디언 — 이벤트 버블링 stop 으로 조상 sec-row 접힘 방지
document.querySelectorAll('.bl-sub-head').forEach(h => {
  h.addEventListener('click', (e) => {
    e.stopPropagation();
    h.parentElement.classList.toggle('open');
  });
});

const refreshBtn = document.getElementById('refreshMarket');
if (refreshBtn) {
  refreshBtn.addEventListener('click', () => {
    refreshBtn.disabled = true; refreshBtn.textContent = '갱신 중...';
    fetch('/api/market/refresh', { credentials: 'include' })
      .then(() => location.reload())
      .catch(() => { refreshBtn.disabled = false; refreshBtn.textContent = '↻ 새로고침'; });
  });
}

// 공통: 가로 막대 차트 그리기. colors 배열은 호출자가 계산해서 넘김.
function _drawBars(canvas, data, freq, colors, onBarClick) {
  const labels = data.map(s => s.name);
  const values = data.map(s => s[freq] || 0);
  return new Chart(canvas, {
    type: 'bar',
    data: { labels, datasets: [{
      data: values,
      backgroundColor: colors,
      borderWidth: 0,
      // 막대 두께 제한 — 항목 1개일 때 영역을 다 채워서 두꺼워지는 것 방지
      maxBarThickness: 22,
      categoryPercentage: 0.7,
      barPercentage: 0.85
    }]},
    options: {
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      onClick: onBarClick ? (e, els) => {
        if (els && els.length) onBarClick(els[0].index);
      } : undefined,
      plugins: { legend: { display: false },
        tooltip: { callbacks: {
          label: c => {
            const s = data[c.dataIndex];
            const v = c.parsed.x;
            return s.name + ': ' + v.toFixed(2) + '%  ('
              + s.stock_count + '종목 · ' + s.cap_eok.toLocaleString('ko-KR') + '억)';
          }
        } } },
      scales: {
        x: { grid: { color: '#eef1f6' }, ticks: { callback: v => v + '%' } },
        y: { grid: { display: false }, ticks: { font: { size: 11.5 }, autoSkip: false } }
      }
    }
  });
}

const COLOR_BIG = '#9ca3af';    // 대분류 — 중립 회색
const COLOR_UP = '#d23b34';     // 부모 대비 상회 — 빨강
const COLOR_DN = '#1565c0';     // 부모 대비 하회 — 파랑
const COLOR_EQ = '#9ca3af';     // 동률 — 회색

function drawBigChart() {
  if (!SEC_BIG || !SEC_BIG.length) return;
  if (_bigChart) { _bigChart.destroy(); _bigChart = null; }
  const colors = SEC_BIG.map(() => COLOR_BIG);
  _bigChart = _drawBars(
    document.getElementById('sectorBarsBig'),
    SEC_BIG, _secFreq, colors,
    (idx) => {
      const code = SEC_BIG[idx].code;
      _selectedBig = (_selectedBig === code) ? null : code;
      drawChildChart();
    }
  );
}

function drawChildChart() {
  const box = document.getElementById('childChartBox');
  const caption = document.getElementById('childCaption');
  if (_childChart) { _childChart.destroy(); _childChart = null; }

  if (!_selectedBig) {
    box.style.height = '0';
    caption.textContent = '대분류 막대를 클릭하세요';
    return;
  }

  // 선택된 대분류의 자식만 필터
  const pool = (_childLevel === 'mid') ? SEC_MID : SEC_SMALL;
  const children = pool.filter(s => s.parent_big === _selectedBig)
                       .sort((a, b) => (b[_secFreq] || 0) - (a[_secFreq] || 0));

  const parent = SEC_BIG.find(s => s.code === _selectedBig) || {};
  const bigName = parent.name || _selectedBig;
  if (!children.length) {
    box.style.height = '80px';
    caption.textContent = bigName + ': 자식 분류 데이터 없음';
    return;
  }

  // 부모 대비 outperform/underperform 색칠
  const pv = parent[_secFreq] || 0;
  const colors = children.map(c => {
    const v = c[_secFreq] || 0;
    if (v > pv) return COLOR_UP;
    if (v < pv) return COLOR_DN;
    return COLOR_EQ;
  });

  // 그룹 수에 따라 차트 높이 동적 조정. 1개일 땐 빈 공간 최소화.
  //   1개 → 90, 2~3개 → 항목당 ~38px, 그 이상 → 28px/항목 + 여백
  let h;
  if (children.length === 1) h = 90;
  else if (children.length <= 3) h = children.length * 40 + 30;
  else h = children.length * 28 + 50;
  box.style.height = h + 'px';
  const pvSign = pv > 0 ? '+' : '';
  caption.innerHTML = bigName + ' (' + pvSign + pv.toFixed(2) + '%) → ' +
                      children.length + '개 ' +
                      (_childLevel === 'mid' ? '중분류' : '소분류') +
                      ' &nbsp;<span style="color:' + COLOR_UP +
                      ';font-weight:700">■</span> 부모 대비 상회 &nbsp;' +
                      '<span style="color:' + COLOR_DN +
                      ';font-weight:700">■</span> 부모 대비 하회';

  // requestAnimationFrame: height 변경 후 canvas 다시 측정해야 함
  requestAnimationFrame(() => {
    _childChart = _drawBars(
      document.getElementById('sectorBarsChild'),
      children, _secFreq, colors, null
    );
  });
}

// 기간 토글 (1D/5D/1M/3M/YTD) — 두 차트 동기 갱신
document.querySelectorAll('#secFreqSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#secFreqSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    _secFreq = b.dataset.f;
    drawBigChart();
    if (_selectedBig) drawChildChart();
  });
});

// 자식 레벨 토글 (중분류 / 소분류)
document.querySelectorAll('#childLevelSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#childLevelSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    _childLevel = b.dataset.lv;
    if (_selectedBig) drawChildChart();
  });
});

// 섹터별 주요 종목 — 토글 시 각 섹터 헤더의 변동률/막대를 해당 기간으로 갱신
function updateSec2(freq) {
  const cap = (freq === 'd1') ? 3.0 : (freq === 'wk1' ? 6.0 : 10.0);
  document.querySelectorAll('.sec-row').forEach(row => {
    const v = parseFloat(row.dataset[freq] || '0') || 0;
    const chg = row.querySelector('.chg');
    if (chg) {
      chg.textContent = (v > 0 ? '+' : '') + v.toFixed(2) + '%';
      chg.classList.remove('up', 'dn', 'flat');
      chg.classList.add(v > 0 ? 'up' : (v < 0 ? 'dn' : 'flat'));
    }
    const bar = row.querySelector('.bar');
    if (bar) {
      const pn = Math.max(-1, Math.min(1, v / cap));
      const bw = (Math.abs(pn) * 50).toFixed(1);
      bar.innerHTML = '<i class="' + (v >= 0 ? 'up' : 'dn') + '" style="width:' + bw + '%"></i>';
    }
  });
}
document.querySelectorAll('#sec2FreqSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#sec2FreqSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    updateSec2(b.dataset.f);
  });
});

// ===== 업종별 투자자 수급 (순매수 억원) =====
const _INV_NAME = {inst:'기관', frgn:'외국인', smfp:'사모펀드', pens:'연기금', fini:'금융투자'};
const _PER_NAME = {d1:'1일', d5:'5일 누적', d20:'20일 누적', d60:'60일 누적'};

// 순매수(+)=빨강, 순매도(-)=파랑
function _flowColor(v){ return v >= 0 ? '#d23b34' : '#1565c0'; }

function drawFlowChart(){
  if (!SEC_FLOW || !SEC_FLOW.length) return;
  const rows = SEC_FLOW
    .map(e => ({name:e.sector, val:(e[_flowPer]||{})[_flowInv] || 0, code:e.sector}))
    .sort((a,b) => b.val - a.val);
  if (_flowChart) { _flowChart.destroy(); _flowChart = null; }
  _flowChart = new Chart(document.getElementById('flowBars'), {
    type:'bar',
    data:{ labels: rows.map(r=>r.name), datasets:[{
      data: rows.map(r=>r.val),
      backgroundColor: rows.map(r=>_flowColor(r.val)), borderWidth:0,
      maxBarThickness: 26,
    }]},
    options:{ responsive:true, maintainAspectRatio:false, indexAxis:'y',
      onClick: (e, els) => { if(els&&els.length){ _flowSel = rows[els[0].index].name; drawFlowChild(); } },
      plugins:{ legend:{display:false},
        tooltip:{ callbacks:{ label: c => {
          const v = c.parsed.x;
          return (v>=0?'순매수 +':'순매도 ') + v.toLocaleString('ko-KR') + '억';
        } } } },
      scales:{
        x:{ grid:{color:'#eef1f6'}, ticks:{ callback:v=> v.toLocaleString('ko-KR') } },
        y:{ grid:{display:false}, ticks:{ font:{size:12}, autoSkip:false } } } }
  });
  document.getElementById('flowCaption').textContent =
    _INV_NAME[_flowInv] + ' · ' + _PER_NAME[_flowPer] + ' 순매수(빨강)/순매도(파랑)';
  if (_flowSel) drawFlowChild();
}

function drawFlowChild(){
  const box = document.getElementById('flowChildBox');
  if (_flowChildChart) { _flowChildChart.destroy(); _flowChildChart = null; }
  const sec = SEC_FLOW.find(e => e.sector === _flowSel);
  if (!sec || !sec.stocks || !sec.stocks.length){ box.style.height='0'; return; }
  const rows = sec.stocks
    .map(s => ({name:s.name, val:(s[_flowPer]||{})[_flowInv] || 0}))
    .sort((a,b) => b.val - a.val);
  const h = Math.max(140, rows.length*26 + 46);
  box.style.height = h + 'px';
  setTimeout(() => {
    _flowChildChart = new Chart(document.getElementById('flowChild'), {
      type:'bar',
      data:{ labels: rows.map(r=>r.name), datasets:[{
        data: rows.map(r=>r.val),
        backgroundColor: rows.map(r=>_flowColor(r.val)), borderWidth:0, maxBarThickness:20,
      }]},
      options:{ responsive:true, maintainAspectRatio:false, indexAxis:'y',
        plugins:{ legend:{display:false},
          title:{ display:true, text: _flowSel + ' — 주요종목 ' + _INV_NAME[_flowInv] + ' 순매수 (' + _PER_NAME[_flowPer] + ', 억)',
            color:'#7c8b9c', font:{size:12} },
          tooltip:{ callbacks:{ label: c => {
            const v=c.parsed.x; return (v>=0?'순매수 +':'순매도 ') + v.toLocaleString('ko-KR') + '억'; } } } },
        scales:{ x:{ grid:{color:'#eef1f6'}, ticks:{ callback:v=> v.toLocaleString('ko-KR') } },
          y:{ grid:{display:false}, ticks:{ font:{size:11}, autoSkip:false } } } }
    });
  }, 30);
}

document.querySelectorAll('#flowInvSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#flowInvSeg button').forEach(x=>x.classList.remove('on'));
    b.classList.add('on'); _flowInv = b.dataset.i; drawFlowChart();
  });
});
document.querySelectorAll('#flowPerSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#flowPerSeg button').forEach(x=>x.classList.remove('on'));
    b.classList.add('on'); _flowPer = b.dataset.p; drawFlowChart();
  });
});

// ===== 신용잔고 vs 예탁금 (일별 시계열) =====
let _creditChart = null;
function drawCreditChart(){
  if (!CREDIT_DEP || !CREDIT_DEP.length) return;
  const el = document.getElementById('creditChart');
  if (!el) return;
  const labels = CREDIT_DEP.map(r => r.d);
  const mk = (key) => CREDIT_DEP.map(r => r[key]);
  if (_creditChart) { _creditChart.destroy(); _creditChart = null; }
  _creditChart = new Chart(el, {
    type: 'line',
    data: { labels, datasets: [
      { label:'KOSPI 신용잔고', data: mk('kospi_credit'), yAxisID:'yCr',
        borderColor:'#e8853a', backgroundColor:'#e8853a', borderWidth:1.8,
        pointRadius:0, tension:.1 },
      { label:'KOSDAQ 신용잔고', data: mk('kosdaq_credit'), yAxisID:'yCr',
        borderColor:'#9b59b6', backgroundColor:'#9b59b6', borderWidth:1.8,
        pointRadius:0, tension:.1 },
      { label:'투자자예탁금', data: mk('deposit'), yAxisID:'yDep',
        borderColor:'#22c7d6', backgroundColor:'#22c7d6', borderWidth:1.8,
        pointRadius:0, tension:.1 },
      { label:'KOSPI 지수', data: mk('kospi'), yAxisID:'yKospi',
        borderColor:'#2e6fd6', backgroundColor:'#2e6fd6', borderWidth:1.2,
        pointRadius:0, tension:.1, borderDash:[] },
    ]},
    options: {
      responsive:true, maintainAspectRatio:false,
      interaction:{ mode:'index', intersect:false },
      plugins:{ legend:{ display:false },
        tooltip:{ callbacks:{ label: c => {
          const v = c.parsed.y;
          if (v == null) return '';
          if (c.dataset.label === 'KOSPI 지수') return 'KOSPI 지수: ' + v.toLocaleString('ko-KR');
          return c.dataset.label + ': ' + v.toFixed(1) + '조';
        } } } },
      scales:{
        x:{ grid:{ display:false }, ticks:{ maxTicksLimit:10, autoSkip:true, maxRotation:0, font:{size:11} } },
        yCr:{ position:'left', title:{ display:true, text:'신용잔고 (조)', color:'#7a8696', font:{size:11} },
              grid:{ color:'#eef1f6' }, ticks:{ font:{size:11} } },
        yDep:{ position:'right', title:{ display:true, text:'예탁금 (조)', color:'#7a8696', font:{size:11} },
               grid:{ display:false }, ticks:{ font:{size:11} } },
        yKospi:{ position:'right', display:false },
      }
    }
  });
}

// ===== 한국추종 ETF 주간 펀드유입액 (EWY/FLKR/EEM 토글, 막대) =====
let _etfChart = null;
function drawEtfChart(){
  if (!_etfTk || !ETF_FLOWS[_etfTk]) return;
  const el = document.getElementById('etfFlowChart');
  if (!el) return;
  const rows = ETF_FLOWS[_etfTk];
  const labels = rows.map(r => r.d);
  const vals = rows.map(r => r.flow);
  if (_etfChart) { _etfChart.destroy(); _etfChart = null; }
  _etfChart = new Chart(el, {
    type:'bar',
    data:{ labels, datasets:[{
      data: vals,
      backgroundColor: vals.map(v => v >= 0 ? '#e8853a' : '#1565c0'),
      borderWidth:0, maxBarThickness:14,
    }]},
    options:{ responsive:true, maintainAspectRatio:false,
      plugins:{ legend:{display:false},
        tooltip:{ callbacks:{ label: c => {
          const v=c.parsed.y;
          return _etfTk + ' ' + (v>=0?'순유입 +':'순유출 ') + v.toLocaleString('ko-KR') + ' 백만달러';
        } } } },
      scales:{
        x:{ grid:{display:false}, ticks:{ maxTicksLimit:12, autoSkip:true, maxRotation:0, font:{size:11} } },
        y:{ grid:{color:'#eef1f6'}, title:{display:true, text:'주간 펀드유입 ($백만)', color:'#7a8696', font:{size:11}},
            ticks:{ font:{size:11}, callback:v=> v.toLocaleString('ko-KR') } } } }
  });
  const desc = document.getElementById('etfDesc');
  if (desc) desc.textContent = _etfTk + ' · ' + (_ETF_DESC[_etfTk]||'') + ' · 단위 $백만';
}
document.querySelectorAll('#etfTkSeg button').forEach(b => {
  b.addEventListener('click', () => {
    document.querySelectorAll('#etfTkSeg button').forEach(x => x.classList.remove('on'));
    b.classList.add('on'); _etfTk = b.dataset.tk; drawEtfChart();
  });
});

// 다크모드 토글 시 차트 다시 그리기
window.addEventListener('theme-changed', () => {
  drawBigChart();
  if (_selectedBig) drawChildChart();
  drawFlowChart();
  drawCreditChart();
  drawEtfChart();
});

drawBigChart();
drawFlowChart();
// 하단 카드 차트들은 초기 레이아웃 확정 후 그려야 폭이 잡힘
setTimeout(drawCreditChart, 80);
setTimeout(drawEtfChart, 90);
"""


def _cell(v, fmt="pct"):
    """변동률 셀 HTML (색상 자동)."""
    if v is None:
        return '<td class="num muted">—</td>'
    cls = "pos" if v > 0 else ("neg" if v < 0 else "muted")
    sign = "+" if v > 0 else ""
    if fmt == "pct":
        return f'<td class="num {cls}">{sign}{v:.2f}</td>'
    return f'<td class="num {cls}">{sign}{v:,.2f}</td>'


def _fmt_close(v):
    if v is None:
        return "—"
    if abs(v) >= 1000:
        return f"{v:,.2f}"
    if abs(v) >= 10:
        return f"{v:,.2f}"
    return f"{v:,.4f}"


def _macro_table(items):
    """매크로 표 한 블록 — 이름 | 현재가 | 1D | 5D | 1M | 3M | YTD | 1YR"""
    if not items:
        return '<div class="muted" style="padding:14px;text-align:center">데이터 없음</div>'
    rows = ""
    for it in items:
        rows += (
            f'<tr><td class="name">{it["name"]}</td>'
            f'<td class="num strong">{_fmt_close(it["close"])}</td>'
            f'{_cell(it["d1"])}{_cell(it["d5"])}{_cell(it["m1"])}'
            f'{_cell(it["m3"])}{_cell(it["ytd"])}{_cell(it["y1"])}</tr>')
    return (
        '<table class="mkt-tbl"><thead><tr>'
        '<th style="text-align:left">항목</th><th>현재가</th>'
        '<th>1D</th><th>5D</th><th>1M</th><th>3M</th><th>YTD</th><th>1Y</th>'
        f'</tr></thead><tbody>{rows}</tbody></table>')


def render_market(fund_names, m):
    if not m:
        # 엑셀 없을 때 안내
        body = (
            '<div class="head"><div><h1>시장 스크리닝</h1>'
            '<div class="subname">data/Market_Screening_*.xlsm 파일이 필요합니다</div></div></div>'
            '<div class="card"><h2>엑셀 파일이 없습니다</h2>'
            '<p style="color:var(--muted);line-height:1.7">'
            '시장 데이터를 표시하려면 <code>Market_Screening_YYMMDD.xlsm</code> 형식의 파일을 '
            '<code>data/</code> 폴더에 넣어주세요. 매일 새 파일을 같은 폴더에 덮어쓰면 자동 갱신됩니다.'
            '</p></div>')
        return page_shell("ANDA 대시보드 · 시장", "__market__", fund_names, body, "")

    sectors = m.get("sectors", [])

    def _serialize_sectors(arr):
        return [
            {"name": s["name"], "code": s.get("code", ""),
             "parent_big": s.get("parent_big", ""),
             "cap_eok": s["cap_eok"], "stock_count": s["stock_count"],
             "d1": s["d1"], "wk1": s.get("wk1", 0),
             "m1": s["m1"], "m3": s["m3"], "ytd": s["ytd"]}
            for s in arr
        ]
    sec_big_json = json.dumps(_serialize_sectors(m.get("sectors_big", [])), ensure_ascii=False)
    sec_mid_json = json.dumps(_serialize_sectors(m.get("sectors_mid", [])), ensure_ascii=False)
    sec_small_json = json.dumps(_serialize_sectors(m.get("sectors_small", [])), ensure_ascii=False)

    # 섹터 아코디언
    sec_rows_html = ""
    for s in sectors:
        d1, wk1, m1, m3, ytd = s["d1"], s.get("wk1", 0) or 0, s["m1"], s["m3"], s["ytd"]
        d1cls = "up" if d1 > 0 else ("dn" if d1 < 0 else "flat")
        d1sign = "+" if d1 > 0 else ""
        # 작은 바: 1D 기준 ±3%
        cap = 3.0
        pn = min(max(d1 / cap, -1.0), 1.0)
        bw = abs(pn) * 50
        bar_html = (f'<i class="up" style="width:{bw:.1f}%"></i>' if d1 >= 0
                    else f'<i class="dn" style="width:{bw:.1f}%"></i>')

        stock_rows = ""
        for i, st in enumerate(s.get("stocks", []), 1):
            stock_rows += (
                f'<tr><td class="num muted">{i}</td>'
                f'<td class="num muted">{st["code"]}</td>'
                f'<td class="name"><a href="/stock/{st["code"]}" '
                f'style="color:inherit;text-decoration:none">{st["name"]}</a></td>'
                f'<td class="num">{st["price"]:,}</td>'
                f'{_cell(st["d1"])}{_cell(st.get("wk1"))}{_cell(st["m1"])}'
                f'{_cell(st["m3"])}{_cell(st["ytd"])}'
                f'<td class="num muted">{st["cap_eok"]:,.0f}</td></tr>')

        # data-* 속성으로 각 기간 변동률 저장 — 토글 시 JS 가 읽어 헤더 변동률/막대 갱신
        sec_rows_html += (
            f'<div class="sec-row" data-d1="{d1}" data-wk1="{wk1}" '
            f'data-m1="{m1}" data-m3="{m3}" data-ytd="{ytd}"><div class="sec-head">'
            f'<span class="arrow">▶</span>'
            f'<span class="nm">{s["name"]}</span>'
            f'<span class="bar">{bar_html}</span>'
            f'<span class="chg {d1cls}">{d1sign}{d1:.2f}%</span>'
            f'<span class="cap">{s["cap_eok"]:,.0f}억 · {s["stock_count"]}종목</span>'
            f'</div>'
            f'<div class="sec-body"><table class="mkt-tbl"><thead><tr>'
            f'<th>#</th><th>코드</th><th style="text-align:left">종목명</th>'
            f'<th>종가</th><th>1D</th><th>5D</th><th>1M</th><th>3M</th><th>YTD</th><th>시총(억)</th>'
            f'</tr></thead><tbody>{stock_rows}</tbody></table></div></div>')

    body = (
        '<div class="head"><div><h1>시장 스크리닝</h1></div>'
        f'<div class="asof">기준일 {m.get("as_of","-")} '
        '<button id="refreshMarket" class="refresh-btn">↻ 새로고침</button></div></div>'

        '<div class="card"><h2>국내 지수</h2>'
        f'{_macro_table(m.get("dom_idx", []))}</div>'

        '<div class="card"><h2>글로벌 지수</h2>'
        f'{_macro_table(m.get("global_idx", []))}</div>'

        '<div class="grid2">'
        '<div class="card"><h2>환율</h2>'
        f'{_macro_table(m.get("fx", []))}</div>'
        '<div class="card"><h2>원자재</h2>'
        f'{_macro_table(m.get("comm", []))}</div>'
        '</div>')

    if m.get("etc"):
        body += ('<div class="card"><h2>기타 (운임 · 가상자산)</h2>'
                 f'{_macro_table(m.get("etc", []))}</div>')

    body += (
        '<div class="card">'
        '<div class="h2row"><h2>WICS 섹터별 변동 (시총 가중)</h2>'
        '<div class="seg" id="secFreqSeg" style="margin-left:auto">'
        '<button data-f="d1" class="on">1D</button>'
        '<button data-f="wk1">5D</button>'
        '<button data-f="m1">1M</button>'
        '<button data-f="m3">3M</button>'
        '<button data-f="ytd">YTD</button>'
        '</div></div>'

        # 대분류 차트 (항상 표시) — 막대 클릭 시 아래 자식 차트 갱신
        '<div class="chart-sub">대분류 (클릭하면 아래에 자식 분류가 펼쳐집니다)</div>'
        '<div class="secchart" style="height:340px"><canvas id="sectorBarsBig"></canvas></div>'

        # 자식 분류 비교 토글
        '<div class="wics-toggle-row">'
        '<span class="wics-label">자식 분류:</span>'
        '<div class="seg" id="childLevelSeg">'
        '<button data-lv="mid" class="on">중분류</button>'
        '<button data-lv="small">소분류</button>'
        '</div>'
        '<span class="wics-count" id="childCaption">대분류 막대를 클릭하세요</span>'
        '</div>'

        # 선택된 대분류의 자식 차트
        '<div class="secchart" id="childChartBox" style="height:0;overflow:hidden;'
        'transition:height .25s"><canvas id="sectorBarsChild"></canvas></div>'

        '</div>'  # /chart card

        '<div class="card"><div class="h2row">'
        '<h2>섹터별 주요 종목 (시총 상위 10 · 클릭해서 펼치기)</h2>'
        '<div class="seg" id="sec2FreqSeg" style="margin-left:auto">'
        '<button data-f="d1" class="on">1D</button>'
        '<button data-f="wk1">5D</button>'
        '<button data-f="m1">1M</button>'
        '<button data-f="m3">3M</button>'
        '<button data-f="ytd">YTD</button>'
        '</div></div>'
        f'<div class="sec-list">{sec_rows_html}</div></div>'

        # 업종별 투자자 수급 (바차트 + 종목 드릴다운)
        '<div class="card">'
        '<div class="h2row"><h2>업종별 투자자 수급 '
        '<span class="src-tag">순매수 억원 · Market_Screening</span></h2>'
        '<div class="seg" id="flowPerSeg" style="margin-left:auto">'
        '<button data-p="d1">1D</button>'
        '<button data-p="d5">5D</button>'
        '<button data-p="d20" class="on">20D</button>'
        '<button data-p="d60">60D</button>'
        '</div></div>'
        '<div class="flow-inv-row">'
        '<span class="wics-label">투자자</span>'
        '<div class="seg" id="flowInvSeg">'
        '<button data-i="inst" class="on">기관</button>'
        '<button data-i="frgn">외국인</button>'
        '<button data-i="smfp">사모펀드</button>'
        '<button data-i="pens">연기금</button>'
        '<button data-i="fini">금융투자</button>'
        '</div>'
        '<span class="wics-count" id="flowCaption">—</span>'
        '</div>'
        '<div class="secchart" style="height:340px"><canvas id="flowBars"></canvas></div>'
        '<div class="chart-sub">업종 막대를 클릭하면 아래에 그 업종 주요 종목(시총 상위 15) 수급이 펼쳐집니다</div>'
        '<div class="secchart" id="flowChildBox" style="height:0;overflow:hidden;'
        'transition:height .25s"><canvas id="flowChild"></canvas></div>'
        '</div>'
    )

    # 신용잔고 vs 예탁금 (금투협 종합통계)
    cd = load_credit_deposit()
    if cd and cd.get("series"):
        body += (
            '<div class="card"><div class="h2row">'
            '<h2>신용잔고 vs 예탁금 <span class="src-tag">금융투자협회 · 일별</span></h2>'
            f'<span class="asof" style="margin-left:auto">기준일 {cd.get("as_of","-")} · 단위 조</span>'
            '</div>'
            '<div class="secchart" style="height:380px"><canvas id="creditChart"></canvas></div>'
            '<div class="tmlegend">'
            '<i style="background:#e8853a"></i>KOSPI 신용잔고'
            '<i style="background:#9b59b6"></i>KOSDAQ 신용잔고'
            '<i style="background:#22c7d6"></i>투자자예탁금'
            '<i style="background:#2e6fd6"></i>KOSPI 지수'
            '<span>· 매일 자동 갱신 (금투협 발표)</span></div>'
            '</div>')

    # 한국추종 ETF 주간 펀드유입액 (EWY / FLKR / EEM 토글)
    etff = load_etf_flows()
    if etff and etff.get("etfs"):
        _ETF_DESC = {"EWY": "iShares MSCI Korea",
                     "DRAM": "DRAM 관련 자금",
                     "FLKR": "Franklin FTSE Korea",
                     "EEM": "iShares MSCI EM (한국 ~11% 편입)"}
        tk0 = etff["tickers"][0]
        tabs = ""
        for tk in etff["tickers"]:
            on = ' class="on"' if tk == tk0 else ''
            tabs += f'<button data-tk="{tk}"{on}>{tk}</button>'
        body += (
            '<div class="card"><div class="h2row">'
            '<h2>한국추종 ETF 주간 펀드유입액 <span class="src-tag">Bloomberg · 주간</span></h2>'
            f'<div class="seg" id="etfTkSeg" style="margin-left:auto">{tabs}</div>'
            '</div>'
            f'<div class="chart-sub" id="etfDesc">{tk0} · {_ETF_DESC.get(tk0,"")} · 단위 $백만</div>'
            '<div class="secchart" style="height:340px"><canvas id="etfFlowChart"></canvas></div>'
            '<div class="tmlegend">'
            '<i style="background:#e8853a"></i>순유입(+)'
            '<i style="background:#1565c0"></i>순유출(−)'
            f'<span>· 외국인의 한국주식 투자심리 프록시 · 기준일 {etff.get("as_of","-")} · 매주 갱신</span></div>'
            '</div>')

    # 업종별 글로벌 피어 수익률 (블룸버그) — 러프 구현
    body += _render_bloomberg_peers_card()

    # 조건별 종목 스크리닝 보드
    body += _render_screening_card()

    sec_flow_json = json.dumps(build_sector_flow(m), ensure_ascii=False)
    credit_json = json.dumps((cd or {}).get("series", []), ensure_ascii=False)
    etf_json = json.dumps((etff or {}).get("etfs", {}), ensure_ascii=False)
    scripts = (MARKET_JS
               .replace("__SECBIG__", sec_big_json)
               .replace("__SECMID__", sec_mid_json)
               .replace("__SECSMALL__", sec_small_json)
               .replace("__SECFLOW__", sec_flow_json)
               .replace("__CREDITDEP__", credit_json)
               .replace("__ETFFLOWS__", etf_json))
    return page_shell("ANDA 대시보드 · 시장", "__market__", fund_names, body, scripts)


def _render_screening_card():
    """조건별 종목 스크리닝 보드 카드 HTML (Market_Screening 'screening' 시트)."""
    sb = load_screening_board()
    if not sb:
        return ""
    # 종목명 → 코드 (클릭 시 종목 페이지 이동)
    name2code = {}
    for code, v in get_stock_master().items():
        nm = (v.get("name") or "").strip()
        if nm and nm not in name2code:
            name2code[nm] = code

    def _pct(v):
        if v is None:
            return '<span class="muted">—</span>'
        cls = "pos" if v > 0 else ("neg" if v < 0 else "muted")
        return f'<span class="{cls}">{"+" if v>0 else ""}{v:.2f}%</span>'
    def _eok(v):
        cls = "pos" if v > 0 else ("neg" if v < 0 else "muted")
        return f'<span class="{cls}">{"+" if v>0 else ""}{v:,.0f}</span>'

    # 상단 KOSPI/KOSDAQ 요약
    sum_html = ""
    for s in sb.get("summary", []):
        sum_html += (
            f'<span class="scr-sum"><b>{s["name"]}</b> {_pct(s.get("ret"))} '
            f'<span class="muted">· 외국인</span> {_eok(s.get("frgn_eok",0))}억 '
            f'<span class="muted">· 기관</span> {_eok(s.get("inst_eok",0))}억</span>')

    # 카테고리 컬럼들 (그룹별)
    groups_html = ""
    for g in sb["groups"]:
        cols = ""
        for cat in g["cats"]:
            items = ""
            for nm in cat["stocks"]:
                code = name2code.get(nm)
                if code:
                    items += (f'<a class="scr-item" href="/stock/{code}">{nm}</a>')
                else:
                    items += f'<span class="scr-item nolink">{nm}</span>'
            if not items:
                items = '<span class="scr-item muted">—</span>'
            cols += (
                f'<div class="scr-col">'
                f'<div class="scr-col-head {cat["cls"]}">{cat["name"]}'
                f'<span class="scr-cnt">{len(cat["stocks"])}</span></div>'
                f'<div class="scr-col-body">{items}</div></div>')
        groups_html += (
            f'<div class="scr-group"><div class="scr-group-title">{g["title"]}</div>'
            f'<div class="scr-cols">{cols}</div></div>')

    return (
        '<div class="card"><div class="h2row"><h2>조건별 종목 스크리닝 '
        '<span class="src-tag">Market_Screening</span></h2>'
        f'<span class="asof" style="margin-left:auto">기준 {sb.get("as_of","-")}</span></div>'
        f'<div class="scr-summary">{sum_html}</div>'
        f'{groups_html}'
        '<div class="muted" style="font-size:11.5px;margin-top:8px">'
        '종목명 클릭 시 상세 페이지로 이동 · 매일 파일 갱신 시 자동 반영</div>'
        '</div>')


def _render_bloomberg_peers_card():
    """블룸버그 valuation table → 업종별 글로벌 피어 아코디언 HTML."""
    bl = load_bloomberg_peers()
    if not bl:
        return ('<div class="card"><h2>업종별 글로벌 피어 수익률 '
                '<span class="src-tag">Bloomberg Valuation Table</span></h2>'
                '<div class="muted" style="padding:20px;text-align:center">'
                'Z 드라이브의 <code>★Bloomberg Valuation Table(수정중).xlsx</code> 파일에 접근할 수 없습니다. '
                'VPN 연결 또는 로컬 사본을 <code>data/</code> 폴더에 배치하세요.'
                '</div></div>')

    import datetime as _dt
    mtime_str = _dt.datetime.fromtimestamp(bl.get("mtime") or 0).strftime("%Y-%m-%d %H:%M")

    import math
    def _finite(v):
        try:
            return v is not None and not math.isnan(float(v))
        except (TypeError, ValueError):
            return False
    def _fmt(v):
        if not _finite(v): return '—'
        return ('+' if v > 0 else '') + f'{v:.2f}%'
    def _fnum(x, fmt):
        if not _finite(x): return '—'
        return format(float(x), fmt)
    def _p(x):
        if not _finite(x): return '—'
        return ('+' if x > 0 else '') + f'{x:.2f}'
    def _num_cls(x):
        if not _finite(x): return 'muted'
        return 'pos' if x > 0 else ('neg' if x < 0 else 'muted')
    def _avg(stocks, field):
        vals = [float(s[field]) for s in stocks if _finite(s.get(field))]
        return round(sum(vals) / len(vals), 2) if vals else None
    def _mid_cls(v):
        if not _finite(v): return 'flat'
        return 'up' if v > 0 else ('dn' if v < 0 else 'flat')

    # Bloomberg 티커 접미사 → 국가/국기 매핑
    #   KS = 한국(KOSPI), KQ = 한국(KOSDAQ), US = 미국, HK = 홍콩,
    #   JT/JP = 일본, TT = 대만, CH = 중국 상하이, LN = 영국
    CTRY = {
        "KS": ("🇰🇷", "KR"), "KQ": ("🇰🇷", "KR"),
        "US": ("🇺🇸", "US"), "UN": ("🇺🇸", "US"), "UQ": ("🇺🇸", "US"), "UW": ("🇺🇸", "US"),
        "HK": ("🇭🇰", "HK"), "CH": ("🇨🇳", "CN"), "C1": ("🇨🇳", "CN"), "C2": ("🇨🇳", "CN"),
        "JT": ("🇯🇵", "JP"), "JP": ("🇯🇵", "JP"),
        "TT": ("🇹🇼", "TW"), "TW": ("🇹🇼", "TW"),
        "LN": ("🇬🇧", "GB"), "GY": ("🇩🇪", "DE"), "FP": ("🇫🇷", "FR"),
        "IN": ("🇮🇳", "IN"), "SP": ("🇸🇬", "SG"),
    }
    def _country(ticker):
        # "005930 KS EQUITY" → 접미사 "KS"
        parts = (ticker or "").split()
        if len(parts) >= 3:
            suf = parts[1].upper()
            return CTRY.get(suf, ("🌐", suf))
        return ("🌐", "-")

    def _sub_html(sub, sub_idx):
        """중분류(sub) 하나에 대한 종목 표 + 헤더 HTML."""
        stocks = sub["stocks"]
        avg_1m = _avg(stocks, "chg_1m")
        avg_3m = _avg(stocks, "chg_3m")
        avg_1y = _avg(stocks, "chg_1yr")
        # 국가별 종목 수 요약 (헤더 표시용)
        ctry_count = {}
        for s in stocks:
            _, code = _country(s.get("ticker"))
            ctry_count[code] = ctry_count.get(code, 0) + 1
        kr_n = ctry_count.get("KR", 0)
        foreign_n = len(stocks) - kr_n

        avg_1w = _avg(stocks, "chg_1w")

        stock_rows = ""
        for i, st in enumerate(stocks, 1):
            flag, ccode = _country(st.get("ticker"))
            row_cls = "kr-row" if ccode == "KR" else ""
            stock_rows += (
                f'<tr class="{row_cls}"><td class="num muted">{i}</td>'
                f'<td class="ctry">{flag}</td>'
                f'<td class="num muted" style="font-family:monospace;font-size:11px">{st["ticker"]}</td>'
                f'<td class="name">{st["name"]}</td>'
                f'<td class="num muted">{st.get("currency","")}</td>'
                f'<td class="num">{(st.get("price") or 0):,.0f}</td>'
                f'<td class="num {_num_cls(st.get("chg_1w"))}">{_p(st.get("chg_1w"))}</td>'
                f'<td class="num {_num_cls(st.get("chg_1m"))}">{_p(st.get("chg_1m"))}</td>'
                f'<td class="num {_num_cls(st.get("chg_3m"))}">{_p(st.get("chg_3m"))}</td>'
                f'<td class="num {_num_cls(st.get("chg_1yr"))}">{_p(st.get("chg_1yr"))}</td>'
                f'<td class="num muted">{_fnum(st.get("mkt_cap_usd"), ",.1f")}</td>'
                f'<td class="num per">{_fnum(st.get("per_26e"), ".1f")}</td>'
                f'<td class="num per">{_fnum(st.get("per_27e"), ".1f")}</td>'
                f'<td class="num muted">{_fnum(st.get("per_28e"), ".1f")}</td>'
                f'<td class="num muted">{_fnum(st.get("pbr_26e"), ".2f")}</td></tr>')

        return (
            f'<div class="bl-sub"><div class="bl-sub-head">'
            f'<span class="arrow">▶</span>'
            f'<span class="nm">{sub["name"]}</span>'
            f'<span class="chg {_mid_cls(avg_1w)}">1W {_fmt(avg_1w)}</span>'
            f'<span class="cap">1M {_fmt(avg_1m)} · 3M {_fmt(avg_3m)} · 1Y {_fmt(avg_1y)} · '
            f'🇰🇷 {kr_n} · 🌐 {foreign_n} · {len(stocks)}종목</span>'
            f'</div>'
            f'<div class="bl-sub-body"><table class="mkt-tbl bl-tbl"><thead><tr>'
            f'<th>#</th><th></th><th>Ticker</th><th style="text-align:left">Company</th>'
            f'<th>CRNCY</th><th>Price</th>'
            f'<th>1W</th><th>1M</th><th>3M</th><th>1Y</th>'
            f'<th>Mkt cap<br>(USD bil)</th>'
            f'<th class="per">PER<br>26E</th><th class="per">PER<br>27E</th>'
            f'<th>PER<br>28E</th><th>PBR<br>26E</th>'
            f'</tr></thead><tbody>{stock_rows}</tbody></table></div></div>')

    rows_html = ""
    for sec in BLOOMBERG_SECTORS:
        dat = bl["sectors"].get(sec) or {}
        subs = dat.get("subs") or []
        if not subs:
            continue
        # 대분류 통계 = 전체 종목 합
        all_stocks = [s for sub in subs for s in sub["stocks"]]
        big_1w = _avg(all_stocks, "chg_1w")
        big_1m = _avg(all_stocks, "chg_1m")
        big_3m = _avg(all_stocks, "chg_3m")
        big_1y = _avg(all_stocks, "chg_1yr")
        n_total = len(all_stocks)

        # 자식 중분류 HTML (여러 개면 nested 아코디언, 하나면 자동 open 상태로 표시)
        subs_html = "".join(_sub_html(sub, i) for i, sub in enumerate(subs))
        # 중분류가 1개이면 대분류 body에서 자동 열림 (nested가 무의미)
        sub_wrapper_cls = "bl-sub-list" + (" one-sub" if len(subs) == 1 else "")

        rows_html += (
            f'<div class="sec-row" style="margin-bottom:8px"><div class="sec-head">'
            f'<span class="arrow">▶</span>'
            f'<span class="nm">{sec}</span>'
            f'<span class="chg {_mid_cls(big_1w)}">1W {_fmt(big_1w)}</span>'
            f'<span class="cap">1M {_fmt(big_1m)} · 3M {_fmt(big_3m)} · 1Y {_fmt(big_1y)} · '
            f'{len(subs)}중분류 · {n_total}종목</span>'
            f'</div>'
            f'<div class="sec-body">'
            f'<div class="{sub_wrapper_cls}">{subs_html}</div>'
            f'</div></div>')

    return (
        '<div class="card"><h2>업종별 글로벌 피어 수익률 '
        '<span class="src-tag">Bloomberg Valuation Table</span></h2>'
        f'<div class="muted" style="font-size:12px;margin-bottom:10px">'
        f'파일 갱신 {mtime_str} · 대분류 클릭 → 중분류 · 중분류 클릭 → 종목 표 · '
        f'수익률(%)·시총(USD bil)·PER(x)</div>'
        f'<div class="sec-list">{rows_html}</div></div>')


# ---------------------------------------------------------------------------
# 4) 라우팅
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    funds = find_funds()
    if not funds:
        present = []
        if os.path.isdir(DATA_DIR):
            present = sorted(f for f in os.listdir(DATA_DIR)
                             if f.lower().endswith(".xlsx"))
        listing = "".join(f"<li><code>{f}</code></li>" for f in present) \
            or "<li>(이 폴더에 .xlsx 파일이 없습니다)</li>"
        return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
        <title>펀드 파일을 찾지 못함</title></head>
        <body style="font-family:'Malgun Gothic',sans-serif;max-width:760px;
        margin:48px auto;line-height:1.75;color:#222;padding:0 20px">
        <h2>펀드 파일을 찾지 못했습니다</h2>
        <p>앱이 실제로 보고 있는 data 폴더는 여기입니다:</p>
        <p style="background:#f4f6f9;padding:12px 14px;border-radius:8px">
        <code>{DATA_DIR}</code></p>
        <p>이 폴더에서 발견한 엑셀 파일:</p>
        <ul>{listing}</ul>
        <p><b>확인할 점</b></p>
        <ol>
          <li>파일을 넣은 폴더가 <u>위 경로</u>와 같은가요?
              다르면, 실행한 app.py 옆의 data 폴더에 파일을 넣어야 합니다.</li>
          <li>이름이 <b>○○○펀드수익률.xlsx</b> / <b>○○○펀드종목.xlsx</b> 형태로,
              수익률·종목 <u>두 개가 한 쌍</u>으로 들어 있나요?</li>
        </ol>
        </body></html>"""
    return redirect(url_for("firm"))


@app.route("/firm")
def firm():
    funds = find_funds()
    if not funds:
        return redirect(url_for("index"))
    fd = process_firm(funds)
    return render_firm(list(funds.keys()), fd)


@app.route("/alpha")
def alpha_page():
    funds = find_funds()
    if not funds:
        return redirect(url_for("index"))
    perf_map = get_stock_perf_map()
    bm = get_market_benchmark()
    k200 = load_kospi200()

    # KOSPI 200 alpha points
    kospi200_alpha = build_alpha_scatters(
        [(s["code"], s["name"], s["weight"]) for s in k200["stocks"]],
        perf_map, bm)

    # 각 펀드 alpha
    fund_alphas = {}
    for fn, paths in funds.items():
        h = read_holdings(paths["holdings"])
        codes_pcts = [(r.get("종목코드"), r.get("종목명"), r.get("편입비", 0))
                      for _, r in h.iterrows()]
        fund_alphas[fn] = build_alpha_scatters(codes_pcts, perf_map, bm)

    # 전사 alpha (all / ex_tmf) — process_firm 재사용
    fd = process_firm(funds)

    tmf_funds = [fn for fn in funds if _is_tmf(fn)]
    payload = {
        "bm": bm,
        "kospi200": kospi200_alpha,
        "kospi200_count": len(k200["codes"]),
        "kospi200_matched": len(kospi200_alpha.get("points", [])),
        "firm_all": fd.get("alpha_all") or {},
        "firm_ex_tmf": fd.get("alpha_ex_tmf") or {},
        "funds": fund_alphas,
        "fund_names": list(funds.keys()),
        "tmf_funds": tmf_funds,
    }
    return render_alpha(list(funds.keys()), payload)


@app.route("/stock/<code>")
def stock_page(code):
    code = (code or "").strip()
    master = get_stock_master()
    if code not in master:
        return ("등록된 종목이 아닙니다. 엑셀(반도체_소부_총정리.xlsx)에 추가 후 캐시 파일을 삭제하세요.", 404)
    funds = find_funds()
    return render_stock(list(funds.keys()), master[code])


# 지수/ETF 등 종목마스터에 없어도 차트 조회 허용할 코드 (KODEX 200 = BM)
_OHLCV_ALLOWLIST = {"069500": "KOSPI"}


@app.route("/api/ohlcv/<code>")
def api_ohlcv(code):
    code = (code or "").strip()
    master = get_stock_master()
    if code in master:
        market = master[code].get("market")
    elif code in _OHLCV_ALLOWLIST:
        market = _OHLCV_ALLOWLIST[code]
    else:
        return {"ok": False, "error": "unknown code"}, 404
    try:
        bars = load_ohlcv(code, years=5, market=market)
        return {"ok": True, "code": code, "bars": bars}
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500


def load_research_reports(date_str=None):
    """data/reports/ 에서 지정 날짜 또는 가장 최신 JSON 로드."""
    if not os.path.isdir(REPORTS_DIR):
        return None, []
    files = sorted([f for f in os.listdir(REPORTS_DIR) if f.endswith(".json")], reverse=True)
    if not files:
        return None, []
    if date_str:
        target = f"{date_str}.json"
        if target in files:
            files = [target] + [f for f in files if f != target]
    with open(os.path.join(REPORTS_DIR, files[0]), encoding="utf-8") as f:
        latest = json.load(f)
    all_dates = [f.replace(".json", "") for f in files]
    return latest, all_dates


RESEARCH_CSS = """
  .rs-head{ display:flex; align-items:flex-end; justify-content:space-between;
    margin-bottom:18px; flex-wrap:wrap; gap:10px; }
  .rs-head h1{ font-size:22px; margin:0; font-weight:800; }
  .rs-head .subname{ color:var(--navy2); font-size:13px; font-weight:600; margin-top:3px; }
  .rs-toolbar{ display:flex; gap:10px; align-items:center; flex-wrap:wrap;
    background:var(--card); padding:12px 14px; border-radius:10px;
    border:1px solid var(--line); margin-bottom:14px; }
  .rs-search{ flex:1; min-width:200px; padding:8px 12px; border-radius:8px;
    border:1px solid var(--line); font-size:13px; font-family:inherit; outline:none; }
  .rs-search:focus{ border-color:var(--navy2); }
  .rs-date{ padding:7px 10px; border-radius:8px; border:1px solid var(--line);
    font-size:13px; font-family:inherit; background:#fff; }
  .rs-count{ color:var(--muted); font-size:13px; }

  .rs-list{ display:grid; grid-template-columns:repeat(auto-fill,minmax(360px,1fr));
    gap:14px; }
  .rs-card{ background:var(--card); border:1px solid var(--line); border-radius:12px;
    padding:14px 16px; display:flex; flex-direction:column; gap:8px;
    transition:transform .12s, box-shadow .12s; }
  .rs-card:hover{ transform:translateY(-2px); box-shadow:0 6px 18px rgba(22,49,79,.08); }

  .rs-card .rs-top{ display:flex; align-items:center; gap:8px; }
  .rs-card .rs-stock{ font-size:16px; font-weight:800; color:var(--ink);
    flex:1; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
  .rs-card .rs-stock .code{ font-size:11px; color:var(--muted); font-weight:600;
    margin-left:5px; font-family:monospace; }
  .rs-card .rs-broker{ font-size:12px; color:var(--muted); font-weight:600;
    padding:3px 8px; background:#f6f8fb; border-radius:6px; white-space:nowrap; }

  .rs-card .rs-title{ font-size:13.5px; color:var(--ink); line-height:1.4;
    font-weight:600; min-height:36px; }

  /* 투자의견 + 목표가 강조 박스 */
  .rs-card .rs-tp{ display:flex; align-items:center; gap:10px;
    padding:9px 12px; background:#f8fafc; border-radius:8px;
    border-left:3px solid var(--navy2); }
  .rs-card .rs-op{ font-size:12px; font-weight:700; padding:3px 9px;
    border-radius:5px; background:#16314f; color:#fff; }
  .rs-card .rs-op.buy{ background:#c93a2e; }
  .rs-card .rs-op.hold{ background:#7c8b9c; }
  .rs-card .rs-op.sell{ background:#1565c0; }
  .rs-card .rs-target{ flex:1; font-size:13px; }
  .rs-card .rs-target b{ font-size:15px; font-weight:800; }
  .rs-card .rs-up{ font-size:12.5px; font-weight:700; }
  .rs-card .rs-up.pos{ color:var(--up); } .rs-card .rs-up.neg{ color:var(--down); }

  /* EPS · 목표가 변동 배지 */
  .rs-card .rs-eps{ display:inline-flex; align-items:center; gap:4px;
    font-size:11.5px; font-weight:700; padding:2px 7px; border-radius:5px;
    margin-right:4px; }
  .rs-card .rs-eps.up{ background:#fee9e7; color:#c93a2e; }
  .rs-card .rs-eps.dn{ background:#e7f0fd; color:#1565c0; }
  .rs-card .rs-eps.flat{ background:#eef2f7; color:#5a6b7e; }
  body[data-theme="dark"] .rs-card .rs-eps.flat{ background:rgba(255,255,255,.08); color:#a5b3c5; }

  .rs-card .rs-meta{ display:flex; justify-content:space-between;
    font-size:11.5px; color:var(--muted); padding-top:8px;
    border-top:1px solid #f0f2f6; }
  .rs-card a.rs-open{ display:inline-block; padding:6px 11px;
    background:var(--navy); color:#fff; border-radius:7px; font-size:12px;
    font-weight:600; text-decoration:none; text-align:center; }
  .rs-card a.rs-open:hover{ background:var(--navy2); }

  .rs-empty{ background:var(--card); border:1px dashed var(--line); border-radius:12px;
    padding:40px; text-align:center; color:var(--muted); }
  @media (max-width:600px){ .rs-list{ grid-template-columns:1fr; } }
"""

RESEARCH_JS = """
const RPS = __REPORTS__;
const $list = document.getElementById('rsList');
const $count = document.getElementById('rsCount');
const $search = document.getElementById('rsSearch');

function fmtKor(n) { return n == null ? '-' : Number(n).toLocaleString('ko-KR'); }

function opCls(opinion) {
  if (!opinion) return '';
  const o = opinion.toUpperCase();
  // 매수/Buy/Outperform/Overweight/Accumulate → 빨강
  if (o.includes('BUY') || o.includes('매수') ||
      o.includes('OUTPERFORM') || o.includes('OVERWEIGHT') ||
      o.includes('ACCUMULATE') || o.includes('STRONG')) return 'buy';
  // 매도/Sell/Underperform/Underweight/Reduce → 파랑
  if (o.includes('SELL') || o.includes('매도') ||
      o.includes('UNDERPERFORM') || o.includes('UNDERWEIGHT') ||
      o.includes('REDUCE')) return 'sell';
  // 보유/Hold/Neutral/Marketperform → 회색
  return 'hold';
}

function render(q) {
  q = (q||'').trim().toLowerCase();
  const filtered = q ? RPS.filter(r =>
    (r.stock_name||'').toLowerCase().includes(q) ||
    (r.broker||'').toLowerCase().includes(q) ||
    (r.analyst||'').toLowerCase().includes(q) ||
    (r.title||'').toLowerCase().includes(q) ||
    (r.opinion||'').toLowerCase().includes(q)
  ) : RPS;
  $count.textContent = `${filtered.length}건${q ? ' (필터 적용)' : ''}`;
  if (!filtered.length) {
    $list.innerHTML = '<div class="rs-empty">일치하는 리포트가 없습니다</div>';
    return;
  }
  $list.innerHTML = filtered.map(r => {
    const code = r.stock_code ? `<span class="code">${r.stock_code}</span>` : '';
    // 투자의견 + 목표가 박스
    let tpBox = '';
    if (r.opinion || r.target_price) {
      const op = r.opinion ? `<span class="rs-op ${opCls(r.opinion)}">${r.opinion}</span>` : '';
      let tgt = '';
      if (r.target_price) {
        const upcls = (r.upside_pct||0) >= 0 ? 'pos' : 'neg';
        const upsign = (r.upside_pct||0) >= 0 ? '+' : '';
        const upTxt = r.upside_pct != null ? ` <span class="rs-up ${upcls}">${upsign}${r.upside_pct.toFixed(1)}%</span>` : '';
        tgt = `<span class="rs-target">목표 <b>${fmtKor(r.target_price)}</b>원${upTxt}</span>`;
      } else if (r.prev_close) {
        tgt = `<span class="rs-target" style="color:var(--muted)">전일종가 ${fmtKor(r.prev_close)}원</span>`;
      }
      tpBox = `<div class="rs-tp">${op}${tgt}</div>`;
    }
    // 목표가 / EPS 변동 배지
    let tgtBadge = '';
    if (r.target_change === '상향') tgtBadge = '<span class="rs-eps up">▲ 목표가 상향</span>';
    else if (r.target_change === '하향') tgtBadge = '<span class="rs-eps dn">▼ 목표가 하향</span>';
    else if (r.target_change === '유지') tgtBadge = '<span class="rs-eps flat">— 목표가 유지</span>';
    let epsBadge = '';
    if (r.eps_change === '상향') epsBadge = '<span class="rs-eps up">▲ EPS 상향</span>';
    else if (r.eps_change === '하향') epsBadge = '<span class="rs-eps dn">▼ EPS 하향</span>';

    return `<div class="rs-card">
      <div class="rs-top">
        <div class="rs-stock">${r.stock_name}${code}</div>
        <div class="rs-broker">${r.broker}</div>
      </div>
      <div class="rs-title">${r.title}</div>
      ${tpBox}
      <div class="rs-meta">
        <span>${r.analyst} · ${r.pages || '?'}p ${tgtBadge}${epsBadge}</span>
        <a class="rs-open" href="${r.source_url || r.pdf_url || '#'}" target="_blank" rel="noopener">원문 →</a>
      </div>
    </div>`;
  }).join('');
}

$search.addEventListener('input', e => render(e.target.value));
render('');
"""


def render_research(fund_names, latest, all_dates):
    if not latest:
        body = (
            '<div class="head"><div><h1>리서치</h1>'
            '<div class="subname">매일 FNGuide 에서 자동 수집된 개별 종목 리포트</div></div></div>'
            '<div class="card"><h2>아직 수집된 리포트가 없습니다</h2>'
            '<p style="color:var(--muted);line-height:1.7">'
            '<code>data/reports/YYYY-MM-DD.json</code> 파일이 필요합니다. '
            '터미널에서 <code>python3 fnguide_scraper.py</code> 를 실행하면 오늘자 리포트를 수집합니다.'
            '</p></div>')
        return page_shell("ANDA 대시보드 · 리서치", "__research__", fund_names, body, "")

    reports = latest.get("reports", [])
    as_of = latest.get("as_of", "-")
    fetched = latest.get("fetched_at", "-")

    # 날짜 드롭다운 (다른 날짜 보기)
    date_opts = "".join(
        f'<option value="{d}" {"selected" if d == as_of else ""}>{d}</option>'
        for d in all_dates[:30]
    )

    body = (
        '<div class="rs-head"><div><h1>오늘의 리서치</h1>'
        f'<div class="subname">FNGuide · 개별 종목 리포트 (자동 수집)</div></div>'
        f'<div class="asof">수집 {fetched}</div></div>'

        '<div class="rs-toolbar">'
        '<input type="text" id="rsSearch" class="rs-search" '
        'placeholder="종목명·증권사·애널리스트·제목으로 검색" autocomplete="off">'
        f'<select class="rs-date" onchange="location.href=\'/research/\'+this.value">{date_opts}</select>'
        '<span class="rs-count" id="rsCount">—</span>'
        '</div>'

        '<div class="rs-list" id="rsList"></div>'
    )

    scripts = ("<style>" + RESEARCH_CSS + "</style>") + \
              RESEARCH_JS.replace("__REPORTS__", json.dumps(reports, ensure_ascii=False))
    return page_shell("ANDA 대시보드 · 리서치", "__research__", fund_names, body, scripts)


@app.route("/research")
@app.route("/research/<date_str>")
def research_page(date_str=None):
    funds = find_funds()
    latest, all_dates = load_research_reports(date_str)
    return render_research(list(funds.keys()), latest, all_dates)


@app.route("/market")
def market_page():
    funds = find_funds()
    data = load_market_data()
    # 차트 드릴다운용: 대/중/소 모두 미리 그룹핑해 둠 (클라이언트 측 전환).
    # 종목 카드는 기존 wics_level 쿠키 따라 한 단위만 표시.
    if data:
        raw = data.get("raw_stocks", [])
        level = request.cookies.get("wics_level", "big")
        if level not in ("big", "mid", "small"):
            level = "big"
        data = dict(data)
        data["sectors"] = _group_sectors(raw, level=level)        # 종목 카드용
        data["sectors_big"] = _group_sectors(raw, level="big")    # 차트 — 대분류
        data["sectors_mid"] = _group_sectors(raw, level="mid")    # 차트 — 중분류
        data["sectors_small"] = _group_sectors(raw, level="small")  # 차트 — 소분류
        data["wics_level"] = level
    return render_market(list(funds.keys()), data)


@app.route("/api/market/refresh")
def api_market_refresh():
    load_market_data(force_refresh=True)
    return {"ok": True}


@app.route("/kosdaq70")
def kosdaq70_page():
    funds = find_funds()
    data = load_kosdaq70()
    if not data:
        return ("코스닥 데이터를 찾지 못했습니다. data/" + KOSDAQ_FILE + " 가 있는지 확인하세요.", 404)
    return render_kosdaq70(list(funds.keys()), data)


@app.route("/fund/<name>")
def fund(name):
    funds = find_funds()
    if name not in funds:
        abort(404)
    risk_rows = load_risk()
    risk_row = match_risk(name, risk_rows)
    data = process_fund(name, funds[name]["returns"], funds[name]["holdings"], risk_row)
    data["risk_scatter"] = build_scatter(risk_rows, risk_row["code"] if risk_row else None)
    data["firm_size"] = get_firm_size(funds)
    data["kospi_size"] = load_kospi_size()
    return render_page(name, list(funds.keys()), data)


# ---------------------------------------------------------------------------
# 백그라운드 리서치 수집 스케줄러
# ---------------------------------------------------------------------------
_SCHEDULER_STARTED = False


def _start_research_scheduler():
    """평일 7~16시에 1시간 간격으로 오늘자 리포트 수집을 시도하는 데몬 스레드.

    동작 규칙:
      - 앱 시작 직후 1회 시도 (오늘자 파일이 없거나 4시간 이상 오래된 경우)
      - 이후 1시간마다 같은 조건 검사 → 매번 wisereport_scraper.fetch() 호출
      - 휴일/장 종료 후에는 조건 불충족이라 fetch 건너뜀 (헛걸음 없음)
    수집 로그는 data/reports/_scheduler.log 에 추가 기록.
    """
    global _SCHEDULER_STARTED
    if _SCHEDULER_STARTED:
        return
    _SCHEDULER_STARTED = True

    import threading
    import time as _time
    import datetime as _dt
    import traceback

    def _log(msg):
        ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {msg}"
        print(line)
        try:
            os.makedirs(REPORTS_DIR, exist_ok=True)
            with open(os.path.join(REPORTS_DIR, "_scheduler.log"),
                     "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def _should_fetch(startup=False):
        now = _dt.datetime.now()
        if now.weekday() >= 5:        # 토/일 제외
            return False
        # 서버 시작 직후엔 시간 제약 없이 오늘 파일이 없으면 즉시 시도
        if not startup and not (7 <= now.hour <= 16):
            return False
        today = now.strftime("%Y-%m-%d")
        today_file = os.path.join(REPORTS_DIR, f"{today}.json")
        if not os.path.isfile(today_file):
            return True
        age_h = (_time.time() - os.path.getmtime(today_file)) / 3600.0
        return age_h > 4.0           # 4시간 이상 오래된 데이터는 재수집

    def _do_fetch():
        try:
            import importlib
            ws = importlib.import_module("wisereport_scraper")
            _log("wisereport 수집 시작")
            ws.fetch(debug=False, today_only=True)
            _log("wisereport 수집 완료")
        except Exception as e:
            _log(f"수집 실패: {e}")
            _log(traceback.format_exc().strip().splitlines()[-1])

    def _loop():
        # 앱 시작 후 잠깐 대기 (Flask 가 먼저 뜨도록)
        _time.sleep(8)
        first_iter = True
        while True:
            try:
                if _should_fetch(startup=first_iter):
                    _do_fetch()
                elif first_iter:
                    _log("스케줄러 초기 체크: 오늘 파일 이미 있음 or 휴일")
            except Exception as e:
                _log(f"loop 오류: {e}")
            first_iter = False
            _time.sleep(3600)        # 1시간 간격

    t = threading.Thread(target=_loop, daemon=True, name="research-scheduler")
    t.start()
    _log("리서치 스케줄러 시작 (평일 7~16시 1시간 간격)")


if __name__ == "__main__":
    host = "0.0.0.0" if SHARE_ON_NETWORK else "127.0.0.1"
    ip = _local_ip()
    bar = "=" * 54
    print(bar)
    print("  ANDA 펀드 대시보드가 실행되었습니다.")
    print(f"  내 PC에서:     http://127.0.0.1:5000")
    if SHARE_ON_NETWORK:
        print(f"  동료 접속용:   http://{ip}:5000   (같은 사내망에서)")
        print(f"  로그인 아이디: {LOGIN_ID}   /   비밀번호: 코드에 설정한 값")
    if LOGIN_PW == "CHANGE_ME_1234":
        print("  [주의] 비밀번호가 기본값입니다. app.py 의 LOGIN_PW 를 꼭 바꾸세요.")
    print("  (종료: 이 창에서 Ctrl + C)")
    print(bar)
    _start_research_scheduler()
    app.run(host=host, port=5000, debug=False)
